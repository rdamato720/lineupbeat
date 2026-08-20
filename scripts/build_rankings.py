#!/usr/bin/env python3
"""2026 Half-PPR rankings: /nfl/rankings/ and the four position boards.

    python3 scripts/build_rankings.py
    python3 scripts/build_rankings.py --export data/site_export.csv
    python3 scripts/build_rankings.py --dry-run

Five URLs, one component, one dataset. /nfl/rankings/ is the overall Top
200; /nfl/rankings/qb|rb|wr|te/ are the same page rendered with a position
filter. There is no second template and no hand-maintained HTML -- a change
to the board is one change here, or the five drift.

WHERE THE NUMBERS COME FROM

`data/projections.xlsx` is the site's projection board. The projections
page, the player-page chips and draft value already read it, and the
standing rule is one number read twice: a rankings page fed by its own
private copy of the projections would be a second source for a number that
already exists, free to disagree with the first.

So the default is to derive rankings from that workbook using the published
formula. `--export` accepts the Site Export sheet as CSV or JSON where a
separately prepared ranking file is the intended source; the same
validation runs either way.

THE RANK CORRECTION

The `Rank` column in each projection sheet is NOT used. It may carry a
different scoring format's order, and a rank that is silently for the wrong
format is worse than no rank. Position rank is computed by sorting within
the position on ranking_score, descending. Overall rank is computed by
sorting everybody on ranking_score, then projected points, then name.

WHAT AN ADJUSTMENT IS AND IS NOT

An editorial adjustment moves draft order. It never touches the frozen
projected-points column, and it cannot be applied without a published
reason -- the build fails on a non-zero adjustment with no reason, because
an unexplained thumb on the scale is indistinguishable from a bug.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import seo  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"
SPORT = "nfl"
SEASON = 2026

POSITIONS = ["QB", "RB", "WR", "TE"]
TOP_N = 200

# What the board is reconciled against is the source it was built from, not a
# number written here. A constant would mean signing a new player is a code
# change, and a code change nobody wants to make is a gate that gets deleted.
#
# The count is counted at read time, carried into the manifest the build
# writes, and pinned in data/nfl_rankings_source.json for the frozen source.
# A legitimate addition moves all three together and shows up in the diff; a
# swapped or truncated artifact moves the SHA and fails.
MANIFEST = ROOT / "data" / "nfl_rankings_source.json"

# Ranking inputs that are not projections: the league shape, the replacement
# ranks, the tier bands and the approved adjustments. They live beside the
# board rather than inside it so a projections update flows straight through,
# and in a tracked file rather than in this module so an editorial decision
# is a data change with a diff, not a deploy.
CONFIG = ROOT / "data" / "nfl_rankings_config.json"

# Provenance, not the production feed. The workbook the rankings were first
# cut from, kept for audit and for `--export` rebuilds of a historical board.
ARCHIVE = ROOT / "data" / "archive"

SCORING_FORMAT = "half_ppr"
SCORING_LABEL = "Half-PPR"

# The default league, and the replacement levels that fall out of it:
# 12 teams starting 1 QB, 2 RB, 2 WR, 1 TE and 2 FLEX.
LEAGUE = {"teams": 12, "qb": 1, "rb": 2, "wr": 2, "te": 1, "flex": 2}
REPLACEMENT_RANK = {"QB": 13, "RB": 37, "WR": 49, "TE": 13}

DATA_JSON = ROOT / "data" / f"nfl_rankings_{SEASON}.json"

# Approved editorial adjustments. Name -> (points, published reason).
#
# Every entry here changes draft order and nothing else. The reason is not
# documentation of the adjustment, it is part of it: validate() rejects a
# non-zero adjustment without one.
ADJUSTMENTS = {
    "Ashton Jeanty": (
        16.0,
        "Upgraded for elite tackle-breaking production behind one of 2025's "
        "least effective blocking environments, plus the arrival of Rick "
        "Dennison and Klint Kubiak's run-oriented system. Frozen statistical "
        "projection unchanged."),
    "Christian McCaffrey": (
        -33.0,
        "Downgraded for age-30 durability risk following 440 regular-season "
        "opportunities and 450 total touches including the playoffs, combined "
        "with an extended health-related absence during training camp. Frozen "
        "statistical projection unchanged."),
}

DEFAULT_REASON = "Projection-only v1.0"
DEFAULT_WHY = "Projection-based ranking"

# The expected top of the RB board after the approved adjustments.
#
# This is a test of the sorting logic, never an input to it. Nothing here is
# hardcoded into a rank: the list is compared against what the sort produced
# and the build stops if they differ.
# The approved board, as published. Tests of the ranking, never inputs to it:
# each is compared against what the sort and the editorial pass produced.
EXPECTED_OVERALL_TOP = [
    "Jahmyr Gibbs", "Bijan Robinson", "Ja'Marr Chase", "Puka Nacua",
    "Ashton Jeanty", "Jonathan Taylor",
]

# Relative order that must hold, whether or not the players are consecutive.
# Love over Montgomery is not a configured constraint -- it falls out of the
# two that are, plus the natural order -- so it is checked rather than
# assumed. If a projection change ever breaks it, the build stops and
# somebody adds an explicit constraint with a published reason.
EXPECTED_ORDER = [
    ("Ja'Marr Chase", "Puka Nacua"),
    ("Ashton Jeanty", "Jonathan Taylor"),
    ("Jeremiyah Love", "David Montgomery"),
    ("David Montgomery", "Josh Jacobs"),
    ("Jadarian Price", "Bucky Irving"),
]

EXPECTED_RB_TOP = [
    "Jahmyr Gibbs", "Bijan Robinson", "Ashton Jeanty", "Jonathan Taylor",
    "Omarion Hampton",
]

# Where the two adjusted players must land. McCaffrey is RB9 at one decimal
# -- 255.1 against Cook's 255.2 -- and that is the answer, not a rounding
# problem to be tuned away.
EXPECTED_POSITION_RANK = {"Ashton Jeanty": 3, "Christian McCaffrey": 9}

CANON = f"https://lineupbeat.com/{SPORT}/rankings/"
OG_IMAGE = "https://lineupbeat.com/og.png"

TITLE = f"{SEASON} Fantasy Football Rankings: Half-PPR Top 200 | LineupBeat"
DESCRIPTION = (
    f"LineupBeat's {SEASON} Half-PPR fantasy football rankings, including an "
    "overall Top 200, positional rankings, projected points, tiers and "
    "documented ranking adjustments.")
POS_TITLE = f"{SEASON} Fantasy {{pos}} Rankings | LineupBeat"


def bands(ranges, fallback):
    """[[1,8],[9,24]] -> [8, 24]: a tier is named by where it ends."""
    if not ranges:
        return list(fallback)
    try:
        return [int(r[1]) for r in ranges]
    except (TypeError, ValueError, IndexError):
        return list(fallback)


def load_config():
    """The tracked ranking configuration, with the built-in defaults beneath.

    Missing file means the defaults, so a checkout without it still builds a
    correct board; a present file wins field by field, so a config that only
    pins adjustments does not have to restate the league.
    """
    cfg = {
        "replacement_ranks": dict(REPLACEMENT_RANK),
        "overall_tiers": list(OVERALL_TIER_MAX),
        "position_tiers": {k: list(v) for k, v in POSITION_TIER_MAX.items()},
        "adjustments": {},
        "editorial": [],
        "league": dict(LEAGUE),
        "sha256": None,
    }
    if not CONFIG.exists():
        print(f"  no {CONFIG.name}, using built-in defaults")
        return cfg
    raw = json.loads(CONFIG.read_text())
    cfg["sha256"] = sha256(CONFIG)
    if raw.get("replacement_ranks"):
        cfg["replacement_ranks"] = {k.upper(): int(v) for k, v
                                    in raw["replacement_ranks"].items()}
    cfg["overall_tiers"] = bands(raw.get("overall_tiers"), OVERALL_TIER_MAX)
    if raw.get("position_tiers"):
        cfg["position_tiers"] = {
            k.upper(): bands(v, POSITION_TIER_MAX.get(k.upper(), OVERALL_TIER_MAX))
            for k, v in raw["position_tiers"].items()}
    if raw.get("league"):
        cfg["league"] = raw["league"]

    # Identity is name + team + position everywhere. A bare name would follow
    # a player to a new team, and a decision written about one situation is
    # not automatically true of another.
    for a in (raw.get("manual_adjustments") or []):
        cfg["adjustments"][(str(a.get("player") or "").strip(),
                            str(a.get("team") or "").strip().upper(),
                            str(a.get("position") or "").strip().upper())] = {
            "value": float(a.get("value") or 0.0),
            "reason": (a.get("reason") or "").strip(),
        }
    # The older keyed form, still read so an existing config keeps working.
    for key, a in (raw.get("adjustments") or {}).items():
        parts = [x.strip() for x in str(key).split("|")]
        if len(parts) != 3:
            print(f"  adjustment key {key!r} is not name|TEAM|POS, ignored")
            continue
        cfg["adjustments"].setdefault(
            (parts[0], parts[1].upper(), parts[2].upper()),
            {"value": float(a.get("value") or 0.0),
             "reason": (a.get("reason") or "").strip()})

    for e in (raw.get("editorial_order") or []):
        cfg["editorial"].append({
            "player": str(e.get("player") or "").strip(),
            "team": str(e.get("team") or "").strip().upper(),
            "position": str(e.get("position") or "").strip().upper(),
            "over": str(e.get("ranked_ahead_of") or "").strip(),
            "over_team": str(e.get("ranked_ahead_of_team") or "").strip().upper(),
            "reason": (e.get("reason") or "").strip(),
        })
    return cfg


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def check_manifest(source: Path, digest: str, count: int) -> list[str]:
    """Reconcile an artifact against the frozen manifest, when it lists one.

    The manifest pins artifacts by filename. Building from something it does
    not list is not an error -- that is what `--export` is for -- so the pin
    applies only where the names match, and says which case it is either way
    rather than passing in silence.
    """
    if not MANIFEST.exists():
        return []
    try:
        m = json.loads(MANIFEST.read_text())
    except (ValueError, OSError) as exc:
        return [f"{MANIFEST.name} is unreadable ({exc})"]

    artifacts = m.get("artifacts")
    if artifacts is None:                       # the older single-entry shape
        artifacts = [{"file": m.get("source_file"),
                      "sha256": m.get("source_sha256"),
                      "player_count": m.get("source_player_count")}]
    entry = next((a for a in artifacts if a.get("file") == source.name), None)
    if entry is None:
        print(f"  {source.name} is not in {MANIFEST.name} -- not reconciled "
              f"against a pinned artifact")
        return []

    bad = []
    if entry.get("sha256") and entry["sha256"] != digest:
        bad.append(f"{source.name} does not match the frozen manifest.\n"
                   f"      manifest {entry['sha256'][:16]}...\n"
                   f"      on disk  {digest[:16]}...\n"
                   f"      If the artifact really was replaced, update "
                   f"{MANIFEST.name} in the same commit.")
    if entry.get("player_count") not in (None, count):
        bad.append(f"{source.name} has {count} players, the manifest records "
                   f"{entry['player_count']}")
    if not bad:
        print(f"  reconciled against {MANIFEST.name}: "
              f"{entry.get('player_count')} players, sha matches")
    return bad


def eastern_now():
    """Eastern, because UTC rolls over at 8pm here and stamps tomorrow."""
    now = datetime.now(timezone.utc)
    year = now.year
    mar = datetime(year, 3, 8, tzinfo=timezone.utc)
    dst_start = mar + timedelta(days=(6 - mar.weekday()) % 7)
    nov = datetime(year, 11, 1, tzinfo=timezone.utc)
    dst_end = nov + timedelta(days=(6 - nov.weekday()) % 7)
    offset = -4 if dst_start <= now < dst_end else -5
    # Carry the offset, do not just shift the clock. Shifting alone leaves a
    # datetime that says UTC while holding Eastern wall time, which is
    # invisible in "August 2026" and wrong the moment it is written into
    # metadata as an ISO timestamp.
    return now.astimezone(timezone(timedelta(hours=offset)))


def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)


def slugify(s):
    s = re.sub(r"[^\w\s-]", "", (s or "").lower())
    return re.sub(r"[\s_]+", "-", s).strip("-")


def norm(name):
    """Match names across sources the way the rest of the site does."""
    n = re.sub(r"[.'`’]", "", (name or "").lower())
    return " ".join(re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", n).split())


# --------------------------------------------------------------- loading

def read_workbook(path: Path):
    """Half-PPR points per player, from the site's projection board.

    Deliberately reads the `Half PPR` column and ignores `Rank` -- see the
    module docstring.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        pos = sheet.split()[0].upper()
        if pos not in POSITIONS:
            continue
        ws = wb[sheet]
        head = [str(c.value or "").strip().lower() for c in ws[1]]

        def col(*names):
            for n in names:
                if n in head:
                    return head.index(n)
            return None

        ci = {"player": col("player", "name"), "team": col("team", "tm"),
              "half": col("half ppr", "half", "half-ppr")}
        if ci["player"] is None or ci["half"] is None:
            print(f"  {sheet}: no player or Half PPR column, skipped")
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if ci["player"] >= len(row) or not row[ci["player"]]:
                continue
            try:
                pts = float(row[ci["half"]])
            except (TypeError, ValueError, IndexError):
                continue
            team = ""
            if ci["team"] is not None and ci["team"] < len(row) and row[ci["team"]]:
                team = str(row[ci["team"]]).strip().upper()
            out.append({"player_name": str(row[ci["player"]]).strip(),
                        "team": team, "position": pos,
                        "projected_points": round(pts, 1)})
    return out


# Sheets worth reading in the rankings workbook, best first.
#
# `Source Data` is the whole board -- all 615 players -- and `Site Export` is
# the published 200. The full board is preferred because replacement level is
# the 37th running back, and you cannot find the 37th running back in a list
# that has already been cut to the top 200 overall.
EXPORT_SHEETS = ["Source Data", "Site Export"]

# The workbook's own column names, mapped onto the record's.
SHEET_COLUMNS = {
    "player_name": ("player_name", "player"),
    "team": ("team", "tm"),
    "position": ("position", "pos"),
    "projected_points": ("projected_points", "half-ppr pts", "half_ppr_pts",
                         "half ppr pts", "proj pts", "half_ppr"),
    "manual_adjustment": ("manual_adjustment", "manual adj", "adjustment"),
    "adjustment_reason": ("adjustment_reason", "adj reason", "reason"),
}

# A placeholder, not a market label. It means "no ADP has been loaded", which
# the published record says with null -- and a gate rejects a label sitting
# on a record whose adp is null, so importing the string verbatim would fail
# the build for a value that is not actually a claim about the market.
NOT_A_LABEL = {"adp_not_loaded", "adp not loaded", "", "none", "null"}


def read_rankings_workbook(path: Path):
    """The rankings workbook, from whichever sheet carries the full board.

    The header is not on row one -- the sheets open with a title and a
    subtitle -- so the header row is found by looking for the one that names
    a player column rather than assumed to be at a fixed offset.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = next((n for n in EXPORT_SHEETS if n in wb.sheetnames), None)
    if sheet is None:
        return [], ""
    ws = wb[sheet]

    head, first = None, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12,
                                         values_only=True), 1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(c in ("player", "player_name") for c in cells):
            head, first = cells, i + 1
            break
    if head is None:
        return [], sheet

    def col(*names):
        for n in names:
            if n in head:
                return head.index(n)
        return None

    ci = {k: col(*v) for k, v in SHEET_COLUMNS.items()}
    if ci["player_name"] is None or ci["projected_points"] is None:
        return [], sheet

    out, skipped = [], 0
    for row in ws.iter_rows(min_row=first, values_only=True):
        if ci["player_name"] >= len(row) or not row[ci["player_name"]]:
            continue

        def val(k):
            i = ci.get(k)
            return row[i] if i is not None and i < len(row) else None

        try:
            pts = float(val("projected_points"))
        except (TypeError, ValueError):
            continue
        adj = val("manual_adjustment")
        try:
            adj = float(adj) if adj not in (None, "") else 0.0
        except (TypeError, ValueError):
            adj = 0.0
        position = str(val("position") or "").strip().upper()
        # Only the four skill positions count as a player row. A kicker or a
        # defence in the sheet is not a shortfall in the board, and counting
        # one would make the reconciliation disagree with itself.
        if position not in POSITIONS:
            skipped += 1
            continue
        out.append({
            "player_name": str(val("player_name")).strip(),
            "team": str(val("team") or "").strip().upper(),
            "position": position,
            "projected_points": round(pts, 1),
            "manual_adjustment": adj,
            "adjustment_reason": (str(val("adjustment_reason")).strip()
                                  if val("adjustment_reason") else None),
        })
    if skipped:
        print(f"  {skipped} row(s) outside {'/'.join(POSITIONS)} not counted")
    return out, sheet


def read_export(path: Path):
    """The Site Export sheet, as CSV or JSON.

    Column names are matched loosely because a sheet exported twice is
    rarely exported identically. Anything the file does not carry -- the
    adjustment, its reason -- is filled from ADJUSTMENTS, so an export
    prepared before an adjustment was approved still ranks correctly.
    """
    if path.suffix.lower() == ".json":
        rows = json.loads(path.read_text())
    else:
        with path.open(newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))

    def pick(row, *names):
        low = {str(k).strip().lower().replace(" ", "_"): v
               for k, v in row.items()}
        for n in names:
            if n in low and low[n] not in (None, ""):
                return low[n]
        return None

    out = []
    for row in rows:
        name = pick(row, "player_name", "player", "name")
        pos = pick(row, "position", "pos")
        if not name or not pos:
            continue
        try:
            pts = float(pick(row, "projected_points", "half_ppr", "half_ppr_points",
                             "half ppr", "points", "proj_pts") or "")
        except (TypeError, ValueError):
            continue
        adj = pick(row, "manual_adjustment", "adjustment")
        try:
            adj = float(adj) if adj not in (None, "") else None
        except (TypeError, ValueError):
            adj = None
        position = str(pos).strip().upper()
        if position not in POSITIONS:
            continue
        rec = {"player_name": str(name).strip(),
               "team": str(pick(row, "team", "tm") or "").strip().upper(),
               "position": position,
               "projected_points": round(pts, 1)}
        if adj is not None:
            rec["manual_adjustment"] = adj
            rec["adjustment_reason"] = pick(row, "adjustment_reason", "reason")
        out.append(rec)
    return out


# ---------------------------------------------------------------- tiers
#
# Tiers are rank ranges, not clusters. That is how the source workbook
# defines them on its Assumptions sheet, and it is why the design brief asks
# for the first eight overall players to be marked: tier one IS the top
# eight. A gap-clustered alternative was tried first and thrown away -- it
# disagreed with the published methodology, which is the thing a reader is
# being asked to trust.
#
# Ranks past the last boundary fall into one more tier rather than off the
# end, so a longer board than the workbook anticipated still tiers.

OVERALL_TIER_MAX = [8, 24, 48, 72, 100, 130, 165, 200]
POSITION_TIER_MAX = {
    # One starter: QB and TE.
    "QB": [3, 8, 14, 24], "TE": [3, 8, 14, 24],
    # Two starters, so twice the shelf at every depth.
    "RB": [6, 18, 36, 60], "WR": [6, 18, 36, 60],
}


def tier_for_rank(rank, bounds):
    for i, top in enumerate(bounds, 1):
        if rank <= top:
            return i
    return len(bounds) + 1


# ---------------------------------------------------------------- ranking

def resolve(recs, name, team="", position=""):
    """Exactly one player, or nothing. Ambiguity is never guessed at.

    The resolver elsewhere in this project refuses to guess for the same
    reason: an editorial decision applied to the wrong Josh is worse than one
    that fails loudly.
    """
    hits = [r for r in recs if norm(r["player_name"]) == norm(name)]
    if team:
        hits = [r for r in hits if r["team"] == team] or hits
    if position:
        hits = [r for r in hits if r["position"] == position] or hits
    exact = [r for r in hits
             if (not team or r["team"] == team)
             and (not position or r["position"] == position)]
    return exact if exact else hits


def find_cycle(edges):
    """A cycle in the "must rank above" graph, as the path that closes it.

    Run before anything is sorted. Two constraints that each demand the other
    player is above cannot both be honoured, and the reorder pass would swap
    them back and forth until it hit its iteration cap and published whichever
    side it happened to stop on.
    """
    graph = {}
    for a, b in edges:
        graph.setdefault(a, []).append(b)
    WHITE, GREY, BLACK = 0, 1, 2
    state, stack = {}, []

    def walk(n):
        state[n] = GREY
        stack.append(n)
        for m in graph.get(n, []):
            if state.get(m, WHITE) == GREY:
                return stack[stack.index(m):] + [m]
            if state.get(m, WHITE) == WHITE:
                found = walk(m)
                if found:
                    return found
        stack.pop()
        state[n] = BLACK
        return None

    for n in list(graph):
        if state.get(n, WHITE) == WHITE:
            found = walk(n)
            if found:
                return found
    return None


def apply_editorial(order, pairs, passes=25):
    """Lift each preferred player to sit immediately above his comparison.

    Lift, not demote. The two produce different boards and the brief settles
    it: with Chase over Nacua and Jeanty over Taylor the published top six is
    Gibbs, Robinson, Chase, Nacua, Jeanty, Taylor -- Jeanty moves up past the
    receivers between them rather than Taylor sinking below them.

    Everyone unaffected keeps his relative order; the only players who move
    are the one being lifted and the ones he passes.

    Repeated until stable because one lift can undo another -- two players
    lifted above the same third player interact. `passes` is a backstop; a
    genuine cycle is caught before this runs.
    """
    order = list(order)
    for _ in range(passes):
        moved = False
        for a, b in pairs:
            ia, ib = order.index(a), order.index(b)
            if ia > ib:
                order.pop(ia)
                order.insert(order.index(b), a)
                moved = True
        if not moved:
            return order, True
    return order, False


def sort_key(r):
    """Ranking score, then projected points, then name -- all at one decimal.

    One decimal because that is what the published record stores and what the
    page prints. Comparing on anything coarser would make two visibly
    different scores tie and hand the decision to the tiebreak, which is a
    rule the methodology does not claim and a reader cannot see.

    The tiebreak is therefore rare by design. That is correct: it exists for
    a genuine tie at the published precision, not as a second sort.
    """
    return (-round(r["ranking_score"], 1), -round(r["projected_points"], 1),
            r["player_name"])


def build_records(rows, cfg):
    """Everything the JSON and the pages need, in published order."""
    adjustments = cfg["adjustments"]
    ranks = cfg["replacement_ranks"]
    applied = set()
    # One row per player. A duplicate in the source is a real problem and is
    # caught by validate() rather than silently collapsed here.
    recs = []
    for r in rows:
        pos = (r.get("position") or "").upper()
        name = (r.get("player_name") or "").strip()
        team = (r.get("team") or "").upper()
        entry = adjustments.get((name, team, pos))
        if entry:
            applied.add((name, team, pos))
            adj, reason = entry["value"], entry["reason"]
        else:
            adj = float(r.get("manual_adjustment") or 0.0)
            reason = r.get("adjustment_reason") or DEFAULT_REASON
        recs.append({
            "scoring_format": SCORING_FORMAT,
            "player_name": name,
            "team": (r.get("team") or "").upper(),
            "position": pos,
            "projected_points": r.get("projected_points"),
            "manual_adjustment": round(float(adj), 1),
            "adjustment_reason": reason,
            "adp": None,
            "value_label": None,
        })

    # Replacement level, per position, from the league shape. Taken on
    # projected points -- the frozen number -- so an adjustment cannot move
    # the baseline everybody else is measured against.
    replacement = {}
    for pos in POSITIONS:
        group = sorted([r for r in recs if r["position"] == pos],
                       key=lambda r: -r["projected_points"])
        n = ranks.get(pos, REPLACEMENT_RANK[pos])
        replacement[pos] = (group[n - 1]["projected_points"]
                            if len(group) >= n else
                            (group[-1]["projected_points"] if group else 0.0))

    for r in recs:
        rep = replacement.get(r["position"], 0.0)
        r["replacement_points"] = round(rep, 1)
        r["vorp"] = round(r["projected_points"] - rep, 1)
        r["ranking_score"] = round(r["vorp"] + r["manual_adjustment"], 1)

    # Natural order first: the ranking-score sort, untouched.
    recs.sort(key=sort_key)
    for i, r in enumerate(recs):
        r["_i"] = i

    # Then the approved editorial constraints, resolved against the board.
    issues, pairs, ahead = [], [], {}
    for e in cfg["editorial"]:
        who = resolve(recs, e["player"], e["team"], e["position"])
        over = resolve(recs, e["over"], e["over_team"])
        label = f"{e['player']} over {e['over']}"
        if len(who) != 1:
            issues.append(f"editorial: {e['player']} ({e['team']} "
                          f"{e['position']}) resolves to {len(who)} players")
            continue
        if len(over) != 1:
            issues.append(f"editorial: comparison player {e['over']} "
                          f"({e['over_team']}) resolves to {len(over)} players")
            continue
        a, b = who[0], over[0]
        if not e["reason"]:
            issues.append(f"editorial: {label} has no published reason")
        if a["position"] != b["position"]:
            # Every constraint here is between players who share a position,
            # and it has to be: a position board cannot express "this running
            # back above that receiver".
            issues.append(f"editorial: {label} compares {a['position']} with "
                          f"{b['position']}; a constraint must be within one "
                          f"position")
            continue
        pairs.append((a["_i"], b["_i"]))
        ahead.setdefault(a["_i"], {"names": [], "reasons": []})
        ahead[a["_i"]]["names"].append(b["player_name"])
        ahead[a["_i"]]["reasons"].append(e["reason"])

    # Cycle detection before any reordering, never after: two constraints
    # that each demand the other player is above would otherwise swap back
    # and forth until the pass gave up and published whichever side it
    # stopped on.
    cycle = find_cycle(pairs)
    if cycle:
        # Full identities, not bare names. The config addresses players as
        # name|TEAM|POS and a cycle is fixed by editing those keys, so the
        # error prints the thing that has to be edited.
        names = " -> ".join(
            f"{recs[i]['player_name']}|{recs[i]['team']}|{recs[i]['position']}"
            for i in cycle)
        issues.append(f"editorial: the constraints contain a cycle: {names}. "
                      f"Nothing was reordered and the previous board stands.")
        order = [r["_i"] for r in recs]
    else:
        order, stable = apply_editorial([r["_i"] for r in recs], pairs)
        if not stable:
            issues.append("editorial: the constraints did not settle; "
                          "nothing was reordered")
            order = [r["_i"] for r in recs]

    by_i = {r["_i"]: r for r in recs}
    ranked = [by_i[i] for i in order]

    for r in recs:
        info = ahead.get(r["_i"])
        r["editorial_override"] = bool(info)
        r["ranked_ahead_of"] = "; ".join(info["names"]) if info else None
        r["override_reason"] = ("\n\n".join(dict.fromkeys(info["reasons"]))
                                if info else None)

    # Both sets of ranks come off the one final order, so the overall board
    # and a position board cannot disagree about two players -- there is only
    # one sequence to read them from.
    seen_pos = {}
    for r in ranked:
        pos = r["position"]
        seen_pos[pos] = seen_pos.get(pos, 0) + 1
        r["position_rank"] = seen_pos[pos]
        r["position_tier"] = tier_for_rank(
            seen_pos[pos], cfg["position_tiers"].get(pos, OVERALL_TIER_MAX))
    for i, r in enumerate(ranked, 1):
        if i <= TOP_N:
            r["overall_rank"] = i
            r["overall_tier"] = tier_for_rank(i, cfg["overall_tiers"])
        else:
            r["overall_rank"] = None
            r["overall_tier"] = None
    recs = ranked
    for r in recs:
        r.pop("_i", None)

    # The published record, in the documented field order.
    order = ["scoring_format", "overall_rank", "overall_tier", "position_rank",
             "position_tier", "player_name", "team", "position",
             "projected_points", "replacement_points", "vorp",
             "manual_adjustment", "ranking_score", "adp", "value_label",
             "adjustment_reason", "editorial_override", "ranked_ahead_of",
             "override_reason"]
    # An adjustment that matched nobody is an editorial decision that
    # silently did not happen -- a traded player, a renamed key. Reported
    # back so validate() can stop the build rather than publish a board that
    # quietly ignores it.
    for key in sorted(set(adjustments) - applied):
        issues.append(f"adjustment {'|'.join(key)} matched no player on the "
                      f"board -- check the team and position in {CONFIG.name}")
    return [{k: r[k] for k in order} for r in recs], replacement, issues


# ------------------------------------------------------------- validation

def validate(records, source_count, issues=()):
    """The published rules, checked before publishing rather than trusted.

    Each of these is something the page or the JSON asserts. A rule nobody
    checks is a rule that quietly stops being true, so every one of them
    stops the build.
    """
    # Configuration problems first. A cycle or an unresolvable name makes
    # every ordering check below fail as a consequence, and a reader who has
    # to scroll past nine symptoms to reach the cause will fix the wrong
    # thing.
    bad = list(issues)
    ranked = [r for r in records if r["overall_rank"] is not None]

    # -- the board against the source it was built from --------------------
    #
    # Both counts, because they fail differently. The first catches a player
    # dropped between reading and publishing; the second catches two rows for
    # one player, which leaves the total intact and the board wrong.
    if len(records) != source_count:
        bad.append(f"{len(records)} players published from a source of "
                   f"{source_count} -- {abs(source_count - len(records))} "
                   f"lost between reading and publishing")
    ids = [norm(r["player_name"]) for r in records]
    if len(set(ids)) != source_count:
        seen = set()
        dupes = sorted({r["player_name"] for r in records
                        if norm(r["player_name"]) in seen
                        or seen.add(norm(r["player_name"]))})
        bad.append(f"{len(set(ids))} unique players from a source of "
                   f"{source_count}"
                   + (f": {', '.join(dupes[:4])}" if dupes else ""))

    # -- the published top 200 --------------------------------------------
    if len(ranked) != TOP_N:
        bad.append(f"{len(ranked)} non-null overall ranks, expected {TOP_N}")
    ranks = sorted(r["overall_rank"] for r in ranked)
    if len(set(ranks)) != len(ranks):
        bad.append("duplicate overall rank(s)")
    if ranks != list(range(1, len(ranked) + 1)):
        bad.append("overall ranks are not sequential from 1")
    for r in records:
        if r["overall_rank"] is None and r["overall_tier"] is not None:
            bad.append(f"{r['player_name']}: no overall rank but a tier")

    # -- ordering, rebuilt from what the JSON publishes and nothing else ----
    #
    # A plain "scores must descend" check cannot survive an approved
    # override, and relaxing it to "unless somebody is overridden" would
    # excuse any reordering at all. So the board is reconstructed: sort every
    # player by the published key, lift each override above the comparison
    # player it names, and require the result to be exactly the published
    # order. That covers the knock-on moves too -- Jeanty passes two
    # receivers on his way over Taylor, and neither is named anywhere.
    published = [norm(r["player_name"]) for r in records]
    where_i = {norm(r["player_name"]): i for i, r in enumerate(records)}
    pairs = []
    for r in records:
        if not (r["editorial_override"] and r.get("ranked_ahead_of")):
            continue
        for nm in str(r["ranked_ahead_of"]).split(";"):
            j = where_i.get(norm(nm))
            if j is not None:
                pairs.append((where_i[norm(r["player_name"])], j))
    natural = [where_i[norm(r["player_name"])]
               for r in sorted(records, key=sort_key)]
    rebuilt, stable = apply_editorial(natural, pairs)
    if not stable:
        bad.append("the editorial constraints do not settle into one order")
    elif [published[i] for i in rebuilt] != published:
        i = next((n for n, (a, b) in
                  enumerate(zip([published[i] for i in rebuilt], published))
                  if a != b), 0)
        bad.append(f"the published order is not the ranking sort plus the "
                   f"approved overrides; they first differ at rank {i + 1} "
                   f"({records[i]['player_name']})")

    # -- every player, ranked at his position ------------------------------
    for pos in POSITIONS:
        group = [r for r in records if r["position"] == pos]
        if not group:
            bad.append(f"no {pos} on the board")
            continue
        pranks = [r["position_rank"] for r in group]
        if any(p is None for p in pranks):
            n = sum(1 for p in pranks if p is None)
            bad.append(f"{pos}: {n} player(s) with no position rank")
            continue
        if sorted(pranks) != list(range(1, len(group) + 1)):
            if len(set(pranks)) != len(pranks):
                bad.append(f"{pos}: duplicate position rank(s)")
            else:
                bad.append(f"{pos}: position ranks are not sequential "
                           f"from 1 to {len(group)}")
        # A position board is a slice of the overall order, checked against
        # it rather than re-sorted: re-sorting would drop every override and
        # disagree with the page it describes.
        by_rank = [r["player_name"] for r in
                   sorted(group, key=lambda r: r["position_rank"])]
        from_overall = [r["player_name"] for r in records
                        if r["position"] == pos]
        if by_rank != from_overall:
            i = next((n for n, (a, b) in enumerate(zip(by_rank, from_overall))
                      if a != b), 0)
            bad.append(f"{pos}: the position board is not a slice of the "
                       f"overall order; they differ at {pos}{i + 1} "
                       f"({by_rank[i]} against {from_overall[i]})")

    # -- every record, on its own ------------------------------------------
    for r in records:
        for field in ("player_name", "team", "position"):
            if not str(r.get(field) or "").strip():
                bad.append(f"{r['player_name'] or '(unnamed)'}: blank {field}")
        if r["position"] not in POSITIONS:
            bad.append(f"{r['player_name']}: position {r['position']!r} is "
                       f"not one of {', '.join(POSITIONS)}")
        pts = r.get("projected_points")
        if pts is None:
            bad.append(f"{r['player_name']}: no projected points")
        elif pts < 0:
            bad.append(f"{r['player_name']}: negative projected points ({pts})")
        if r["manual_adjustment"] and not str(r.get("adjustment_reason") or "").strip():
            bad.append(f"{r['player_name']}: adjustment of "
                       f"{r['manual_adjustment']} with no reason")
        if r["manual_adjustment"] and r["adjustment_reason"] == DEFAULT_REASON:
            bad.append(f"{r['player_name']}: adjustment of "
                       f"{r['manual_adjustment']} still carries the default reason")
        if r.get("adp") is None and r.get("value_label") is not None:
            bad.append(f"{r['player_name']}: value label {r['value_label']!r} "
                       f"with no ADP behind it")


    # -- the sorting test --------------------------------------------------
    #
    # Not an input to the ranking. It is the check on it, and it has already
    # earned its place once.
    # An editorial override moves a player's rank and must move nothing else.
    for r in records:
        if not r["editorial_override"]:
            continue
        if not str(r.get("override_reason") or "").strip():
            bad.append(f"{r['player_name']}: editorial override with no "
                       f"published reason")
        if not str(r.get("ranked_ahead_of") or "").strip():
            bad.append(f"{r['player_name']}: editorial override naming no "
                       f"comparison player")
        vorp = round(r["projected_points"] - r["replacement_points"], 1)
        if abs(vorp - r["vorp"]) > 0.051:
            bad.append(f"{r['player_name']}: vorp {r['vorp']} is not "
                       f"projected minus replacement ({vorp}) -- an override "
                       f"has moved a number it must not touch")
        score = round(r["vorp"] + r["manual_adjustment"], 1)
        if abs(score - r["ranking_score"]) > 0.051:
            bad.append(f"{r['player_name']}: ranking_score {r['ranking_score']} "
                       f"is not vorp plus adjustment ({score}) -- an override "
                       f"has moved a number it must not touch")

    where = {norm(r["player_name"]): r for r in records}
    for a, b in EXPECTED_ORDER:
        ra, rb = where.get(norm(a)), where.get(norm(b))
        if ra is None or rb is None:
            bad.append(f"{a if ra is None else b} is not on the board")
            continue
        if ra["position"] == rb["position"] and \
                ra["position_rank"] > rb["position_rank"]:
            bad.append(f"{a} is {ra['position']}{ra['position_rank']} and {b} "
                       f"is {rb['position']}{rb['position_rank']}: {a} must "
                       f"rank above {b}")
        # The two boards are cut from one order, so a disagreement here means
        # that stopped being true.
        if ra["overall_rank"] and rb["overall_rank"]:
            same_pos = ra["position"] == rb["position"]
            by_pos = ra["position_rank"] < rb["position_rank"]
            by_all = ra["overall_rank"] < rb["overall_rank"]
            if not by_all:
                bad.append(f"{a} is overall {ra['overall_rank']} and {b} is "
                           f"overall {rb['overall_rank']}: {a} must rank above")
            if same_pos and by_pos != by_all:
                bad.append(f"the overall and {ra['position']} boards disagree "
                           f"about {a} and {b}")

    top = [r["player_name"] for r in
           sorted([x for x in records if x["overall_rank"]],
                  key=lambda x: x["overall_rank"])[:len(EXPECTED_OVERALL_TOP)]]
    for i, (want, have) in enumerate(zip(EXPECTED_OVERALL_TOP, top), 1):
        if want != have:
            bad.append(f"overall {i} is {have}, expected {want}")

    rbs = sorted([r for r in records if r["position"] == "RB"],
                 key=lambda r: r["position_rank"])
    got = [r["player_name"] for r in rbs[:len(EXPECTED_RB_TOP)]]
    for i, (want, have) in enumerate(zip(EXPECTED_RB_TOP, got), 1):
        if want != have:
            bad.append(f"RB{i} is {have}, expected {want}")
    for name, want in EXPECTED_POSITION_RANK.items():
        r = next((x for x in records if x["player_name"] == name), None)
        if r is None:
            bad.append(f"{name} is not on the board")
        elif r["position_rank"] != want:
            bad.append(f"{name} is {r['position']}{r['position_rank']}, "
                       f"expected {r['position']}{want}")
    return bad


# -------------------------------------------------------------------- CSS

PAGE_CSS = """
/* Text for a screen reader and nobody else. The adjustment column carries
   its direction in words here, so "+16.0" is not the only way to know which
   way a player moved -- colour and an arrow glyph are both invisible to
   somebody listening to the page. */
.visually-hidden{position:absolute!important;width:1px;height:1px;
  margin:-1px;padding:0;overflow:hidden;clip:rect(0 0 0 0);
  clip-path:inset(50%);white-space:nowrap;border:0}

/* ---- rankings ----
   The board is the page. Everything here is in service of reading down a
   column of two hundred players quickly: compact rows, one rule between
   them, no cell borders, and colour held back for the two things that
   genuinely need it -- the tier labels and the adjustments. */
.rkwrap{max-width:var(--content-wide-table);margin:0 auto;padding:0 1.1rem 4rem}
.rkhead{padding:1.6rem 0 .4rem}
.rkeyebrow{font:600 .7rem/1 var(--agate);letter-spacing:.16em;
  text-transform:uppercase;color:var(--signal);margin:0 0 .5rem}
.rkwrap h1{font:400 clamp(2rem,4.4vw,3rem)/1.05 var(--text);color:var(--ink);
  margin:0 0 .7rem;letter-spacing:-.01em}
.rkintro{font:400 1.02rem/1.55 var(--text);color:var(--muted);
  max-width:60ch;margin:0 0 .9rem}
.rkstatus{font:500 .72rem/1.4 var(--agate);letter-spacing:.09em;
  text-transform:uppercase;color:var(--quiet);margin:0 0 .5rem}
.rkmethlink{display:inline-flex;align-items:center;gap:.35rem;min-height:44px;
  font:600 .74rem/1 var(--agate);letter-spacing:.09em;text-transform:uppercase;
  color:var(--signal);text-decoration:none;border-bottom:1px solid transparent}
.rkmethlink:hover,.rkmethlink:focus-visible{border-bottom-color:var(--signal)}

/* ---- controls ---- */
.rkctl{position:sticky;top:0;z-index:6;background:var(--paper);
  padding:.7rem 0 .6rem;border-bottom:1px solid var(--rule);margin:0 0 .2rem}
.rktabs{display:flex;gap:.4rem;overflow-x:auto;-webkit-overflow-scrolling:touch;
  scrollbar-width:none;padding:0 0 .15rem}
.rktabs::-webkit-scrollbar{display:none}
.rktab{flex:0 0 auto;display:inline-flex;align-items:center;min-height:44px;
  padding:0 .95rem;border:1px solid var(--rule);border-radius:8px;
  background:transparent;color:var(--muted);text-decoration:none;
  font:600 .78rem/1 var(--agate);letter-spacing:.09em;text-transform:uppercase;
  white-space:nowrap}
.rktab:hover{border-color:var(--quiet);color:var(--ink)}
.rktab:focus-visible{outline:2px solid var(--signal);outline-offset:2px}
.rktab[aria-current="page"]{background:var(--signal);border-color:var(--signal);
  color:#08090B}
.rkfilters{display:flex;gap:.5rem;align-items:center;margin:.55rem 0 0;
  flex-wrap:wrap}
.rkfilters input,.rkfilters select{min-height:44px;background:var(--card);
  border:1px solid var(--rule);border-radius:8px;color:var(--ink);
  padding:0 .7rem;font:400 .95rem/1 var(--text)}
.rkfilters input{flex:1 1 15rem;min-width:0}
.rkfilters select{flex:0 0 auto;font-family:var(--agate);font-size:.85rem;
  letter-spacing:.05em}
.rkfilters input:focus-visible,.rkfilters select:focus-visible,
.rkclear:focus-visible{outline:2px solid var(--signal);outline-offset:2px}
.rkclear{min-height:44px;padding:0 .9rem;border:1px solid var(--rule);
  border-radius:8px;background:transparent;color:var(--quiet);cursor:pointer;
  font:600 .72rem/1 var(--agate);letter-spacing:.09em;text-transform:uppercase}
.rkclear:hover{color:var(--ink);border-color:var(--quiet)}
.rkcount{font:500 .72rem/1 var(--agate);letter-spacing:.08em;
  text-transform:uppercase;color:var(--quiet);margin:.55rem 0 0}
/* "Show more" is a real control, not a link pretending to be one: the pool
   below it is already in the markup and it only clears the hidden state. */
.rkmore{display:block;width:100%;min-height:48px;margin:.9rem 0 0;
  background:transparent;border:1px solid var(--rule);border-radius:8px;
  color:var(--ink);cursor:pointer;font:600 .76rem/1 var(--agate);
  letter-spacing:.1em;text-transform:uppercase}
.rkmore:hover{border-color:var(--signal);color:var(--signal)}
.rkmore:focus-visible{outline:2px solid var(--signal);outline-offset:2px}

/* ---- table ---- */
.rktable{width:100%;border-collapse:collapse;margin:.2rem 0 0}
.rktable th{position:sticky;top:calc(var(--rkctlh,124px));z-index:4;
  background:var(--paper);text-align:right;
  font:600 .66rem/1 var(--agate);letter-spacing:.1em;text-transform:uppercase;
  color:var(--quiet);padding:.6rem .5rem;border-bottom:1px solid var(--rule);
  white-space:nowrap}
.rktable th.l,.rktable td.l{text-align:left}
/* Numbers right, words left, headings and cells the same way. They were
   not: the headings were right-aligned and the cells took the browser
   default, so every column label sat over the gap to the right of the
   figures it named. */
.rktable td{padding:.5rem;border-bottom:1px solid var(--rule);
  font:400 .92rem/1.3 var(--text);color:var(--muted);vertical-align:top;
  white-space:nowrap;text-align:right}
.rktable tbody tr:hover td{background:rgba(255,255,255,.02)}
.rkrank{font:600 .95rem/1.3 var(--agate);color:var(--ink);width:3rem}
.rktier{font:600 .7rem/1.3 var(--agate);letter-spacing:.06em;color:var(--quiet);
  width:3rem}
.rkname{color:var(--ink);font-weight:600;white-space:normal;min-width:9rem}
.rkname a{color:var(--ink);text-decoration:none;
  border-bottom:1px solid rgba(198,245,60,.28)}
.rkname a:hover{color:var(--signal);border-bottom-color:var(--signal)}
.rkteam,.rkpos{font:500 .78rem/1.3 var(--agate);letter-spacing:.06em;
  color:var(--quiet)}
.rkpts{font:600 .92rem/1.3 var(--agate);color:var(--ink)}
.rkposrk{font:500 .82rem/1.3 var(--agate);color:var(--muted)}
/* On a position board the row's rank IS the position rank, so this column
   would print RB4 next to a 4. It stays in the markup because the phone
   layout reads it back into "LV . RB4 . 246.0 projected points", and it is
   hidden here rather than dropped so the two layouts stay one table. */
.rktable th.rkempty,.rktable td.rkempty{display:none}
/* The first eight, marked but not boxed. A rail on the rank cell reads as
   "the top of the board" without turning eight rows into eight cards. */
.rktable tr.top8 td{background:rgba(198,245,60,.035)}
.rktable tr.top8 td.rkrank{box-shadow:inset 2px 0 0 var(--signal)}

/* ---- adjustments ----
   Never colour alone: the sign is in the text and the accessible name
   spells out the direction, so the column survives being read aloud or
   printed in greyscale. */
.rkadj{font:600 .88rem/1.3 var(--agate);letter-spacing:.02em}
.rkadj.up{color:var(--signal)}
.rkadj.dn{color:#E2705C}
.rkadj.flat{color:var(--quiet)}
/* An editorial decision, set in the label face rather than the figure face,
   so it does not read as a number that happens to be spelled out. */
.rkadj.ed{color:var(--quiet);font:500 .72rem/1.3 var(--agate);
  letter-spacing:.07em;text-transform:uppercase;white-space:normal}
.rktable td.rkwhy{white-space:normal;max-width:26rem;font-size:.86rem;
  color:var(--quiet)}
.rktable td.rkname{white-space:normal}
.rkwhy details>summary{cursor:pointer;list-style:none;color:var(--quiet)}
.rkwhy details>summary::-webkit-details-marker{display:none}
.rkwhy details>summary:focus-visible{outline:2px solid var(--signal);
  outline-offset:2px}
.rkwhy details[open]>summary .whylong{color:var(--muted)}
.rkwhy .whyshort{display:none;color:var(--signal);font:600 .78rem/1.3 var(--agate);
  letter-spacing:.07em;text-transform:uppercase}
.rkwhy .whylong{display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;
  overflow:hidden}
.rkwhy details[open] .whylong{-webkit-line-clamp:unset}
.rkwhy p{margin:.35rem 0 0;color:var(--muted);font-size:.86rem;line-height:1.45}

/* ---- tier separators ----
   Subtle by design. A tier is context for the rows beneath it, not an
   event, and a badge on all two hundred rows is noise on a phone. */
.rktable td.rktier-cell,.rktier-row td{background:var(--card);
  padding:.4rem .6rem;border-bottom:1px solid var(--rule);text-align:left}
.rktier-lab{font:600 .68rem/1 var(--agate);letter-spacing:.14em;
  text-transform:uppercase;color:var(--signal)}
.rktier-note{font:400 .78rem/1.3 var(--text);color:var(--quiet);
  margin-left:.6rem}

/* ---- methodology ---- */
.rkmeth{max-width:var(--content-reading);margin:2.6rem auto 0;
  padding:1.6rem 0 0;border-top:1px solid var(--rule)}
.rkmeth h2{font:400 1.6rem/1.15 var(--text);color:var(--ink);margin:0 0 .8rem}
.rkmeth p{font:400 1rem/1.6 var(--text);color:var(--muted);max-width:66ch;
  margin:0 0 .9rem}
.rkmeth .rkdisc{border-left:2px solid var(--rule);padding-left:.9rem;
  color:var(--quiet)}
.rknote{font:400 .9rem/1.5 var(--text);color:var(--quiet);margin:.9rem 0 0}

/* ---- mobile ----
   Not the desktop table squeezed. Below 720px each row becomes a compact
   block: rank and tier, then the name, then the line a drafter actually
   scans -- team, position rank, points. The adjustment and its reason
   appear only where there is one, because an empty line on 190 of 200
   cards is 190 wasted rows of screen. */
@media (max-width:720px){
  .rkwrap{padding:0 .85rem 3rem}
  .rktable thead{display:none}
  .rktable,.rktable tbody,.rktable tr,.rktable td{display:block;width:auto}
  .rktable tr.r{position:relative;padding:.7rem .2rem .75rem;
    border-bottom:1px solid var(--rule)}
  .rktable tr.top8{background:rgba(198,245,60,.035);
    box-shadow:inset 2px 0 0 var(--signal)}
  .rktable tr.top8 td{background:transparent}
  /* The desktop right-alignment for figures makes no sense once a row is a
     block: every line starts at the left edge of the card. */
  .rktable td{padding:0;border:0;white-space:normal;text-align:left}
  .rktable td.rkpos{display:none}
  /* Line one: #4 . TIER 1 */
  /* 14px floor. These carry the rank and the tier, which is meaningful
     information, not a caption -- .8rem computed to 12.8px. */
  .rktable td.rkrank{display:inline;font-size:.875rem;color:var(--quiet);
    letter-spacing:.06em}
  .rktable td.rkrank::before{content:"#"}
  /* Padding, not spaces inside content: a leading or trailing space in a
     generated string is collapsed and the dots end up welded to the word
     after them -- "DET .RB1". */
  .rktable td.rktier::before,.rktable td.rkposrk::before,
  .rktable td.rkpts::before{padding:0 .34rem}
  .rktable tr.top8 td.rkrank{box-shadow:none}
  .rktable td.rktier{display:inline;font-size:.875rem;color:var(--quiet)}
  /* "#4 . TIER 1" spelled out, where the desktop column says T1. */
  .rktable td.rktier .tpre{display:none}
  /* The space is an escape terminator and is eaten, so "\\00b7 TIER" renders
     as ".TIER". A non-breaking space escape survives because the T that
     follows it is not a hex digit and ends the escape on its own. */
  .rktable td.rktier::before{content:"\\00b7\\00a0TIER"}
  /* Line two: the name, at a size that is actually the point of the row. */
  .rktable td.rkname{display:block;font-size:1.05rem;margin:.15rem 0 .1rem}
  /* Line three: LV . RB4 . 246.0 projected points */
  .rktable td.rkteam{display:inline;font-size:.88rem}
  .rktable td.rkposrk,.rktable td.rkempty{display:inline;font-size:.88rem}
  .rktable td.rkposrk::before{content:"\\00b7"}
  .rktable td.rkpts{display:inline;font-size:.88rem}
  .rktable td.rkpts::before{content:"\\00b7"}
  .rktable td.rkpts::after{content:" projected points";color:var(--quiet);
    font-weight:400}
  /* Line four, only where there is an adjustment. */
  .rktable td.rkadj{display:block;margin:.3rem 0 0;font-size:.85rem}
  .rktable td.rkadj.flat{display:none}
  .rktable td.rkadj::after{content:" ranking adjustment";color:var(--quiet);
    font-weight:400;letter-spacing:0}
  /* "Editorial ranking decision" is already a whole phrase. */
  .rktable td.rkadj.ed::after{content:none}
  .rktable td.rkwhy{display:block;margin:.25rem 0 0;max-width:none}
  .rktable td.rkwhy.plain{display:none}
  .rkwhy .whyshort{display:inline-flex;align-items:center;min-height:44px}
  .rkwhy .whylong{display:none}
  .rktable td.rkwhy p{font-size:.9rem}
  .rktier-row td{display:block;padding:.55rem .2rem}
  .rkctl{padding:.6rem 0 .55rem}
  .rkfilters{flex-direction:column;align-items:stretch}
  .rkfilters input,.rkfilters select,.rkclear{width:100%;flex:0 0 auto}
}
"""

PAGE_JS = """
<script>
(function(){
  var t = document.getElementById('rktable');
  if (!t) return;
  var q = document.getElementById('rkq'),
      team = document.getElementById('rkteam'),
      clear = document.getElementById('rkclear'),
      more = document.getElementById('rkmore'),
      count = document.getElementById('rkcount'),
      rows = Array.prototype.slice.call(t.querySelectorAll('tr.r')),
      tiers = Array.prototype.slice.call(t.querySelectorAll('tr.rktier-row')),
      total = rows.length,
      capped = rows.filter(function(r){ return r.classList.contains('beyond'); }).length,
      expanded = capped === 0,
      // From an attribute, not by reading the button's own label back: the
      // label is display text and may be rewritten, and parsing it here
      // needed an escape that Python's string literal ate.
      noun = (more && more.getAttribute('data-noun')) || 'players';

  function apply(){
    var s = (q.value || '').trim().toLowerCase(),
        tm = team.value,
        filtering = !!(s || tm),
        shown = 0, matched = 0;

    rows.forEach(function(r){
      var hit = (!s || r.getAttribute('data-name').indexOf(s) > -1) &&
                (!tm || r.getAttribute('data-team') === tm);
      if (hit) matched++;
      // The cap is a display default, not a filter. A search reaches the
      // whole pool -- a player the board projects must be findable whether
      // or not he happens to sit past the fiftieth row.
      var visible = hit && (expanded || filtering || !r.classList.contains('beyond'));
      r.hidden = !visible;
      if (visible) shown++;
    });

    // A tier heading with nothing under it is a label for nothing.
    tiers.forEach(function(h){
      var n = h.nextElementSibling, any = false;
      while (n && !n.classList.contains('rktier-row')) {
        if (n.classList.contains('r') && !n.hidden) { any = true; break; }
        n = n.nextElementSibling;
      }
      h.hidden = !any;
    });

    count.textContent = filtering
      ? matched + ' of ' + total + ' ' + noun
      : (expanded ? total + ' ' + noun
                  : shown + ' of ' + total + ' ' + noun);
    clear.hidden = !filtering;
    // Nothing left to reveal once a search is showing the whole pool.
    if (more) more.hidden = expanded || filtering;
  }

  q.addEventListener('input', apply);
  team.addEventListener('change', apply);
  clear.addEventListener('click', function(){
    q.value = ''; team.value = ''; apply(); q.focus();
  });
  if (more) {
    more.addEventListener('click', function(){
      expanded = true;
      more.setAttribute('aria-expanded', 'true');
      apply();
      // Focus the first row revealed, or the button's replacement would be
      // nowhere and a keyboard lands back at the top of the document.
      var first = rows.filter(function(r){ return r.classList.contains('beyond'); })[0];
      if (first) { var a = first.querySelector('a, td'); if (a) { a.setAttribute('tabindex','-1'); a.focus(); } }
    });
  }
  clear.hidden = true;
  apply();

  // aria-expanded on every disclosure, kept in step with the element.
  Array.prototype.forEach.call(t.querySelectorAll('details.whyd'), function(d){
    var s = d.querySelector('summary');
    if (!s) return;
    s.setAttribute('aria-expanded', d.open ? 'true' : 'false');
    d.addEventListener('toggle', function(){
      s.setAttribute('aria-expanded', d.open ? 'true' : 'false');
    });
  });

  // The sticky column heads sit under the sticky control bar, so the offset
  // is what that bar actually measures rather than a guess that breaks when
  // the filters wrap onto two lines.
  var bar = document.querySelector('.rkctl');
  function offset(){
    if (!bar) return;
    document.documentElement.style.setProperty(
      '--rkctlh', bar.getBoundingClientRect().height + 'px');
  }
  offset();
  window.addEventListener('resize', offset);
})();
</script>
"""


# ------------------------------------------------------------------ page

def existing_slugs():
    """Player directories that exist, so the board only links where a page is.

    The projections board shipped 614 plain names because it ran before the
    player pages did. Same rule here: link where there is something to link
    to, and print plain text everywhere else.
    """
    d = SITE / SPORT
    if not d.is_dir():
        return set()
    return {p.name for p in d.glob("*") if p.is_dir()}


def adj_cell(r):
    """The adjustment column: a number only where there is a number.

    An editorial override moves a player's rank and no figure exists for it,
    so none is printed. Inventing one -- "+4.2" for a decision that was never
    arithmetic -- would be a fabricated statistic on a page whose whole claim
    is that the numbers are real.

    And the players who were passed carry nothing at all. Being ranked below
    somebody is not a downgrade, and printing one against Taylor or Jacobs
    would invent a second fiction to explain the first.
    """
    a = r["manual_adjustment"]
    if not a:
        if r.get("editorial_override"):
            return ('<td class="rkadj ed" data-lab="Adjustment">'
                    'Editorial ranking decision</td>')
        return ('<td class="rkadj flat" data-lab="Adjustment">'
                '<span class="visually-hidden">No adjustment</span></td>')
    up = a > 0
    # U+2212 for the minus sign, because a hyphen next to a number reads as
    # a dash and is announced as one.
    txt = f"+{a:.1f}" if up else f"−{abs(a):.1f}"
    arrow = "▲" if up else "▼"
    word = "raised" if up else "lowered"
    return (f'<td class="rkadj {"up" if up else "dn"}" data-lab="Adjustment">'
            f'<span aria-hidden="true">{arrow} </span>{txt}'
            f'<span class="visually-hidden"> points, {word}</span></td>')


def why_cell(r):
    """Why the player sits where he does.

    A player can carry both a scored adjustment and an editorial override --
    Jeanty does -- and they are separate claims, so both are published rather
    than one standing in for the other.

    Unadjusted, unoverridden players get one flat line and no disclosure.
    There is nothing to disclose, and 200 empty expanders is how a board
    starts looking like a form.
    """
    adj, override = r["manual_adjustment"], r.get("editorial_override")
    if not adj and not override:
        return f'<td class="l rkwhy plain" data-lab="Why">{esc(DEFAULT_WHY)}</td>'

    parts = []
    if adj and str(r.get("adjustment_reason") or "").strip():
        parts.append(str(r["adjustment_reason"]).strip())
    if override and str(r.get("override_reason") or "").strip():
        parts.extend(str(r["override_reason"]).strip().split("\n\n"))

    higher = override or adj > 0
    short = "Why we're higher" if higher else "Why we're lower"
    body = "".join(f"<p>{esc(x)}</p>" for x in parts)
    return (f'<td class="l rkwhy" data-lab="Why"><details class="whyd">'
            f'<summary><span class="whyshort">{short}</span>'
            f'<span class="whylong">{esc(parts[0]) if parts else ""}</span>'
            f'</summary>{body}</details></td>')


INITIAL_ROWS = 50


def page_records(records, pos):
    """What a page lists.

    Overall is the published 200. A position page is the whole pool at that
    position -- every player the board projects, ranked against each other --
    because a position rank that only counted players who had made a
    different cut would be a different number wearing the same name.
    """
    if not pos:
        return [r for r in records if r["overall_rank"] is not None]
    return sorted([r for r in records if r["position"] == pos],
                  key=lambda r: r["position_rank"])


def rows_html(records, pos, slugs):
    """The board, written into the HTML at build time.

    Proprietary data belongs in the initial markup: a board rendered by
    JavaScript is a board a crawler reads as an empty table. The script
    filters these rows; it never creates them.
    """
    out, ncols = [], (7 if pos else 9)
    seen_tier = None
    # Everything is written into the HTML; the tail is written hidden.
    #
    # Not paginated and not lazy-loaded: a crawler and a search box both need
    # the whole pool in the markup, and 255 rows of eight short cells is a
    # smaller page than one screenshot. `hidden` is the initial state only --
    # "Show more" clears it, and a search clears it for anything that matches.
    cap = INITIAL_ROWS if pos else None
    for n, r in enumerate(records, 1):
        tier = r["position_tier"] if pos else r["overall_tier"]
        rank = r["position_rank"] if pos else r["overall_rank"]
        beyond = cap is not None and n > cap
        if tier != seen_tier:
            seen_tier = tier
            note = ""
            if tier == 1:
                note = ('<span class="rktier-note">Clear top of the board.'
                        '</span>')
            out.append(
                f'<tr class="rktier-row{" beyond" if beyond else ""}"'
                f'{" hidden" if beyond else ""}>'
                f'<td class="rktier-cell" colspan="{ncols}">'
                f'<span class="rktier-lab">Tier {tier}</span>{note}</td></tr>')

        sl = slugify(r["player_name"])
        name = (f'<a href="/{SPORT}/{sl}/">{esc(r["player_name"])}</a>'
                if sl in slugs else esc(r["player_name"]))
        top8 = " top8" if (not pos and r["overall_rank"] <= 8) else ""
        cells = [
            f'<td class="rkrank">{rank}</td>',
            f'<td class="l rktier" data-lab="Tier">'
            f'<span class="tpre">T</span>{tier}</td>',
            f'<td class="l rkname" data-lab="Player">{name}</td>',
            f'<td class="l rkteam" data-lab="Team">{esc(r["team"])}</td>',
        ]
        if not pos:
            cells.append(f'<td class="l rkpos" data-lab="Pos">'
                         f'{esc(r["position"])}</td>')
            cells.append(f'<td class="rkposrk" data-lab="Pos rank">'
                         f'{esc(r["position"])}{r["position_rank"]}</td>')
        else:
            # The position board's own rank IS the position rank, so the
            # column would repeat the first one. It stays out of the table
            # and comes back on the phone, where the line reads
            # "LV . RB4 . 246.0 projected points".
            cells.append(f'<td class="rkposrk rkempty" data-lab="Pos rank">'
                         f'{esc(r["position"])}{r["position_rank"]}</td>')
        cells.append(f'<td class="rkpts" data-lab="Proj pts">'
                     f'{r["projected_points"]:.1f}</td>')
        cells.append(adj_cell(r))
        cells.append(why_cell(r))
        out.append(
            f'<tr class="r{top8}{" beyond" if beyond else ""}"'
            f'{" hidden" if beyond else ""} '
            f'data-name="{esc(r["player_name"].lower())}" '
            f'data-team="{esc(r["team"])}">' + "".join(cells) + "</tr>")
    return "\n".join(out), ncols


def head_row(pos):
    cols = ["RK", "TIER", "PLAYER", "TEAM"]
    if not pos:
        cols += ["POS", "POS RK"]
    cols += ["PROJ PTS", "ADJUSTMENT", "WHY"]
    ths = []
    for c in cols:
        left = (' class="l"' if c in ("PLAYER", "TEAM", "POS", "TIER", "WHY")
                else "")
        ths.append(f'<th scope="col"{left}>{c}</th>')
    if pos:
        # Kept in the markup for the mobile line, hidden in the table.
        ths.insert(4, '<th scope="col" class="rkempty">POS RK</th>')
    return "<tr>" + "".join(ths) + "</tr>"


METHODOLOGY = """
<section class="rkmeth" id="methodology">
  <h2>How LineupBeat Rankings Work</h2>
  <p>LineupBeat projections estimate each player's full-season statistical
     production. Rankings convert those projections into draft value by
     comparing each player with the expected replacement-level option at his
     position.</p>
  <p>The default rankings assume a 12-team, one-quarterback Half-PPR league
     starting one quarterback, two running backs, two wide receivers, one
     tight end and two flex players.</p>
  <p>Replacement levels are QB13, RB37, WR49 and TE13. A player's ranking
     score is his projected points above that replacement level, plus any
     documented editorial adjustment.</p>
  <p>Editorial adjustments do not silently change a player's statistical
     projection. They are reserved for circumstances such as unusual role
     security, durability risk or evidence that the median projection may not
     fully represent the player's draft value. Every adjustment must include
     a published reason.</p>
  <p class="rkdisc">ADP is not currently included in these rankings.
     Market-value labels will be added only after a verified and consistently
     updated ADP source is available.</p>
  <p class="rknote">Projected points stay frozen. A documented ranking
     adjustment can move a player's draft order without changing the
     projection it was applied to.</p>
</section>
"""


def page_html(records, pos, built, replacement):
    """One component. Overall when `pos` is None, otherwise that position."""
    shown = page_records(records, pos)
    slugs = existing_slugs()
    body_rows, ncols = rows_html(shown, pos, slugs)
    teams = sorted({r["team"] for r in shown if r["team"]})
    tabs = [("Overall", f"/{SPORT}/rankings/", pos is None)]
    for p in POSITIONS:
        tabs.append((p, f"/{SPORT}/rankings/{p.lower()}/", pos == p))
    tabs_html = "".join(
        f'<a class="rktab" href="{href}"'
        + (' aria-current="page"' if cur else "")
        + f">{esc(label)}</a>"
        for label, href, cur in tabs)

    h1 = (f"{SEASON} Fantasy Football Rankings" if not pos
          else f"{SEASON} Fantasy {pos} Rankings")
    crumb_last = "Rankings" if not pos else f"{pos} rankings"
    crumbs = (
        '<nav class="crumbs" aria-label="Breadcrumb">'
        '<a href="/">Home</a><span>/</span>'
        f'<a href="/{SPORT}/data/">NFL</a><span>/</span>'
        + (f'<a href="/{SPORT}/rankings/">Rankings</a><span>/</span>'
           f'<b aria-current="page">{esc(pos)}</b>' if pos else
           '<b aria-current="page">Rankings</b>')
        + "</nav>")

    intro = ("Overall and position-by-position Half-PPR draft rankings built "
             "from LineupBeat's full statistical projections. Rankings account "
             "for positional replacement value and clearly documented "
             "editorial adjustments.")
    if pos:
        intro = (f"Half-PPR {pos} draft rankings built from LineupBeat's full "
                 f"statistical projections, with positional replacement value "
                 f"({pos}{REPLACEMENT_RANK[pos]} at "
                 f"{replacement.get(pos, 0):.1f} points) and clearly "
                 f"documented editorial adjustments.")

    team_opts = "".join(f'<option value="{esc(t)}">{esc(t)}</option>'
                        for t in teams)

    hidden = max(0, len(shown) - INITIAL_ROWS) if pos else 0
    more_button = ""
    count_label = f"{len(shown)} players"
    if hidden:
        count_label = f"{min(INITIAL_ROWS, len(shown))} of {len(shown)} {pos}s"
        more_button = (
            f'  <button class="rkmore" id="rkmore" type="button" '
            f'data-noun="{esc(pos)}s" '
            f'aria-expanded="false" aria-controls="rktable">'
            f'Show all {len(shown)} {pos}s'
            f'<span class="visually-hidden"> in these rankings</span>'
            f'</button>')
    return f"""
<main class="rkwrap">
{crumbs}
  <header class="rkhead">
    <p class="rkeyebrow">{SEASON} FANTASY FOOTBALL</p>
    <h1>{esc(h1)}</h1>
    <p class="rkintro">{esc(intro)}</p>
    <p class="rkstatus">Updated {built:%B %Y} &middot; Half-PPR &middot;
       12-team, one-QB leagues</p>
    <a class="rkmethlink" href="#methodology">How these rankings work</a>
  </header>

  <div class="rkctl">
    <nav class="rktabs" aria-label="Rankings by position">{tabs_html}</nav>
    <div class="rkfilters">
      <label class="visually-hidden" for="rkq">Search players</label>
      <input id="rkq" type="search" placeholder="Search players"
             autocomplete="off">
      <label class="visually-hidden" for="rkteam">Filter by team</label>
      <select id="rkteam"><option value="">All teams</option>{team_opts}</select>
      <button class="rkclear" id="rkclear" type="button">Clear filters</button>
    </div>
    <p class="rkcount" id="rkcount">{count_label}</p>
  </div>

  <table class="rktable" id="rktable">
    <caption class="visually-hidden">{esc(h1)}, Half-PPR, sorted by ranking
      score</caption>
    <thead>{head_row(pos)}</thead>
    <tbody>
{body_rows}
    </tbody>
  </table>
{more_button}

{METHODOLOGY}
{seo.related_html('rankings')}
</main>
"""


def site_chrome():
    tpl = SITE / "template.html"
    if not tpl.exists():
        return "", "", ""
    src = tpl.read_text()
    css = re.search(r"<style>(.*?)</style>", src, re.S)
    foot = re.search(r"<footer.*?</footer>", src, re.S)
    return (css.group(1) if css else ""), seo.site_nav("data"), \
           (foot.group(0) if foot else "")


def render(records, pos, built, replacement, css, header, footer):
    shown = page_records(records, pos)
    url = CANON if not pos else f"{CANON}{pos.lower()}/"
    title = TITLE if not pos else POS_TITLE.format(pos=pos)
    desc = DESCRIPTION if not pos else (
        f"LineupBeat's {SEASON} Half-PPR fantasy football {pos} rankings, with "
        f"projected points, tiers, replacement value and documented ranking "
        f"adjustments.")

    trail = [("LineupBeat", "/"), ("Fantasy data", f"/{SPORT}/data/"),
             ("Rankings", f"/{SPORT}/rankings/")]
    if pos:
        trail.append((f"{pos} rankings", f"/{SPORT}/rankings/{pos.lower()}/"))

    # The players actually on this page, in the order the page shows them.
    items = [(r["position_rank"] if pos else r["overall_rank"],
              f'{r["player_name"]} ({r["team"]}, {r["position"]})',
              f'/{SPORT}/{slugify(r["player_name"])}/'
              if slugify(r["player_name"]) in existing_slugs() else None)
             for r in shown]

    dataset = {
        "@type": "Dataset",
        "name": (f"{SEASON} Half-PPR fantasy football rankings"
                 + (f", {pos}" if pos else ", overall top 200")),
        "description": desc,
        "url": url,
        "dateModified": built.strftime("%Y-%m-%d"),
        "creator": {"@type": "Organization", "name": "LineupBeat"},
        "variableMeasured": ["Projected fantasy points", "Value over "
                             "replacement", "Ranking score", "Tier"],
        **seo.dataset_extras(temporal=str(SEASON)),
    }
    ldjson = seo.graph(
        dataset,
        seo.breadcrumbs(trail),
        seo.itemlist_schema(title, url, items),
        seo.ORGANISATION)

    body = page_html(records, pos, built, replacement)
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;500;600&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600&display=swap" rel="stylesheet">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical" href="{url}">
{seo.social_meta(title, desc, url, OG_IMAGE)}
<script type="application/ld+json">{ldjson}</script>
<style>{css}{PAGE_CSS}{seo.CRUMB_CSS}{seo.UI_CSS}{seo.RELATED_CSS}{seo.TEAMS_CSS}</style>
</head>
<body>
{header}
{body}
{footer}
{seo.TRACKING}{seo.VIEW_CONTENT}
{PAGE_JS}
</body>
</html>"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--projections", default="data/projections.xlsx",
                    help="the site's projection workbook (the default source)")
    ap.add_argument("--export", default=None,
                    help="Site Export sheet as .csv or .json, used instead")
    ap.add_argument("--base", default="https://lineupbeat.com")
    ap.add_argument("--dry-run", action="store_true",
                    help="build and validate, write nothing")
    args = ap.parse_args()

    # Production reads the live projection board. The frozen workbook is
    # provenance, reachable through --export, and deliberately not the
    # default: pinning it would freeze the rankings while the projections
    # page kept moving, and the two would disagree about the same player.
    historical = bool(args.export)
    if historical:
        src = Path(args.export)
        if not src.is_absolute():
            src = ROOT / src
        if not src.exists() and (ARCHIVE / src.name).exists():
            src = ARCHIVE / src.name
        if not src.exists():
            sys.exit(f"  no {args.export}")
        print(f"  explicit-source build: {src.name}. This is a historical or "
              f"audit rebuild, not the production feed.")
    else:
        src = ROOT / args.projections
        if not src.exists():
            sys.exit(f"  no {args.projections}")

    if src.suffix.lower() in (".xlsx", ".xlsm"):
        rows, sheet = read_rankings_workbook(src)
        if rows:
            origin = f"{src.name} [{sheet}]"
        else:
            # A projection workbook has no Source Data sheet; it has QB/RB/WR/TE
            # sheets with a Half PPR column. Same file extension, different
            # shape, so the fallback is by content rather than by name.
            rows = read_workbook(src)
            origin = f"{src.name} (projection board, Half PPR column)"
    else:
        rows = read_export(src)
        origin = f"{src.name} (site export)"

    if not rows:
        sys.exit(f"  no players read from {src.name}")

    # Counted, not assumed: whatever the source actually carries at the four
    # skill positions is what the board must reconcile against.
    source_count = len(rows)
    digest = sha256(src)

    cfg = load_config()
    records, replacement, unmatched = build_records(rows, cfg)
    built = eastern_now()
    print(f"  {source_count} players from {origin}")
    print(f"  projections sha256 {digest[:16]}...")
    if cfg["sha256"]:
        print(f"  config      sha256 {cfg['sha256'][:16]}...  "
              f"({len(cfg['adjustments'])} adjustment(s))")
    print("  replacement: " + "  ".join(
        f"{p}{REPLACEMENT_RANK[p]}={replacement.get(p, 0):.1f}"
        for p in POSITIONS))

    # The frozen artifact is reconciled against its manifest; the live
    # projection board is not pinned, because it is supposed to move.
    bad = (check_manifest(src, digest, source_count) if historical else [])
    bad += validate(records, source_count, unmatched)
    if bad:
        print(f"\n  {len(bad)} validation problem(s):")
        for x in bad[:12]:
            print(f"    {x}")
        if len(bad) > 12:
            print(f"    ... and {len(bad) - 12} more")
        sys.exit("\n  Not publishing. These are the page's own rules.")

    from collections import Counter
    c = Counter(r["position"] for r in records)
    print("  published: " + "  ".join(f"{c.get(p, 0)} {p}" for p in POSITIONS)
          + f"  ({len(records)} total)")
    adjusted = [r for r in records if r["manual_adjustment"]]
    for r in adjusted:
        print(f"    {r['player_name']}: {r['manual_adjustment']:+.1f} -> "
              f"{r['position']}{r['position_rank']}, overall {r['overall_rank']}")

    if args.dry_run:
        print("\n  dry run, nothing written")
        return 0

    # The published file is the last-known-good board.
    #
    # Every CI run starts from a fresh checkout, so when a gate fails the only
    # rankings that exist are the ones in the repository. That makes two
    # things matter: the file must never be half-written, and it must not
    # churn for no reason. A rebuilt-but-identical board that rewrote its own
    # timestamp would put a diff on twelve thousand lines every two hours and
    # bury the one commit where a rank actually moved.
    meta = {
        "season": SEASON,
        "scoring_format": SCORING_FORMAT,
        "projection_source_file": src.name,
        "projection_source_sha256": digest,
        "ranking_config_file": CONFIG.name if cfg["sha256"] else None,
        "ranking_config_sha256": cfg["sha256"],
        "source_player_count": source_count,
        "published_overall_count": TOP_N,
        "generated_at": built.replace(microsecond=0).isoformat(),
        **({"build_mode": "explicit-source"} if historical else {}),
    }

    previous = {}
    if DATA_JSON.exists():
        try:
            previous = json.loads(DATA_JSON.read_text())
        except (ValueError, OSError):
            previous = {}
    old_meta = previous.get("metadata") or {}
    same_inputs = (
        old_meta.get("projection_source_sha256") == digest
        and old_meta.get("ranking_config_sha256") == cfg["sha256"])
    if same_inputs and previous.get("players") == records:
        # Both inputs and every player identical: the board did not change,
        # so neither does the file. The timestamp says when the rankings last
        # moved, which is the honest thing for it to say.
        meta["generated_at"] = old_meta.get("generated_at", meta["generated_at"])
        # The pages carry the same stamp, so an unchanged board does not
        # advertise itself as freshly updated.
        try:
            built = datetime.fromisoformat(meta["generated_at"])
        except (TypeError, ValueError):
            pass
        print(f"\n  {DATA_JSON.relative_to(ROOT)} unchanged "
              f"(same inputs, same board) -- not rewritten, "
              f"generated_at still {meta['generated_at']}")
    else:
        payload = {"metadata": meta, "players": records}
        DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
        tmp = DATA_JSON.with_name(DATA_JSON.name + ".tmp")
        try:
            tmp.write_text(json.dumps(payload, indent=1) + "\n")
            # Validate what was actually written, not what we meant to write.
            # A serialisation fault would otherwise replace a good board with
            # a broken one, and this file is the fallback.
            reread = json.loads(tmp.read_text())
            again = validate(reread.get("players") or [], source_count)
            if again:
                for x in again[:6]:
                    print(f"    {x}")
                raise ValueError("the written file does not pass its own gates")
            os.replace(tmp, DATA_JSON)
        except Exception:
            tmp.unlink(missing_ok=True)
            raise
        why = ("inputs changed" if not same_inputs else "the board changed")
        print(f"\n  wrote {DATA_JSON.relative_to(ROOT)} ({why}: "
              f"{len(records)} players, {DATA_JSON.stat().st_size:,} bytes)")

    css, header, footer = site_chrome()
    for pos in [None] + POSITIONS:
        page = render(records, pos, built, replacement, css, header, footer)
        out = (SITE / SPORT / "rankings" / "index.html" if not pos
               else SITE / SPORT / "rankings" / pos.lower() / "index.html")
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(seo.check_page(page, str(out)))
        n = len(page_records(records, pos))
        print(f"  wrote {out.relative_to(ROOT)}  ({n} players, "
              f"{len(page):,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
