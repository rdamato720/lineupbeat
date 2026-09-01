#!/usr/bin/env python3
"""Build the projections page from a simple spreadsheet.

    python3 scripts/build_projections.py ~/Downloads/2026_Projections.xlsx
    python3 scripts/build_projections.py FILE --season 2026

One sheet per position, six columns: Rank, Player, Team, PPR, Half PPR,
Non-PPR. Drop in a new file every few days and rebuild; nothing else needs
to know.

WHY THIS IS SEPARATE FROM THE ENGINE

The engine, the frozen release and the evidence layer are still there and
still correct. They are how a projection gets derived when a projection
needs deriving. This is how a spreadsheet gets onto a website, which is a
different job, and running the first machine to do the second one was the
mistake.

If the sheet later comes from the engine rather than from research, nothing
here changes: it reads six columns and does not care who wrote them.

WHAT IT DOES WITH THE NUMBERS

Nothing. Ranks are the ones in the file, points are the ones in the file.
A page that quietly recomputes what it was given is a page you cannot check
against the source, and the source is the thing being reviewed every few
days.

The one thing it adds is a link: where a name matches the roster, the row
links to that player's wire page, so a projection and the news about it are
one click apart.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sqlite3
import sys
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))
import seo
import seo_faqs

ROOT = Path(__file__).resolve().parent.parent


def eastern_now():
    """The date a reader in the league's own time zone would call today.

    UTC rolls over at 8pm Eastern, so a page built in the evening was
    stamped tomorrow and looked a day ahead of the data it was showing.
    """
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/New_York"))
    except Exception:
        return eastern_now() - timedelta(hours=4)

SITE = ROOT / "site"
SPORT = "nfl"
POSITIONS = ["QB", "RB", "WR", "TE"]

# Stat columns per position. One definition, in seo.py, because the chip on
# seven hundred player pages reads the same list -- and did not, until it
# was moved there: the board dropped targets and reordered, the chip kept
# the old list, and a receiver's page showed a Targets figure the board no
# longer carried.
STAT_COLUMNS = seo.STAT_COLUMNS


def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)


def key(n):
    n = re.sub(r"[.'`’]", "", (n or "").lower())
    return " ".join(re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", n).split())


def slug(s: str) -> str:
    """The same slug build_pages.py uses for player page directories.

    The projections page was linking to /nfl/{player_id}/, which is the
    roster id -- /nfl/nfl-4984/. The pages are at /nfl/aamaris-brown/. Every
    one of those 610 links was a 404, and nothing checks a link that a build
    script writes.
    """
    s = re.sub(r"[^\w\s-]", "", (s or "").lower())
    return re.sub(r"[\s_]+", "-", s).strip("-")


def read_sheet(path: Path):
    """Every position sheet, as rows. Six columns, named or positional."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        pos = name.split()[0].upper()
        if pos not in POSITIONS:
            continue
        ws = wb[name]
        head = [str(c.value or "").strip().lower() for c in ws[1]]

        def col(*names, default=None):
            for n in names:
                if n in head:
                    return head.index(n)
            return default

        ci = {
            "rank": col("rank", default=0),
            "player": col("player", "name", default=1),
            "team": col("team", "tm", default=2),
            "bye": col("bye"),
            # Totals, by name only.
            #
            # There used to be a positional fallback to columns 3, 4 and 5,
            # which is where the simple sheet keeps them. On the full
            # workbook those columns are PATT, CMP and Pass Yds, so every
            # quarterback was published with 3,812 "standard points". A
            # header that does not name a column is a column that is not
            # there.
            "ppr": col("ppr", "ppr points"),
            "half": col("half ppr", "half", "half-ppr"),
            "std": col("non-ppr", "standard", "std", "non ppr"),
            # The stat line, where the sheet carries one.
            "targets": col("targets", "tgt"),
            "rec": col("rec", "receptions"),
            "recyd": col("rec yds", "receiving yards", "rec yards"),
            "rectd": col("rec td", "receiving td"),
            "ruatt": col("rush att", "rushing attempts", "carries"),
            "ruyd": col("rush yds", "rushing yards", "rush yards"),
            "rutd": col("rush td", "rushing td"),
            "patt": col("patt", "pass att", "attempts"),
            "cmp": col("cmp", "comp", "completions"),
            "payd": col("pass yds", "passing yards"),
            "patd": col("pass td", "passing td"),
            "int": col("int", "ints", "interceptions"),
            "fl": col("fl", "fumbles", "fumbles lost"),
        }
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            nm = r[ci["player"]] if ci["player"] < len(r) else None
            if not nm or not str(nm).strip():
                continue

            def num(k):
                i = ci[k]
                if i is None or i >= len(r) or r[i] is None:
                    return None
                try:
                    return float(r[i])
                except (TypeError, ValueError):
                    return None

            # QB sheets often carry a Locked/Calc points pair rather than a
            # PPR column; a quarterback scores the same in every format, so
            # one number fills all three.
            ppr = num("ppr")
            if ppr is None:
                lk = col("locked fpts", "calc fpts", "fpts")
                if lk is not None and lk < len(r) and r[lk] is not None:
                    try:
                        ppr = float(r[lk])
                    except (TypeError, ValueError):
                        ppr = None
            half = num("half")
            std = num("std")
            if ppr is not None and half is None and std is None:
                half = std = ppr

            row = {
                "rank": int(num("rank") or 0) or len(rows) + 1,
                "name": str(nm).strip(),
                "team": (str(r[ci["team"]]).strip().upper()
                         if ci["team"] is not None and ci["team"] < len(r)
                         and r[ci["team"]] else ""),
                "ppr": ppr, "half": half, "std": std,
                "pos": pos,
            }
            for k in ("bye", "targets", "rec", "recyd", "rectd", "ruatt",
                      "ruyd", "rutd", "patt", "cmp", "payd", "patd",
                      "int", "fl"):
                row[k] = num(k)
            rows.append(row)
        if rows:
            out[pos] = rows
    if not out:
        sys.exit(f"  no position sheets in {path.name}. Expected one of "
                 f"{POSITIONS} as sheet names.")
    return out


# ---------------------------------------------------------------- analysis
#
# Prose generated from the board rather than written once and left to rot.
#
# A projections page is a table with a caption, and a table is close to
# invisible to a search engine: it holds the numbers and none of the
# questions somebody typed to find them. What ranks is the paragraph
# explaining where the cliff falls and how deep the position is -- and that
# paragraph has to change when the numbers do, or it becomes a lie by the
# second update.

def tiers(rows, fmt="ppr"):
    """Where the position naturally breaks.

    The largest point gaps inside the top forty. Not a fixed cutoff: a
    year where the top two backs are miles clear looks different from one
    where twelve are level, and a page that says "tier one is the top five"
    every season is describing a convention rather than the board.
    """
    pts = sorted((r[fmt] for r in rows if r.get(fmt) is not None), reverse=True)
    if len(pts) < 12:
        return []
    top = pts[:min(40, len(pts))]
    gaps = [(top[i] - top[i + 1], i + 1) for i in range(len(top) - 1)]
    gaps.sort(reverse=True)
    # A break only counts if it is well clear of the ordinary step between
    # neighbours; otherwise every position "breaks" everywhere.
    typical = sorted(g for g, _ in gaps)[len(gaps) // 2]
    out = [(rank, round(gap, 1)) for gap, rank in gaps[:4]
           if gap > max(typical * 2.5, 4)]
    return sorted(out)


def depth_note(pos, rows, fmt="ppr"):
    """How far the useful players run, in the language of a starting spot."""
    starters = {"QB": 12, "RB": 24, "WR": 36, "TE": 12}[pos]
    pts = sorted((r[fmt] for r in rows if r.get(fmt) is not None), reverse=True)
    if len(pts) <= starters:
        return None
    last = pts[starters - 1]
    # How many are within a tenth of the last starter: the size of the
    # group a manager is really choosing between late.
    close = sum(1 for p in pts[starters:] if p >= last * 0.9)
    return starters, round(last, 1), close


def position_prose(pos, rows, season, fmt="ppr"):
    """Two or three sentences about the shape of the position."""
    name = {"QB": "quarterback", "RB": "running back",
            "WR": "wide receiver", "TE": "tight end"}[pos]
    plural = {"QB": "quarterbacks", "RB": "running backs",
              "WR": "wide receivers", "TE": "tight ends"}[pos]
    ranked = sorted([r for r in rows if r.get(fmt) is not None],
                    key=lambda x: -x[fmt])
    if not ranked:
        return ""

    out = []
    top = ranked[0]
    out.append(
        f"{top['name']} leads our {season} {name} projections at "
        f"{top[fmt]:.1f} PPR points.")

    br = tiers(rows, fmt)
    if br:
        rank, gap = br[0]
        after = ranked[rank]["name"] if rank < len(ranked) else ""
        if rank <= 3:
            out.append(
                f"The board separates early: {gap:.0f} points divide "
                f"{pos}{rank} from {pos}{rank + 1}, so the first "
                f"{'two' if rank == 2 else 'few'} {plural} sit in a tier of "
                f"their own before {after} begins the next group.")
        else:
            out.append(
                f"The clearest break comes after {pos}{rank}, where "
                f"{gap:.0f} points separate that group from {after} and the "
                f"tier below.")
        if len(br) > 1:
            later = ", ".join(f"{pos}{r}" for r, _g in br[1:3])
            out.append(f"Smaller steps follow after {later}.")

    d = depth_note(pos, rows, fmt)
    if d:
        starters, last, close = d
        if close >= 8:
            out.append(
                f"Depth is the story below that: {pos}{starters} projects "
                f"{last:.1f}, and {close} more {plural} land within ten "
                f"percent of him, so the difference between a late starter "
                f"and a waiver add is small.")
        else:
            out.append(
                f"It thins quickly after {pos}{starters} at {last:.1f} "
                f"points, with only {close} {plural} projected within ten "
                f"percent of that line.")
    return " ".join(out)


POS_FULL = {"QB": "quarterback", "RB": "running back",
            "WR": "wide receiver", "TE": "tight end"}
POS_PLURAL = {"QB": "Quarterback", "RB": "Running Back",
              "WR": "Wide Receiver", "TE": "Tight End"}

POS_FAQ = {
    "QB": [
        ("How many quarterbacks should I draft?",
         "Most single-quarterback leagues need one, and many managers take "
         "a second late as insurance or a matchup play. Because the "
         "position is deep, waiting is a defensible strategy: the gap "
         "between QB6 and QB12 is usually far smaller than the gap between "
         "RB6 and RB12."),
        ("Do quarterback projections change between PPR and standard?",
         "Barely. Quarterbacks rarely catch passes, so their totals are "
         "close to identical in all three formats. The scoring toggle "
         "matters far more at running back, receiver and tight end."),
        ("What makes a quarterback projection move most?",
         "Rushing volume. Passing yards and touchdowns are relatively "
         "predictable across a season; a quarterback who runs adds points "
         "no pocket passer can match, which is why the top of the position "
         "is dominated by mobile starters."),
    ],
    "RB": [
        ("How many running backs should I draft?",
         "Most lineups start two plus a flex, so four to six is typical. "
         "The position carries the highest injury rate, and a backup who "
         "inherits a starting role is the most valuable thing on a waiver "
         "wire, which is why managers draft depth here they would not at "
         "other positions."),
        ("Why do running back rankings change so much between PPR and "
         "standard?",
         "Because receiving work is worth a point a catch in PPR and "
         "nothing in standard. A back with 70 receptions gains 70 points "
         "in PPR, which can move him ten or more spots. Switch the scoring "
         "toggle and the board reorders to match."),
        ("What is a workhorse back?",
         "One who handles both early-down carries and passing-down work "
         "rather than splitting a backfield. Those players carry the "
         "highest projections because volume is the single most reliable "
         "input to a running back's season."),
    ],
    "WR": [
        ("How many wide receivers should I draft?",
         "Most lineups start two or three plus a flex, so five to seven is "
         "common. Receiver is the deepest position, which means the cost "
         "of waiting is lower than at running back and the late rounds "
         "still return usable starters."),
        ("Do wide receiver projections favour PPR?",
         "Strongly. A receiver with 100 catches gains 100 points in PPR "
         "and none in standard, so possession receivers rise sharply in "
         "full point formats while big-play receivers hold their value "
         "better in standard."),
        ("What drives a wide receiver projection most?",
         "Targets. Catch rate and yards per reception vary year to year, "
         "but a receiver who commands a large share of his team's throws "
         "has a floor that efficiency alone cannot provide."),
    ],
    "TE": [
        ("Should I draft a tight end early?",
         "It depends entirely on the shape of the position in a given "
         "year. When the top one or two are clear of the field by a wide "
         "margin, the advantage is real; when the tier is flat, waiting "
         "costs almost nothing. The gaps on this page show which of those "
         "years it is."),
        ("How many tight ends should I draft?",
         "One in most formats, two if you take a late flyer rather than an "
         "early starter. The position is shallow, so the difference "
         "between TE12 and TE20 is often small enough that streaming "
         "works."),
        ("Why are tight end projections lower than other positions?",
         "Tight ends block as well as run routes, so they see fewer "
         "targets than receivers on the same offense. A top tight end is "
         "valuable relative to his position, not relative to a top "
         "receiver."),
    ],
}


def position_page(pos, board, links, css, header, footer, season, stats_json):
    """A page per position, because that is what people search.

    "Fantasy football projections" is a term owned by four national sites.
    "Fantasy football RB projections 2026" is a question with real volume
    and far less competition, and it is a page this board can answer better
    than they can because the stat line is on it.
    """
    rows = sorted(board.get(pos, []), key=lambda r: r["rank"])
    if not rows:
        return None

    full = POS_FULL[pos]
    label = POS_PLURAL[pos]
    prose = position_prose(pos, rows, season)
    cols = stats_json.get(pos, [])

    body_rows = []
    for r in rows:
        pid = links.get(key(r["name"]))
        nm = (f'<a href="/{SPORT}/{pid}/">{esc(r["name"])}</a>' if pid
              else esc(r["name"]))
        cells = "".join(
            f'<td class="stat">'
            + ("" if r.get(k) is None else
               (f"{round(r[k]):,}" if k in ("payd", "recyd", "ruyd", "patt",
                                            "ruatt", "targets")
                else f"{r[k]:.1f}"))
            + "</td>" for k, _lab in cols)
        alt = "".join(
            f'<td class="alt dim">'
            + ("" if r.get(k) is None else f"{r[k]:.1f}") + "</td>"
            for k in ("half", "std"))
        main = "" if r.get("ppr") is None else f'{r["ppr"]:.1f}'
        body_rows.append(
            f'<tr><td class="l rk c-rk">{r["rank"]}</td>'
            f'<td class="l nm c-nm">{nm}</td>'
            f'<td class="l tm">{esc(r["team"])}</td>'
            + f'<td class="pt">{main}</td>' + cells + alt + "</tr>")

    heads = "".join(f'<th class="stat">{esc(lab)}</th>' for _k, lab in cols)
    return rows, prose, "\n".join(body_rows), heads



def check(board):
    """Say what is odd about the file before publishing it.

    Not a gate. A rank that skips, a missing format, a player listed twice --
    these are worth seeing on the way past, and none of them is a reason to
    refuse a page.
    """
    notes = []
    for pos, rows in board.items():
        seen = {}
        for r in rows:
            k = key(r["name"])
            if k in seen:
                notes.append(f"{pos}: {r['name']} appears twice "
                             f"(rows {seen[k]} and {r['rank']})")
            seen[k] = r["rank"]
        missing = [r["name"] for r in rows
                   if r["ppr"] is None or r["half"] is None
                   or r["std"] is None]
        if missing:
            notes.append(f"{pos}: {len(missing)} row(s) missing a format: "
                         f"{', '.join(missing[:3])}")
        ranks = [r["rank"] for r in rows]
        if sorted(ranks) != list(range(1, len(rows) + 1)):
            notes.append(f"{pos}: ranks are not 1..{len(rows)} without gaps")
        # PPR should never be below standard for a pass catcher.
        for r in rows:
            if (r["ppr"] is not None and r["std"] is not None
                    and r["ppr"] < r["std"] - 0.05):
                notes.append(f"{pos}: {r['name']} scores less in PPR "
                             f"({r['ppr']}) than standard ({r['std']})")
    return notes


def roster_links(conn):
    """player name -> wire page slug, where the wire knows the player."""
    links = {}
    # Only names the wire actually built a page for. Checking the directory
    # rather than the roster means a link is never written to a page that
    # does not exist, however the slug is derived.
    pages = SITE / SPORT
    if pages.exists():
        have = {d.name for d in pages.iterdir() if d.is_dir()}
    else:
        have = set()
    rp = ROOT / "rosters" / f"{SPORT}.csv"
    if rp.exists():
        for r in csv.DictReader(rp.open()):
            name = r.get("name", "")
            s = slug(name)
            if s in have:
                links[key(name)] = s
    return links


PAGE_CSS = """
/* The nav, identical on every page.
   These are anchors on a static page and buttons in the app, so the browser
   underlined them here and not there -- the same bar looking different
   depending which page you were on. And the accent pill marks where you
   are, which is what the app does and what these pages were not doing. */
.topbar .logo,.topbar .vbtn{text-decoration:none}
.topbar .vbtn:hover{text-decoration:none; color:var(--ink)}
.vbtn[aria-current="page"]{color:#0A0C08; background:var(--signal);
  border-color:var(--signal)}

/* ---- projections ----
   A board, not a dashboard. One table, three formats, four positions, and a
   search box; everything else people came here to read is a number in a
   row. */
.pbwrap{max-width:1080px; margin:0 auto; padding:0 1rem 4rem}
.pbhead{display:flex; align-items:flex-end; gap:1rem; flex-wrap:wrap;
        margin:1.6rem 0 .6rem}
.pbhead h1{font-size:1.7rem; margin:0; letter-spacing:-.01em;
        font-family:var(--text)}
.pbsub{color:var(--quiet); font-size:.84rem; margin:.35rem 0 0; max-width:64ch}
/* Up here rather than only in the footnote. A projection page that does not
   say when it was made is asking to be trusted on nothing, and the answer
   should not be at the bottom of a 614-row table. */
.pbdate{display:inline-block; margin-left:.5rem; font-family:var(--agate);
  text-transform:uppercase; letter-spacing:.06em; font-size:.7rem;
  color:var(--signal); border:1px solid var(--rule); border-radius:999px;
  padding:.1rem .5rem; vertical-align:.05em}
.pbctl{display:flex; gap:1rem; flex-wrap:wrap; align-items:center;
       margin:1.1rem 0 .5rem}
.pbtabs{display:flex; gap:.3rem; flex-wrap:wrap}
/* Links, not four paragraphs.
   The analysis of how deep each position is lives on that position's own
   page, where it is the reason the page ranks. Repeating all four here put
   a wall of prose between the headline and the board, and said the same
   thing twice. */
.boardnotes{display:flex; gap:.4rem; flex-wrap:wrap; align-items:center;
  margin:1.1rem 0 .2rem}
.boardnotes .posnav b{font-family:var(--data); color:var(--ink);
  font-weight:600; margin-left:.25rem}
.boardnotes .posnav:hover b{color:var(--signal)}
.pbsublab{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.07em; font-size:.66rem; color:var(--quiet);
  margin-right:.3rem}
.posnav{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.04em; font-size:.74rem; color:var(--quiet);
  border:1px solid var(--rule); border-radius:999px; padding:.3rem .7rem;
  text-decoration:none}
.posnav:hover{color:var(--signal); border-color:var(--signal)}
@media (max-width:760px){
  .boardnotes .posnav{min-height:44px; display:inline-flex;
    align-items:center}
}
/* `font:inherit` sat after `font-family` here and reset it, so these were
   the only filter pills on the site set in the serif rather than the
   agate. The shorthand is gone; seo.UI_CSS owns the rest. */
.pbtab{background:transparent; border:1px solid var(--rule);
       color:var(--quiet); cursor:pointer}
.pbtab:hover{color:var(--ink); border-color:var(--ink)}
.pbtab[aria-pressed="true"]{background:var(--signal); border-color:var(--signal);
       color:#0b0f0a; font-weight:600}
.pbfmt{display:flex; gap:.3rem; margin-left:auto}
.pbsearch input{background:var(--card); border:1px solid var(--rule);
       color:var(--ink); font:inherit; font-size:.82rem; padding:.36rem .7rem;
       border-radius:6px; min-width:190px}
/* 16px on a phone, because anything smaller makes iOS zoom the page on
   focus and the reader has to pinch back out to see the table he was
   filtering. */
@media (max-width:760px){
  .pbsearch{width:100%}
  .pbsearch input{width:100%; font-size:16px; min-height:44px}
}
.pbtbl{width:100%; border-collapse:collapse; font-size:.9rem;
       font-variant-numeric:tabular-nums}
.pbtbl .pt,.pbtbl .alt,.pbtbl .rk,.pbtbl .stat{font-family:var(--data)}
.pbtbl .stat{color:var(--quiet); font-size:.82rem}
.pbtbl tbody tr:hover .stat{color:var(--ink)}
.pbtbl th{text-align:right; font-family:var(--agate); font-size:.72rem; letter-spacing:.09em;
       text-transform:uppercase; color:var(--quiet); font-weight:600;
       padding:.5rem .55rem; border-bottom:1px solid var(--rule);
       position:sticky; top:0; background:var(--paper)}
.pbtbl th.l,.pbtbl td.l{text-align:left}
.pbtbl td{padding:.42rem .55rem; border-bottom:1px solid var(--rule);
       text-align:right}
.pbtbl tbody tr:hover{background:var(--card)}
.pbtbl .rk{color:var(--quiet); font-size:.78rem; width:3.2rem}
.pbtbl .nm{font-weight:500}
.pbtbl .nm a{color:var(--ink); text-decoration:none;
       border-bottom:1px solid transparent}
.pbtbl .nm a:hover{border-bottom-color:var(--signal); color:var(--signal)}
.pbtbl .tm{color:var(--quiet); font-size:.76rem; letter-spacing:.05em;
       width:3.4rem}
.pbtbl .pt{font-weight:600; width:4.6rem; color:var(--ink)}
.pbtbl .alt{width:4.4rem}
.pbtbl .dim{color:var(--quiet); font-weight:400}
.pbnote{color:var(--quiet); font-size:.78rem; margin:1.2rem 0 0;
       max-width:72ch; line-height:1.55}
.pbempty{color:var(--quiet); padding:1.4rem .55rem; font-size:.86rem}

/* Touch targets on a phone.
   These pills are ~30px tall, which is fine for a cursor and small for a
   thumb -- the platform guidance is 44. Padding rather than height, so the
   text stays where it is and only the box a finger can hit grows. */
@media (max-width:760px){
  .pbtab{min-height:44px; display:inline-flex; align-items:center;
    padding-top:.5rem; padding-bottom:.5rem}
}
@media (max-width:640px){
  /* The stat columns used to be display:none here, so a phone got rank,
     player, team and a points total and none of the numbers the ranking
     is built from -- on the layout most people read this in. They scroll
     sideways now, inside .xtab, and nothing is dropped.

     The two alternate scoring formats do still go. They are the same
     player scored three ways rather than three facts about him, the
     format buttons above the table already switch between them, and at
     390px they are two columns of width that push the stat line further
     from the reader for no new information. */
  .pbtbl .alt{display:none}
  .pbfmt{margin-left:0; width:100%}
}

/* Inside a scroll container the table stops being width-constrained, so
   it may take the width its columns need. Without this it compresses
   back to the viewport and the columns wrap after all. */
.xtab .pbtbl{width:auto; min-width:100%}
/* The shared component draws the row lines as inset shadows, which is what
   lets them stick with a pinned cell. The table's own borders would sit
   directly on top of them and read as a double rule. */
.xtab .pbtbl td, .xtab .pbtbl th{border-bottom:0}
/* Points, the number the ranking is built on, reads as a total rather
   than as one more stat: the stats beside it are deliberately quiet. */
.pbtbl .pt{color:var(--signal)}
.pbtbl thead th.pt{color:var(--signal)}
@media (max-width:900px){
  /* 13px floor, from a .82rem stat cell that measured 13.1 and a rank and
     team that measured 12.5 and 12.2. Small numbers are the whole content
     of this page. */
  .pbtbl .stat{font-size:.84rem}
  .pbtbl .rk{font-size:.82rem}
  .pbtbl .tm{font-size:.82rem}
}
"""


def build_html(board, links, css, header, footer, season, source_name, notes):
    total = sum(len(v) for v in board.values())
    built = eastern_now()

    rows_json = {}
    for pos, rows in board.items():
        rows_json[pos] = [
            {"r": r["rank"], "n": r["name"], "t": r["team"],
             "p": r["ppr"], "h": r["half"], "s": r["std"],
             "id": links.get(key(r["name"])),
             **{k: r.get(k) for k in
                ("targets", "rec", "recyd", "rectd", "ruatt", "ruyd",
                 "rutd", "patt", "cmp", "payd", "patd", "int", "fl")}}
            for r in sorted(rows, key=lambda x: x["rank"])
        ]

    # Which stat columns each position gets, and only where the sheet
    # actually carried them. A table of empty columns is worse than a
    # narrower table: it looks like missing data rather than data that was
    # never part of this position.
    STATS = STAT_COLUMNS
    stats_json = {}
    for pos, rows in board.items():
        cols = [(k, lab) for k, lab in STATS.get(pos, [])
                if any(r.get(k) is not None for r in rows)]
        stats_json[pos] = cols
    has_stats = any(stats_json.values())

    counts = "  ".join(f"{p} {len(board.get(p, []))}" for p in POSITIONS
                       if board.get(p))

    # Only where the page exists. A link to a changelog that was never
    # built is worse than no link.
    changes_link = (
        f'<a class="posnav" href="/{SPORT}/projections/changes/">'
        f'What changed</a>'
        if (SITE / SPORT / "projections" / "changes" / "index.html").exists()
        else "")

    # The default view, in the HTML.
    #
    # An empty tbody means a crawler reads a page about 614 players that
    # contains none of them. QB, PPR, rank order -- the same thing draw()
    # produces on load, so there is no flicker when the script runs.
    _first = next((p for p in POSITIONS if board.get(p)), None)
    _cols = stats_json.get(_first, [])
    _static = []
    for r in sorted(board.get(_first, []), key=lambda x: x["rank"]):
        pid = links.get(key(r["name"]))
        nm = (f'<a href="/{SPORT}/{pid}/">{esc(r["name"])}</a>' if pid
              else esc(r["name"]))
        cells = "".join(
            f'<td class="stat">'
            f'{"" if r.get(k) is None else (f"{round(r[k]):,}" if k in ("payd","recyd","ruyd","patt","ruatt","targets") else f"{r[k]:.1f}")}'
            f'</td>' for k, _lab in _cols)
        alt = "".join(f'<td class="alt dim">'
                      f'{"" if r.get(k) is None else f"{r[k]:.1f}"}</td>'
                      for k in ("half", "std"))
        main = "" if r.get("ppr") is None else f'{r["ppr"]:.1f}'
        _static.append(
            f'<tr><td class="l rk c-rk">{r["rank"]}</td>'
            f'<td class="l nm c-nm">{nm}</td>'
            f'<td class="l tm">{esc(r["team"])}</td>'
            + cells
            + f'<td class="pt">{main}</td>'
            + alt + '</tr>')
    static = "\n".join(_static)

    body = f"""<main class="pbwrap">
  <nav class="crumbs" aria-label="Breadcrumb">
    <a href="/">LineupBeat</a><span>/</span>
    <a href="/{SPORT}/data/">Fantasy data</a><span>/</span>
    <b>Projections</b></nav>

  <div class="pbhead">
    <div>
      <h1>{season} Fantasy Football Projections</h1>
      <p class="pbsub">Full-season point projections for
        {total} players, in PPR, half PPR and standard scoring.
        Ranked within position.
        <span class="pbdate">Updated {built:%-m/%-d}</span></p>
    </div>
  </div>

{seo.byline_html(built)}
  <div class="boardnotes">
    <span class="pbsublab">By position</span>
    {"".join(f'<a class="posnav" href="/{SPORT}/projections/{p.lower()}/">'
             f'{POS_PLURAL[p]}s <b>{len(board[p])}</b></a>'
             for p in POSITIONS if board.get(p))}
    {changes_link}
  </div>

  <div class="pbctl">
    <div class="pbtabs" role="group" aria-label="Position">
      {''.join(f'<button class="pbtab" data-pos="{p}" '
               f'aria-pressed="{"true" if p == "QB" else "false"}">{p}</button>'
               for p in POSITIONS if board.get(p))}
    </div>
    <div class="pbfmt pbtabs" role="group" aria-label="Scoring">
      <button class="pbtab" data-fmt="p" aria-pressed="true">PPR</button>
      <button class="pbtab" data-fmt="h" aria-pressed="false">Half</button>
      <button class="pbtab" data-fmt="s" aria-pressed="false">Standard</button>
    </div>
    <div class="pbsearch">
      <input id="pbq" type="search" placeholder="Find a player"
             autocomplete="off" aria-label="Find a player">
    </div>
  </div>

  {seo.scroll_hint("the full stat line")}
  <div class="xtab" tabindex="0" role="region" aria-label="Projection board">
  <table class="pbtbl">
    <thead><tr id="pbhead"></tr></thead>
    <tbody id="pbbody">
{static}
    </tbody>
  </table>
  </div>
  <p class="pbempty" id="pbempty" hidden>No player by that name.</p>

  <p class="pbnote">
    Projections are full-season totals, not per-game. Points come straight
    from the source file. Ranks are within position for the scoring format
    you have selected, so a back who catches nothing rises in standard and
    falls in PPR. Last updated {built:%B %-d, %Y}.
  </p>
{seo.faq_html(seo_faqs.PROJECTIONS)}{seo.related_html('projections')}
</main>

<script>
const PB = {json.dumps(rows_json, separators=(',', ':'))};
const STATS = {json.dumps(stats_json, separators=(',', ':'))};
let pos = "QB", fmt = "p";
const body = document.getElementById("pbbody");
const empty = document.getElementById("pbempty");
const q = document.getElementById("pbq");
const LABEL = {{p: "PPR", h: "Half", s: "Standard"}};

function num(v) {{ return v === null || v === undefined ? "\\u2014" : v.toFixed(1); }}
// Yards and attempts are whole numbers; receptions and touchdowns are not,
// because a projection is an average of seasons that did not happen.
const WHOLE = new Set(["payd", "recyd", "ruyd", "patt", "ruatt", "targets"]);
function stat(v, k) {{
  if (v === null || v === undefined) return "\\u2014";
  return WHOLE.has(k) ? Math.round(v).toString() : v.toFixed(1);
}}

function draw() {{
  // The two formats you are not looking at, so the selected one never
  // appears twice in the same row.
  const alt = ["p", "h", "s"].filter(k => k !== fmt);
  const cols = STATS[pos] || [];
  document.getElementById("pbhead").innerHTML =
    `<th class="l rk c-rk">#</th><th class="l c-nm">Player</th>`
    + `<th class="l tm">Team</th>`
    + `<th class="pt">${{LABEL[fmt]}}</th>`
    + cols.map(c => `<th class="stat">${{c[1]}}</th>`).join("")
    + alt.map(k => `<th class="alt">${{LABEL[k]}}</th>`).join("");
  const term = (q.value || "").trim().toLowerCase();
  // Sorted by the format you are looking at.
  //
  // Switching to standard used to relabel the column and leave the order
  // alone, so the board stayed in PPR sequence while showing standard
  // points: the top of a standard list had a receiver above a back who
  // outscores him in that format. A ranking is a ranking under a scoring
  // rule, so changing the rule reorders the board.
  //
  // The name breaks ties, so two identical totals always land the same way.
  const rows = (PB[pos] || [])
    .filter(r => !term || r.n.toLowerCase().includes(term) ||
                 (r.t || "").toLowerCase().includes(term))
    .slice()
    .sort((a, b) => (b[fmt] ?? -1) - (a[fmt] ?? -1) || a.n.localeCompare(b.n));

  // Position within this format, computed over the whole board rather than
  // the filtered view, so searching for one player still shows his rank
  // rather than "1".
  const order = (PB[pos] || []).slice()
    .sort((a, b) => (b[fmt] ?? -1) - (a[fmt] ?? -1) || a.n.localeCompare(b.n));
  const rank = new Map(order.map((r, i) => [r.n + "|" + r.t, i + 1]));
  body.innerHTML = rows.map(r => {{
    const name = r.id
      ? `<a href="/{SPORT}/${{r.id}}/">${{r.n}}</a>` : r.n;
    return `<tr>
      <td class="l rk c-rk">${{rank.get(r.n + "|" + r.t) || r.r}}</td>
      <td class="l nm c-nm">${{name}}</td>
      <td class="l tm">${{r.t || ""}}</td>`
      + `<td class="pt">${{num(r[fmt])}}</td>`
      + cols.map(c => `<td class="stat">${{stat(r[c[0]], c[0])}}</td>`).join("")
      + alt.map(k => `<td class="alt dim">${{num(r[k])}}</td>`).join("")
      + `</tr>`;
  }}).join("");
  empty.hidden = rows.length > 0;
}}

document.querySelectorAll("[data-pos]").forEach(b =>
  b.addEventListener("click", () => {{
    pos = b.dataset.pos;
    document.querySelectorAll("[data-pos]").forEach(x =>
      x.setAttribute("aria-pressed", x === b ? "true" : "false"));
    draw();
  }}));
document.querySelectorAll("[data-fmt]").forEach(b =>
  b.addEventListener("click", () => {{
    fmt = b.dataset.fmt;
    document.querySelectorAll("[data-fmt]").forEach(x =>
      x.setAttribute("aria-pressed", x === b ? "true" : "false"));
    draw();
  }}));
q.addEventListener("input", draw);
draw();
</script>"""

    schema = {
        "@context": "https://schema.org",
        "@type": "Dataset",
        "name": f"{season} Fantasy Football Projections",
        "description": (f"Full-season fantasy football point projections for "
                        f"{total} NFL players in PPR, half PPR and standard "
                        f"scoring, ranked within position."),
        "url": f"https://lineupbeat.com/{SPORT}/projections/",
        "dateModified": built.strftime("%Y-%m-%d"),
        "creator": {"@type": "Organization", "name": "LineupBeat"},
        **seo.dataset_extras(temporal=str(season)),
        "variableMeasured": ["PPR points", "Half PPR points",
                             "Standard points", "Positional rank"],
    }
    crumbs = {
        "@context": "https://schema.org", "@type": "BreadcrumbList",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "LineupBeat",
             "item": "https://lineupbeat.com/"},
            {"@type": "ListItem", "position": 2, "name": "Fantasy data",
             "item": f"https://lineupbeat.com/{SPORT}/data/"},
            {"@type": "ListItem", "position": 3, "name": "Projections",
             "item": f"https://lineupbeat.com/{SPORT}/projections/"},
        ],
    }

    title = f"{season} Fantasy Football Projections | LineupBeat"
    desc = (f"Full-season fantasy projections for {total} NFL players. "
            f"PPR, half PPR and standard scoring, ranked by position. "
            f"Updated {built:%B %-d, %Y}.")

    # The board as a declared ranking, not just a table of numbers. Capped
    # at fifty because a 614-item list is noise, and the top of a ranking
    # is the part anybody searches for.
    _flat = sorted([r for rows in board.values() for r in rows if r.get("ppr")],
                   key=lambda x: -x["ppr"])[:50]
    itemlist = seo.itemlist_schema(
        f"{season} fantasy football projections, top 50 by PPR points",
        f"https://lineupbeat.com/{SPORT}/projections/",
        [(i, r["name"],
          f"/{SPORT}/{links[key(r['name'])]}/" if links.get(key(r["name"]))
          else None)
         for i, r in enumerate(_flat, 1)])

    # One graph rather than loose blocks: these are facets of one page, and
    # saying so lets a crawler connect the dataset to the site that
    # publishes it and to the questions it answers.
    ldjson = seo.graph(
        {k: v for k, v in schema.items() if k != "@context"},
        {k: v for k, v in crumbs.items() if k != "@context"},
        seo.faq_schema(seo_faqs.PROJECTIONS),
        seo.ORGANISATION,
        itemlist)

    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;500;600&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600&display=swap" rel="stylesheet">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical" href="https://lineupbeat.com/{SPORT}/projections/">
{seo.social_meta(title, desc, f"https://lineupbeat.com/{SPORT}/projections/")}
<script type="application/ld+json">{ldjson}</script>
<style>{css}{PAGE_CSS}{seo.CRUMB_CSS}{seo.UI_CSS}{seo.SCROLLTABLE_CSS}{seo.RELATED_CSS}{seo.TEAMS_CSS}{seo.TEAMS_CSS}{seo.BYLINE_CSS}</style>
</head>
<body>
{header}
{body}
{footer}
{seo.TRACKING}
</body>
</html>"""


def site_chrome():
    """The homepage's own CSS, header and footer, so this matches the site."""
    tpl = SITE / "template.html"
    if not tpl.exists():
        return "", "", ""
    src = tpl.read_text()
    css = re.search(r"<style>(.*?)</style>", src, re.S)
    header = seo.site_nav("projections")
    return (css.group(1) if css else ""), header, seo.site_footer()


def add_to_sitemap(url):
    """A page nobody links to and nobody lists is a page nobody finds."""
    sm = SITE / "sitemap.xml"
    if not sm.exists():
        return False
    text = sm.read_text()
    if url in text:
        return False
    today = eastern_now().strftime("%Y-%m-%d")
    entry = (f"  <url><loc>{url}</loc><lastmod>{today}</lastmod>"
             f"<changefreq>weekly</changefreq>"
             f"<priority>0.8</priority></url>\n")
    text = text.replace("</urlset>", entry + "</urlset>")
    sm.write_text(text)
    return True


def link_from_hub():
    """Add a card to the Fantasy Data hub, if there is one."""
    hub = SITE / SPORT / "data" / "index.html"
    if not hub.exists():
        return False
    text = hub.read_text()
    if "/projections/" in text:
        return False
    card = (
        f'<a class="dcard" href="/{SPORT}/projections/">'
        f'<h3>Projections</h3>'
        f'<p>Full-season points for every relevant player, in PPR, half PPR '
        f'and standard scoring, ranked by position.</p></a>\n'
    )
    # Insert before the closing of the card grid, or before </main>.
    for anchor in ('</div>\n</main>', '</main>'):
        if anchor in text:
            text = text.replace(anchor, card + anchor, 1)
            hub.write_text(text)
            return True
    return False



def write_position_pages(board, links, css, header, footer, season,
                         stats_json, built):
    """One page per position, each targeting its own query."""
    written = []
    for pos in POSITIONS:
        got = position_page(pos, board, links, css, header, footer, season,
                            stats_json)
        if not got:
            continue
        rows, prose, body_rows, heads = got
        full = POS_FULL[pos]
        label = POS_PLURAL[pos]
        n = len(rows)
        others = [x for x in POSITIONS if x != pos and board.get(x)]

        faq = POS_FAQ[pos] + [
            (f"How many {full}s are projected here?",
             f"All {n} {full}s with a meaningful projected role for "
             f"{season}, ranked by PPR points with the stat line behind "
             f"each total."),
            ("How often is this updated?",
             f"Whenever the underlying projections are revised, which is "
             f"typically every few days through the preseason and after "
             f"significant news. This page was last updated "
             f"{built:%B %-d, %Y}."),
        ]

        title = (f"{season} Fantasy Football {label} Projections | "
                 f"LineupBeat")
        # Under 158 characters or Google truncates it, and the truncated
        # half is the half nobody wrote.
        desc = (f"{season} fantasy football {full} projections for {n} "
                f"players. PPR, half PPR and standard points with the "
                f"projected stat line behind every total.")

        schema = {"@type": "Dataset",
                  "name": f"{season} fantasy football {full} projections",
                  "description": desc,
                  "url": f"{seo.SITE_URL}/{SPORT}/projections/{pos.lower()}/",
                  "dateModified": built.strftime("%Y-%m-%d"),
                  "creator": {"@type": "Organization", "name": "LineupBeat"},
                  **seo.dataset_extras(temporal=str(season))}
        crumbs = seo.breadcrumbs([
            ("LineupBeat", "/"), ("Fantasy data", f"/{SPORT}/data/"),
            ("Projections", f"/{SPORT}/projections/"),
            (label, f"/{SPORT}/projections/{pos.lower()}/")])
        itemlist = seo.itemlist_schema(
            f"{season} {full} projections, top 50 by PPR points",
            f"{seo.SITE_URL}/{SPORT}/projections/{pos.lower()}/",
            [(r["rank"], r["name"],
              f"/{SPORT}/{links[key(r['name'])]}/"
              if links.get(key(r["name"])) else None)
             for r in rows[:50]])

        siblings = "".join(
            f'<a class="posnav" href="/{SPORT}/projections/{o.lower()}/">'
            f'{POS_PLURAL[o]}s</a>' for o in others)

        body = f"""<main class="pbwrap">
  <nav class="crumbs" aria-label="Breadcrumb">
    <a href="/">LineupBeat</a><span>/</span>
    <a href="/{SPORT}/data/">Fantasy data</a><span>/</span>
    <a href="/{SPORT}/projections/">Projections</a><span>/</span>
    <b>{esc(label)}</b></nav>

  <div class="pbhead">
    <div>
      <h1>{season} Fantasy Football {esc(label)} Projections</h1>
      <p class="pbsub">Full-season projections for {n} {esc(full)}s, in
        PPR, half PPR and standard scoring, with the stat line behind every
        total.
        <span class="pbdate">Updated {built:%-m/%-d}</span></p>
    </div>
  </div>

  <p class="posprose">{esc(prose)}</p>

  <div class="posnavrow">
    <span class="pbsublab">Other positions</span>{siblings}
    <a class="posnav" href="/{SPORT}/projections/">All positions</a>
  </div>

  {seo.scroll_hint("the full stat line")}
  <div class="xtab" tabindex="0" role="region" aria-label="{esc(label)} projections">
  <table class="pbtbl">
    <thead><tr>
      <th class="l rk c-rk">#</th>
      <th class="l c-nm">Player</th>
      <th class="l tm">Team</th>
      <th class="pt">PPR</th>
      {heads}
      <th class="alt">Half</th>
      <th class="alt">Standard</th>
    </tr></thead>
    <tbody>
{body_rows}
    </tbody>
  </table>
  </div>

  <p class="pbnote">
    Full-season totals, not per game. Ranks are within position. A
    {esc(full)}'s rank differs between scoring formats, which is why all
    three are shown.
  </p>
{seo.faq_html(faq)}{seo.related_html('projections')}
</main>"""

        page = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;500;600&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600&display=swap" rel="stylesheet">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical"
      href="{seo.SITE_URL}/{SPORT}/projections/{pos.lower()}/">
{seo.social_meta(title, desc, f"{seo.SITE_URL}/{SPORT}/projections/{pos.lower()}/")}
<script type="application/ld+json">{seo.graph(
    schema, crumbs, seo.faq_schema(faq), seo.ORGANISATION, itemlist)}</script>
<style>{css}{PAGE_CSS}{seo.CRUMB_CSS}{seo.UI_CSS}{seo.SCROLLTABLE_CSS}{seo.RELATED_CSS}{seo.TEAMS_CSS}{seo.TEAMS_CSS}{seo.BYLINE_CSS}{POS_CSS}</style>
</head>
<body>
{header}
{body}
{footer}
{seo.TRACKING}
</body>
</html>"""

        out = SITE / SPORT / "projections" / pos.lower() / "index.html"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(seo.check_page(page, str(out)))
        written.append((pos, n, len(page)))
        add_to_sitemap(
            f"{seo.SITE_URL}/{SPORT}/projections/{pos.lower()}/")
    return written


POS_CSS = """
/* The paragraph that does the ranking work. A table has the numbers and
   none of the questions somebody typed to find them. */
.posprose{font-size:.95rem; line-height:1.65; color:var(--ink);
  max-width:74ch; margin:1.2rem 0 0}
.posnavrow{display:flex; gap:.4rem; flex-wrap:wrap; align-items:center;
  margin:1.3rem 0 .3rem}
.pbsublab{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.07em; font-size:.66rem; color:var(--quiet);
  margin-right:.3rem}
/* .posnav is styled once, in seo.UI_CSS. This was a second copy that
   came after it in the cascade and quietly won. */
@media (max-width:760px){
  .posnav{min-height:44px; display:inline-flex; align-items:center}
}
"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--db", default="beatwire.db")
    ap.add_argument("--season", type=int, default=2026)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    path = Path(args.workbook).expanduser()
    if not path.exists():
        sys.exit(f"  no file at {path}")

    board = read_sheet(path)
    total = sum(len(v) for v in board.values())
    print(f"\n  {path.name}")
    print(f"  {total} players: "
          + "  ".join(f"{p} {len(board[p])}" for p in POSITIONS
                      if p in board))

    notes = check(board)
    if notes:
        print(f"\n  {len(notes)} thing(s) worth a look:")
        for n in notes[:8]:
            print(f"    {n}")
        print(f"  None of these stops the page being built.")

    conn = None
    db = ROOT / args.db
    if db.exists():
        conn = sqlite3.connect(db)
        conn.row_factory = sqlite3.Row
    links = roster_links(conn)
    matched = sum(1 for rows in board.values() for r in rows
                  if key(r["name"]) in links)
    print(f"\n  {matched} of {total} link to a player page")

    css, header, footer = site_chrome()
    page = build_html(board, links, css, header, footer, args.season,
                      path.name, notes)

    out = Path(args.out) if args.out else SITE / SPORT / "projections" / "index.html"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(seo.check_page(page, str(out)))
    print(f"  wrote {out.relative_to(ROOT)}  ({len(page):,} bytes)")

    if add_to_sitemap(f"https://lineupbeat.com/{SPORT}/projections/"):
        print(f"  added to sitemap.xml")

    # A page per position.
    #
    # "Fantasy football projections" belongs to four national sites.
    # "Fantasy football RB projections 2026" is a real query with far less
    # competition, and it is one this board answers better than they do
    # because the stat line is on the page.
    STATS = STAT_COLUMNS
    stats_json = {
        pos: [(k, lab) for k, lab in STATS.get(pos, [])
              if any(r.get(k) is not None for r in rows)]
        for pos, rows in board.items()}

    from datetime import datetime as _d
    built = eastern_now()
    made = write_position_pages(board, links, css, header, footer,
                                args.season, stats_json, built)
    if made:
        print(f"\n  position pages:")
        for pos, n, size in made:
            print(f"    /{SPORT}/projections/{pos.lower()}/  "
                  f"{n} players, {size:,} bytes")
    if link_from_hub():
        print(f"  linked from the Fantasy Data hub")

    print(f"\n  look at it:")
    print(f"    cd site && python3 -m http.server 8000")
    print(f"    open http://localhost:8000/{SPORT}/projections/")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
