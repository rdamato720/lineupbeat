#!/usr/bin/env python3
"""Build the ADP and draft value page.

    python3 scripts/build_draft_value.py
    python3 scripts/build_draft_value.py --projections data/projections.xlsx

Answers one question: is a player being drafted earlier or later than the
board says he should be?

WHY POSITIONAL RANKS AND NOT OVERALL ADP

Overall ADP 107 and "our QB7" are not comparable quantities. A quarterback
going at pick 107 might be the fourteenth quarterback off the board, and
fourteen against seven is a gap you can reason about. So both sides are
converted to a rank within position first, and the gap is the difference.

    market_position_rank - lineupbeat_position_rank

Positive means the board likes him more than the market does.

WHAT IS DELIBERATELY NOT IN HERE

Durability, coaching, strength of schedule. Each of those is a real signal
and each is on its own page. Folding them in would make the number stop
being checkable: a reader could not work out why a player was labelled
Strong Value, and a value calculation nobody can follow is worth less than
a simple one they can argue with.

THE IMPLIED ADP IS CONTEXT, NOT THE LABEL

Ranking a player QB7 and finding where the market's QB7 actually goes gives
a pick number, and the difference is how many picks of discount the board
implies. That is useful and it is second: the public label comes from the
positional gap, which is the simpler number.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))
import seo
import seo_faqs

ROOT = Path(__file__).resolve().parent.parent


def eastern_now():
    """The date a reader in the league's own time zone would call today.

    UTC rolls over at 8pm Eastern, so a page built at 9pm on the 9th was
    stamped the 10th and looked a day ahead of the ADP it was showing.
    Football runs on Eastern; so should the date on the page.
    """
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/New_York"))
    except Exception:
        # No tz database. Approximate rather than fall back to UTC, which
        # is the failure this exists to avoid.
        return datetime.now(timezone.utc) - timedelta(hours=4)
SITE = ROOT / "site"
SPORT = "nfl"
POSITIONS = ["QB", "RB", "WR", "TE"]

# Thresholds live here, not scattered through the markup. Moving a boundary
# should be one edit, and the labels on the page should never be able to
# disagree with the rule that produced them.
STRONG_VALUE_MIN = 5
VALUE_MIN = 2
FAIR_MIN = -1
FAIR_MAX = 1
PRICEY_MAX = -2
OVERPRICED_MAX = -5

TEAMS_PER_ROUND = 12

SIGNALS = ["Strong Value", "Value", "Fair Price", "Pricey", "Overpriced"]


def signal_for(gap):
    """One place the five labels are decided."""
    if gap is None:
        return None
    if gap >= STRONG_VALUE_MIN:
        return "Strong Value"
    if gap >= VALUE_MIN:
        return "Value"
    if FAIR_MIN <= gap <= FAIR_MAX:
        return "Fair Price"
    if gap > OVERPRICED_MAX:
        return "Pricey"
    return "Overpriced"


def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)


def key(n):
    n = re.sub(r"[.'`’]", "", (n or "").lower())
    return " ".join(re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", n).split())


def slug(s):
    s = re.sub(r"[^\w\s-]", "", (s or "").lower())
    return re.sub(r"[\s_]+", "-", s).strip("-")


def read_projections(path: Path):
    """The board, by scoring format. One row per player per format."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        pos = sheet.split()[0].upper()
        if pos not in POSITIONS:
            continue
        ws = wb[sheet]
        head = [str(c.value or "").strip().lower() for c in ws[1]]

        def col(*names):
            for n in names:
                if n in head:
                    return head.index(n)
            return None

        ci = {"player": col("player", "name"), "team": col("team", "tm"),
              "ppr": col("ppr", "ppr points"),
              "half": col("half ppr", "half"),
              "std": col("non-ppr", "standard", "std")}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if ci["player"] is None or ci["player"] >= len(row):
                continue
            nm = row[ci["player"]]
            if not nm:
                continue

            def val(k):
                i = ci.get(k)
                if i is None or i >= len(row) or row[i] is None:
                    return None
                try:
                    return float(row[i])
                except (TypeError, ValueError):
                    return None

            ppr = val("ppr")
            if ppr is None:
                continue
            out.append({
                "name": str(nm).strip(), "pos": pos,
                "team": (str(row[ci["team"]]).strip().upper()
                         if ci["team"] is not None and ci["team"] < len(row)
                         and row[ci["team"]] else ""),
                "ppr": ppr, "half": val("half") if val("half") is not None else ppr,
                "std": val("std") if val("std") is not None else ppr,
            })
    return out


# Which scoring formats have a compatible ADP dataset.
#
# The roster carries one ADP column and it is PPR, so comparing a standard
# projection against it would put standard ranks on one side and PPR market
# prices on the other -- a comparison that looks fine and means nothing.
#
# Adding "adp_half" or "adp_std" columns to the roster is all that is
# needed to switch the other two on; the page reads this map rather than
# assuming.
ADP_COLUMNS = {"ppr": "adp", "half": "adp_half", "std": "adp_std"}
FORMAT_LABEL = {"ppr": "PPR", "half": "Half PPR", "std": "Standard"}


def available_formats(sample_row):
    """Formats where we hold ADP for that same scoring rule."""
    out = []
    for fmt, col in ADP_COLUMNS.items():
        v = (sample_row.get(col) or "").strip()
        try:
            if float(v) > 0:
                out.append(fmt)
        except (TypeError, ValueError):
            continue
    return out or ["ppr"]


def read_adp():
    """The canonical ADP, the same file the durability board reads.

    Not a second scraper. If Justin Herbert is 107.0 on durability he is
    107.0 here, because it is one number read twice rather than two numbers
    that happen to agree today.
    """
    rp = ROOT / "rosters" / f"{SPORT}.csv"
    if not rp.exists():
        sys.exit(f"  no {rp}")
    adp, meta = {}, {}
    formats = ["ppr"]
    for r in csv.DictReader(rp.open()):
        if formats == ["ppr"]:
            formats = available_formats(r)
        v = (r.get("adp") or "").strip()
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        if v <= 0:
            continue
        k = key(r.get("name", ""))
        adp[k] = {"adp": v, "id": r.get("id"), "slug": slug(r.get("name", "")),
                  "team": (r.get("team") or "").upper(),
                  "pos": (r.get("position") or "").upper()}
    mp = ROOT / "rosters" / "adp_meta.json"
    if mp.exists():
        try:
            meta = json.loads(mp.read_text())
        except ValueError:
            meta = {}
    return adp, meta, formats


def build_board(proj, adp, fmt):
    """One scoring format, ranked both ways, gap computed.

    Ranks are recomputed per format because they genuinely differ: a pass
    catching back can be RB8 in PPR and RB16 in standard, and his draft
    value has to move with it or the page is answering the wrong question
    for two of the three formats.
    """
    rows = []
    for p in proj:
        a = adp.get(key(p["name"]))
        rows.append({**p, "pts": p[fmt],
                     "adp": a["adp"] if a else None,
                     "slug": a["slug"] if a else None,
                     "team": p["team"] or (a["team"] if a else "")})

    # Our rank within position, on the format's own points.
    by_pos = {}
    for r in rows:
        by_pos.setdefault(r["pos"], []).append(r)
    for pos, group in by_pos.items():
        group.sort(key=lambda x: (-x["pts"], x["name"]))
        for i, r in enumerate(group, 1):
            r["lb_rank"] = i

    # The market's rank within position, from overall ADP. A player with no
    # ADP is not ranked at all rather than ranked last: "not currently
    # drafted" is a different statement from "drafted 200th".
    for pos, group in by_pos.items():
        drafted = sorted([r for r in group if r["adp"] is not None],
                         key=lambda x: x["adp"])
        for i, r in enumerate(drafted, 1):
            r["mkt_rank"] = i
        # Where the market has its Nth player at this position, for the
        # implied-ADP lookup below.
        market_at = {r["mkt_rank"]: r["adp"] for r in drafted}
        for r in group:
            r.setdefault("mkt_rank", None)
            if r["mkt_rank"] is None:
                r["gap"] = None
                r["signal"] = None
                r["implied"] = None
                r["pick_value"] = None
                r["round_value"] = None
                continue
            r["gap"] = r["mkt_rank"] - r["lb_rank"]
            r["signal"] = signal_for(r["gap"])
            # Where the market drafts the player we rank him equal to.
            r["implied"] = market_at.get(r["lb_rank"])
            if r["implied"] is not None:
                r["pick_value"] = r["adp"] - r["implied"]
                r["round_value"] = r["pick_value"] / TEAMS_PER_ROUND
            else:
                r["pick_value"] = None
                r["round_value"] = None

    return rows


def explain(r):
    """The expanded row, in the language a person would use.

    Analytical, not promotional. "League winner" and "steal of the century"
    are the sort of phrase that makes a number look like a sales pitch, and
    the number is doing the work here.
    """
    if r["gap"] is None:
        return ("Not currently being drafted in the sample, so there is no "
                "market price to compare our projection against.")
    pos = r["pos"]
    if r["gap"] > 0:
        lead = (f"LineupBeat ranks him {r['gap']} "
                f"spot{'s' if abs(r['gap']) != 1 else ''} higher at his "
                f"position than the current market.")
    elif r["gap"] < 0:
        lead = (f"The market ranks him {abs(r['gap'])} "
                f"spot{'s' if abs(r['gap']) != 1 else ''} higher at his "
                f"position than our projection does.")
    else:
        lead = ("Our projection and the market put him in the same place at "
                "his position.")

    if r["pick_value"] is None:
        return lead
    rounds = abs(r["round_value"])
    if abs(r["pick_value"]) < 6:
        return lead + " His draft price is close to what that ranking implies."
    when = "later" if r["pick_value"] > 0 else "earlier"
    approx = ("about half a round" if rounds < 0.75
              else f"roughly {rounds:.0f} round{'s' if rounds >= 1.5 else ''}")
    return (lead + f" Based on where {pos}s with that ranking are being "
            f"selected, he is going {approx} {when} than our projection "
            f"would imply.")


PAGE_CSS = """
.topbar .logo,.topbar .vbtn{text-decoration:none}
.topbar .vbtn:hover{text-decoration:none; color:var(--ink)}
.vbtn[aria-current="page"]{color:#0A0C08; background:var(--signal);
  border-color:var(--signal)}

/* ---- draft value ----
   A draft board first: ADP ascending, so it reads the way the room reads.
   The value column is what it adds. */
.dvwrap{max-width:1080px; margin:0 auto; padding:0 1rem 4rem}
.dvhead{margin:1.6rem 0 .4rem}
.dvhead h1{font-size:1.7rem; margin:0; letter-spacing:-.01em;
  font-family:var(--text)}
.dvsub{color:var(--quiet); font-size:.86rem; margin:.4rem 0 0; max-width:72ch;
  line-height:1.55}
.dvdate{display:inline-block; margin-left:.4rem; font-family:var(--agate);
  text-transform:uppercase; letter-spacing:.06em; font-size:.7rem;
  color:var(--signal); border:1px solid var(--rule); border-radius:999px;
  padding:.1rem .5rem; vertical-align:.05em}
.dvcards{display:grid; grid-template-columns:repeat(3, 1fr); gap:.7rem;
  margin:1.2rem 0 0}
/* Four terms, because ADP and Market were reading as the same thing: one
   is a pick number and the other is a rank derived from it. */
.dv4{grid-template-columns:repeat(4, 1fr)}
@media (max-width:900px){ .dv4{grid-template-columns:repeat(2, 1fr)} }
@media (max-width:820px){ .dvcards{grid-template-columns:1fr} }
@media (max-width:560px){ .dv4{grid-template-columns:1fr} }
.dvcard{background:var(--card); border:1px solid var(--rule);
  border-radius:8px; padding:.75rem .9rem}
.dvcard p{margin:.3rem 0 0; font-size:.84rem; line-height:1.45;
  color:var(--quiet)}
.dvcard p b{color:var(--ink)}
.dvk{font-family:var(--agate); text-transform:uppercase; letter-spacing:.07em;
  font-size:.68rem; color:var(--signal)}
.dvwhen{color:var(--quiet); font-size:.76rem; margin:.9rem 0 0}
.dvwhen b{color:var(--ink)}

.dvrow{display:flex; gap:.3rem; flex-wrap:wrap; align-items:center;
  margin:.6rem 0 0}
.dvlab{font-family:var(--agate); text-transform:uppercase; letter-spacing:.07em;
  font-size:.66rem; color:var(--quiet); margin-right:.3rem}
.dvtab{font-family:var(--agate); text-transform:uppercase;
  background:transparent; border:1px solid var(--rule); color:var(--quiet);
  font-size:.76rem; padding:.3rem .7rem; border-radius:999px; cursor:pointer;
  letter-spacing:.04em}
.dvtab:hover{color:var(--ink); border-color:var(--ink)}
.dvtab[aria-pressed="true"]{background:var(--signal); border-color:var(--signal);
  color:#0b0f0a; font-weight:600}
.dvsearch input{background:var(--card); border:1px solid var(--rule);
  color:var(--ink); font:inherit; font-size:.82rem; padding:.34rem .7rem;
  border-radius:6px; min-width:180px}

.dvtbl{width:100%; border-collapse:collapse; font-size:.88rem;
  font-variant-numeric:tabular-nums; margin-top:1.2rem}
.dvtbl th{text-align:right; font-family:var(--agate); font-size:.7rem;
  letter-spacing:.08em; text-transform:uppercase; color:var(--quiet);
  font-weight:600; padding:.5rem .5rem; border-bottom:1px solid var(--rule)}
.dvtbl th.l,.dvtbl td.l{text-align:left}
.dvtbl td{padding:.42rem .5rem; border-bottom:1px solid var(--rule);
  text-align:right}
.dvtbl tbody tr.r:hover{background:var(--card); cursor:pointer}
.dvadp{font-family:var(--data); color:var(--quiet); width:4.4rem}
.dvnm{font-weight:500}
.dvnm a{color:var(--ink); text-decoration:none}
.dvnm a:hover{color:var(--signal)}
.dvpos{font-family:var(--data); font-size:.74rem; color:var(--quiet);
  width:3rem}
.dvrk{font-family:var(--data); font-size:.8rem; width:4.2rem}
.dvgap{font-family:var(--data); font-weight:600; width:4rem}
.dvpts{font-family:var(--data); width:5rem}
.g-up{color:#8BE04E} .g-dn{color:#FF6B4A}
.sig{font-family:var(--agate); text-transform:uppercase; font-size:.62rem;
  letter-spacing:.06em; border:1px solid var(--rule); border-radius:999px;
  padding:.1rem .45rem; white-space:nowrap; color:var(--quiet)}
.sig.v-strong-value{color:#8BE04E; border-color:#8BE04E}
.sig.v-value{color:#B9DE7E; border-color:#B9DE7E}
.sig.v-pricey{color:#E09478; border-color:#E09478}
.sig.v-overpriced{color:#FF6B4A; border-color:#FF6B4A}
.sig.v-none{opacity:.6}

.dvx td{background:var(--card); border-bottom:1px solid var(--rule)}
.dvxin{padding:.5rem .3rem .8rem; text-align:left}
.dvxgrid{display:flex; gap:1.4rem; flex-wrap:wrap; margin-bottom:.6rem}
.dvxg span{display:block; font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.05em; font-size:.6rem; color:var(--quiet)}
.dvxg b{font-family:var(--data); font-size:.94rem; color:var(--signal)}
.dvxwhy{font-size:.86rem; line-height:1.55; color:var(--ink); max-width:74ch;
  margin:0}
.dvempty{color:var(--quiet); padding:1.4rem .5rem; font-size:.86rem}
.dvonly{font-size:.82rem; line-height:1.55; color:var(--quiet);
  margin:.7rem 0 0; max-width:74ch}
.dvonly b{color:var(--ink)}
.dvonly a{color:var(--quiet); text-decoration:underline}
.dvonly a:hover{color:var(--signal)}
/* Definitions below the table, not above it.
   Four cards, a scoring note and two methodology paragraphs stood between
   the headline and the data. Somebody arriving here wants the board; the
   explanation is for when a number surprises him. */
.dvterms{display:grid; grid-template-columns:repeat(2, 1fr); gap:.6rem 1.4rem;
  margin:0 0 1.4rem; max-width:74ch}
@media (max-width:640px){ .dvterms{grid-template-columns:1fr} }
.dvterms dt{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.06em; font-size:.68rem; color:var(--ink)}
.dvterms dd{margin:.1rem 0 0; font-size:.82rem; line-height:1.5;
  color:var(--quiet)}
.dvterms dd b{color:var(--signal)}
.dvfoot p b{color:var(--ink)}
.dvcount{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.06em; font-size:.68rem; color:var(--quiet); margin:1rem 0 0}
.dvfoot{color:var(--quiet); font-size:.78rem; margin:2rem 0 0; max-width:74ch;
  line-height:1.55}
.dvfoot h2{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.07em; font-size:.78rem; color:var(--quiet);
  margin:0 0 .5rem}

@media (max-width:760px){
  .dvtab{min-height:44px; display:inline-flex; align-items:center}
  .dvsearch{width:100%}
  .dvsearch input{width:100%; font-size:16px; min-height:44px}
  .dvtbl{font-size:.8rem}
  .dvtbl th,.dvtbl td{padding:.4rem .25rem}
  .dvpts,.dvteam{display:none}
}
"""


def site_chrome():
    tpl = SITE / "template.html"
    if not tpl.exists():
        return "", "", ""
    src = tpl.read_text()
    css = re.search(r"<style>(.*?)</style>", src, re.S)
    foot = re.search(r"<footer.*?</footer>", src, re.S)
    header = seo.site_nav("data")
    return (css.group(1) if css else ""), header, (foot.group(0) if foot else "")


def static_rows(rows):
    """The default view, written into the HTML.

    A crawler that does not run JavaScript saw an empty tbody and the
    "no player matches that" line, which is the opposite of what the page
    is for: the whole point is 600 rows of proprietary comparison and none
    of it was in the source.
    So the ADP-ascending view is rendered at build time. JavaScript still
    owns filtering, search, sorting and the expanded rows -- it replaces
    this table rather than creating it.
    """
    DASH = "\u2014"
    out = []
    for r in rows:
        gap = r["gap"]
        gcls = "" if gap is None else ("g-up" if gap > 0
                                       else "g-dn" if gap < 0 else "")
        gtxt = DASH if gap is None else (f"+{gap}" if gap > 0 else str(gap))
        sig = r["signal"] or "No ADP"
        scls = (("v-" + sig.lower().replace(" ", "-")) if r["signal"]
                else "v-none")
        name = (f'<a href="/{SPORT}/{r["slug"]}/">{esc(r["name"])}</a>'
                if r.get("slug") else esc(r["name"]))
        adp = DASH if r["adp"] is None else f'{r["adp"]:.1f}'
        mkt = DASH if r["mkt_rank"] is None else f'{r["pos"]}{r["mkt_rank"]}'
        out.append(
            "<tr class=\"r\">"
            f'<td class="l dvadp">{adp}</td>'
            f'<td class="l dvnm">{name}</td>'
            f'<td class="l dvpos">{esc(r["pos"])}</td>'
            f'<td class="l dvteam dvpos">{esc(r["team"])}</td>'
            f'<td class="dvrk">{esc(mkt)}</td>'
            f'<td class="dvrk">{esc(r["pos"])}{r["lb_rank"]}</td>'
            f'<td class="dvgap {gcls}">{gtxt}</td>'
            f'<td class="dvpts">{r["pts"]:.1f}</td>'
            f'<td class="l"><span class="sig {scls}">{esc(sig)}</span></td>'
            "</tr>")
    return "\n".join(out)


def build_html(boards, meta, css, header, footer, formats):
    built = eastern_now()

    # One payload, three formats. The gap has to be recomputed per format
    # and doing it here rather than in the browser keeps the arithmetic in
    # one place where it can be checked.
    payload = {}
    for fmt, rows in boards.items():
        payload[fmt] = [
            {"a": r["adp"], "n": r["name"], "p": r["pos"], "t": r["team"],
             "m": r["mkt_rank"], "l": r["lb_rank"], "g": r["gap"],
             "s": r["signal"], "pt": round(r["pts"], 1),
             "im": round(r["implied"], 1) if r["implied"] is not None else None,
             "pv": round(r["pick_value"], 1) if r["pick_value"] is not None else None,
             "rv": round(r["round_value"], 1) if r["round_value"] is not None else None,
             "id": r["slug"], "w": explain(r)}
            for r in sorted(rows, key=lambda x: (x["adp"] is None,
                                                 x["adp"] or 0, x["name"]))
        ]

    # The default view, in the HTML. ADP ascending, PPR, no filter -- the
    # same thing draw() produces on load, so the page does not flicker into
    # a different table when the script runs.
    static = static_rows(sorted(
        boards["ppr"],
        key=lambda x: (x["adp"] is None, x["adp"] or 0, x["name"])))

    # Only formats where both sides use the same scoring rule.
    #
    # Offering a Standard button against a PPR ADP would compare standard
    # ranks to PPR market prices: a comparison that renders cleanly and
    # means nothing. Where only one format is available the row says so
    # rather than quietly disappearing.
    if len(formats) > 1:
        scoring_row = (
            '<div class="dvrow"><span class="dvlab">Scoring</span>'
            + "".join(
                f'<button class="dvtab" data-fmt="{f}" '
                f'aria-pressed="{"true" if f == formats[0] else "false"}">'
                f'{FORMAT_LABEL[f]}</button>' for f in formats)
            + "</div>")
    else:
        scoring_row = (
            f'<p class="dvonly"><b>{FORMAT_LABEL[formats[0]]} scoring.</b> '
            f'Our ADP sample is {FORMAT_LABEL[formats[0]]}, and a ranking is '
            f'only comparable to a market price drawn under the same rules. '
            f'All three formats are on the '
            f'<a href="/{SPORT}/projections/">projections page</a>.</p>')

    n = len([x for x in boards["ppr"] if x["adp"] is not None])
    when = ""
    if meta.get("end"):
        def short(s):
            y, m_, d_ = s.split("-")
            return f"{int(m_)}/{int(d_)}"
        shape = f", {meta['teams']} teams" if meta.get("teams") else ""
        # Two dates, because they are two different facts. "ADP through" is
        # how current the market data is; "updated" is when this page was
        # last built. Showing one number for both invites the reader to
        # assume the market moved when only the build did.
        def longform(s):
            y, m_, d_ = (int(x) for x in s.split("-"))
            return datetime(y, m_, d_).strftime("%B %-d, %Y")
        # The end date once, not twice. "ADP through August 7" and
        # "7/31 - 8/7" were saying the same thing in the same sentence.
        when = (f'<p class="dvwhen">'
                f'<b>ADP through {longform(meta["end"])}</b> '
                f'&middot; {meta.get("drafts", 0):,} drafts since '
                f'{short(meta["start"])}{shape} &middot; PPR</p>')

    body = f"""<main class="dvwrap">
  <nav class="crumbs" aria-label="Breadcrumb">
    <a href="/">LineupBeat</a><span>/</span>
    <a href="/{SPORT}/data/">Fantasy data</a><span>/</span>
    <b>ADP &amp; draft value</b></nav>

  <div class="dvhead">
    <h1>2026 Fantasy Football ADP &amp; Draft Value</h1>
    <p class="dvsub">Where the market is drafting every player compared
      with our projections, and where the biggest gaps are.</p>
    {when}
  </div>

{seo.byline_html(built, data_through=(longform(meta["end"]) if meta.get("end") else None))}
  <div class="dvrow">
    <span class="dvlab">Position</span>
    <button class="dvtab" data-pos="ALL" aria-pressed="true">All</button>
    {''.join(f'<button class="dvtab" data-pos="{p}" aria-pressed="false">{p}</button>'
             for p in POSITIONS)}
    <span class="dvsearch"><input id="dvq" type="search"
      placeholder="Find a player" autocomplete="off"
      aria-label="Find a player"></span>
  </div>

  <div class="dvrow">
    <span class="dvlab">Value</span>
    <button class="dvtab" data-val="ALL" aria-pressed="true">All</button>
    <button class="dvtab" data-val="BEST" aria-pressed="false">Best values</button>
    {''.join(f'<button class="dvtab" data-val="{s}" aria-pressed="false">{s}</button>'
             for s in SIGNALS)}
  </div>

  <div class="dvrow">
    <span class="dvlab">Sort</span>
    <button class="dvtab" data-sort="adp" aria-pressed="true">ADP</button>
    <button class="dvtab" data-sort="gap" aria-pressed="false">Best value</button>
    <button class="dvtab" data-sort="gapdn" aria-pressed="false">Most overpriced</button>
    <button class="dvtab" data-sort="lb" aria-pressed="false">LineupBeat rank</button>
    <button class="dvtab" data-sort="pts" aria-pressed="false">Projected points</button>
  </div>

  <p class="dvcount" id="dvcount"></p>

  <table class="dvtbl">
    <thead><tr>
      <th class="l dvadp">ADP</th>
      <th class="l">Player</th>
      <th class="l dvpos">Pos</th>
      <th class="l dvteam dvpos">Team</th>
      <th class="dvrk">Market</th>
      <th class="dvrk">LineupBeat</th>
      <th class="dvgap">Gap</th>
      <th class="dvpts">Pts</th>
      <th class="l">Draft value</th>
    </tr></thead>
    <tbody id="dvbody">
{static}
    </tbody>
  </table>
  <p class="dvempty" id="dvempty" hidden>No player matches that filter.</p>

  <section class="dvfoot">
    <h2>What the columns mean</h2>
    <dl class="dvterms">
      <div><dt>ADP</dt><dd>Average overall draft pick. One number for the
        whole draft, like 72.4.</dd></div>
      <div><dt>Market</dt><dd>That ADP as a rank within the position. If
        72.4 is the seventh quarterback taken, Market is QB7.</dd></div>
      <div><dt>LineupBeat</dt><dd>Our projected rank at that position, like
        QB4.</dd></div>
      <div><dt>Gap</dt><dd>Market minus LineupBeat. QB7 against QB4 is
        <b>+3</b>: we rank him three spots higher than the market prices
        him.</dd></div>
    </dl>

    <h2>How draft value works</h2>
    <p>Draft Value compares where the market is selecting a player at his
       position with where LineupBeat's full season projection ranks him. A
       positive gap means LineupBeat ranks the player higher than the
       market. A negative gap means the market is paying a premium relative
       to our projection.</p>
    <p>Draft Value is not an instruction to draft or avoid a player at any
       cost. It identifies differences between LineupBeat's projections and
       current market prices. ADP changes, so player value can change with
       it.</p>
    <p><b>{FORMAT_LABEL[formats[0]]} scoring only.</b> Our ADP sample is
       {FORMAT_LABEL[formats[0]]}, and a ranking is only comparable to a
       market price drawn under the same rules. All three formats are on
       the <a href="/{SPORT}/projections/">projections page</a>.</p>
    <p>Durability, coaching and strength of schedule are deliberately not
       part of this calculation. Each has its own page, and folding them in
       would make this number impossible to check.</p>
  </section>
{seo.faq_html(seo_faqs.DRAFT_VALUE)}{seo.related_html('draft-value')}
</main>

<script>
const DV = {json.dumps(payload, separators=(',', ':'))};
let fmt = "ppr", pos = "ALL", val = "ALL", sort = "adp", open = null;
const body = document.getElementById("dvbody");
const empty = document.getElementById("dvempty");
const count = document.getElementById("dvcount");
const q = document.getElementById("dvq");

function sigClass(s){{
  return s ? "v-" + s.toLowerCase().replace(/ /g, "-") : "v-none";
}}
function plus(n){{
  return n === null ? "\\u2014" : (n > 0 ? "+" + n : String(n));
}}

function rows(){{
  const term = (q.value || "").trim().toLowerCase();
  let out = DV[fmt].filter(r => {{
    if(pos !== "ALL" && r.p !== pos) return false;
    if(val === "BEST" && !(r.s === "Strong Value" || r.s === "Value"))
      return false;
    if(val !== "ALL" && val !== "BEST" && r.s !== val) return false;
    if(term && !r.n.toLowerCase().includes(term)
            && (r.t || "").toLowerCase() !== term) return false;
    return true;
  }});
  // Undrafted players sort last whatever the key, because "no ADP" is not a
  // position on the board.
  const nul = out.filter(r => r.g === null);
  const has = out.filter(r => r.g !== null);
  const by = {{
    adp: (a, b) => a.a - b.a,
    gap: (a, b) => b.g - a.g || a.a - b.a,
    gapdn: (a, b) => a.g - b.g || a.a - b.a,
    lb: (a, b) => a.p.localeCompare(b.p) || a.l - b.l,
    pts: (a, b) => b.pt - a.pt,
  }}[sort];
  has.sort(by);
  return has.concat(nul);
}}

function draw(){{
  const list = rows();
  body.innerHTML = list.map((r, i) => {{
    const name = r.id
      ? `<a href="/{SPORT}/${{r.id}}/">${{r.n}}</a>` : r.n;
    const gcls = r.g === null ? "" : r.g > 0 ? "g-up" : r.g < 0 ? "g-dn" : "";
    return `<tr class="r" data-i="${{i}}">
      <td class="l dvadp">${{r.a === null ? "\\u2014" : r.a.toFixed(1)}}</td>
      <td class="l dvnm">${{name}}</td>
      <td class="l dvpos">${{r.p}}</td>
      <td class="l dvteam dvpos">${{r.t || ""}}</td>
      <td class="dvrk">${{r.m === null ? "\\u2014" : r.p + r.m}}</td>
      <td class="dvrk">${{r.p}}${{r.l}}</td>
      <td class="dvgap ${{gcls}}">${{plus(r.g)}}</td>
      <td class="dvpts">${{r.pt.toFixed(1)}}</td>
      <td class="l"><span class="sig ${{sigClass(r.s)}}">${{
        r.s || "No ADP"}}</span></td>
    </tr>`;
  }}).join("");
  empty.hidden = list.length > 0;
  count.textContent = `${{list.length}} player${{list.length === 1 ? "" : "s"}}`;

  body.querySelectorAll("tr.r").forEach(tr =>
    tr.addEventListener("click", () => expand(tr, list[+tr.dataset.i])));
}}

function expand(tr, r){{
  const next = tr.nextElementSibling;
  if(next && next.classList.contains("dvx")){{ next.remove(); return; }}
  body.querySelectorAll(".dvx").forEach(x => x.remove());
  const cells = [
    ["Current ADP", r.a === null ? "\\u2014" : r.a.toFixed(1)],
    ["Market", r.m === null ? "\\u2014" : r.p + r.m],
    ["LineupBeat", r.p + r.l],
    ["Projected pts", r.pt.toFixed(1)],
    ["Implied ADP", r.im === null ? "\\u2014" : r.im.toFixed(1)],
    ["Draft pick value", plus(r.pv)],
    ["Round value", r.rv === null ? "\\u2014" : plus(r.rv)],
  ].map(([k, v]) =>
    `<div class="dvxg"><span>${{k}}</span><b>${{v}}</b></div>`).join("");
  const row = document.createElement("tr");
  row.className = "dvx";
  row.innerHTML = `<td colspan="9"><div class="dvxin">
    <div class="dvxgrid">${{cells}}</div>
    <p class="dvxwhy">${{r.w}}</p></div></td>`;
  tr.after(row);
}}

function wire(attr, set){{
  document.querySelectorAll(`[data-${{attr}}]`).forEach(b =>
    b.addEventListener("click", () => {{
      set(b.dataset[attr]);
      document.querySelectorAll(`[data-${{attr}}]`).forEach(x =>
        x.setAttribute("aria-pressed", x === b ? "true" : "false"));
      draw();
    }}));
}}
wire("pos", v => pos = v);
wire("fmt", v => fmt = v);
wire("val", v => val = v);
wire("sort", v => sort = v);
q.addEventListener("input", draw);
draw();
</script>"""
    return body, built


def add_to_sitemap(url):
    sm = SITE / "sitemap.xml"
    if not sm.exists():
        return False
    text = sm.read_text()
    if url in text:
        return False
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sm.write_text(text.replace(
        "</urlset>",
        f"  <url><loc>{url}</loc><lastmod>{today}</lastmod>"
        f"<changefreq>daily</changefreq><priority>0.8</priority></url>\n"
        "</urlset>"))
    return True


def validate(boards):
    """The checks from the spec, run before publishing rather than trusted.

    Every one of these is a rule the page claims to follow, and a rule
    nobody checks is a rule that quietly stops being true.
    """
    bad = []
    for fmt, rows in boards.items():
        for r in rows:
            g, s = r["gap"], r["signal"]
            if g is None:
                if s is not None:
                    bad.append(f"{r['name']} has a label with no ADP")
                continue
            if s != signal_for(g):
                bad.append(f"{r['name']} {fmt}: {s} for a gap of {g}")
            if r["mkt_rank"] is not None and g != r["mkt_rank"] - r["lb_rank"]:
                bad.append(f"{r['name']} {fmt}: gap does not equal the "
                           f"difference in ranks")
        # Ranks must be within position, never across.
        for pos in POSITIONS:
            group = [r for r in rows if r["pos"] == pos]
            ranks = sorted(r["lb_rank"] for r in group)
            if ranks and ranks != list(range(1, len(group) + 1)):
                bad.append(f"{fmt} {pos}: our ranks are not 1..{len(group)}")
    return bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--projections", default="data/projections.xlsx")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    wb = ROOT / args.projections
    if not wb.exists():
        sys.exit(f"  no {args.projections}")
    proj = read_projections(wb)
    adp, meta, adp_formats = read_adp()
    if not proj:
        sys.exit("  no projections read")

    boards = {f: build_board(proj, adp, f) for f in ("ppr", "half", "std")}
    formats = adp_formats
    if len(formats) == 1:
        print(f"  ADP is {FORMAT_LABEL[formats[0]]} only, so the page offers "
              f"that format alone")

    drafted = [r for r in boards["ppr"] if r["adp"] is not None]
    print(f"\n  {len(proj)} projected players, {len(drafted)} with an ADP")
    if not drafted:
        print(f"  No ADP in rosters/{SPORT}.csv, so every row will read "
              f"\"No ADP\".\n  Run scripts/import_adp.py first.")

    bad = validate(boards)
    if bad:
        print(f"\n  {len(bad)} validation problem(s):")
        for x in bad[:8]:
            print(f"    {x}")
        sys.exit("\n  Not publishing. These are the page's own rules.")

    from collections import Counter
    c = Counter(r["signal"] for r in boards["ppr"] if r["signal"])
    if c:
        print("  " + "  ".join(f"{c.get(s, 0)} {s}" for s in SIGNALS))

    css, header, footer = site_chrome()
    body, built = build_html(boards, meta, css, header, footer, formats)

    title = ("2026 Fantasy Football ADP & Draft Values | LineupBeat")
    desc = ("Compare 2026 fantasy football ADP with LineupBeat projections. "
            "Find draft values, overpriced players and the biggest QB, RB, "
            "WR and TE ranking gaps.")
    schema = {"@context": "https://schema.org", "@type": "Dataset",
              "name": "2026 Fantasy Football ADP and Draft Value",
              "description": desc,
              "url": f"https://lineupbeat.com/{SPORT}/draft-value/",
              "dateModified": built.strftime("%Y-%m-%d"),
              "creator": {"@type": "Organization", "name": "LineupBeat"},
              **seo.dataset_extras(temporal="2026")}
    crumbs = {"@context": "https://schema.org", "@type": "BreadcrumbList",
              "itemListElement": [
                  {"@type": "ListItem", "position": 1, "name": "LineupBeat",
                   "item": "https://lineupbeat.com/"},
                  {"@type": "ListItem", "position": 2, "name": "Fantasy data",
                   "item": f"https://lineupbeat.com/{SPORT}/data/"},
                  {"@type": "ListItem", "position": 3,
                   "name": "ADP and draft value",
                   "item": f"https://lineupbeat.com/{SPORT}/draft-value/"}]}

    # One graph rather than loose blocks: these are facets of one page,
    # and saying so lets a crawler connect the dataset to the site that
    # publishes it and to the questions it answers.
    # The biggest gaps, which is the answer the page exists to give.
    _best = sorted([r for r in boards["ppr"] if r["gap"] is not None],
                   key=lambda x: -x["gap"])[:25]
    globals()["_itemlist"] = seo.itemlist_schema(
        "2026 fantasy football draft values, biggest gaps between our "
        "projection and ADP",
        f"https://lineupbeat.com/{SPORT}/draft-value/",
        [(i, f"{r['name']} ({r['pos']}, {r['signal']})",
          f"/{SPORT}/{r['slug']}/" if r.get("slug") else None)
         for i, r in enumerate(_best, 1)])

    ldjson = seo.graph(
        {k: v for k, v in schema.items() if k != "@context"},
        {k: v for k, v in crumbs.items() if k != "@context"},
        seo.faq_schema(seo_faqs.DRAFT_VALUE),
        seo.ORGANISATION,
        globals().get("_itemlist"))

    page = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;500;600&family=Source+Serif+4:opsz,wght@8..60,400;8..60,600&display=swap" rel="stylesheet">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical" href="https://lineupbeat.com/{SPORT}/draft-value/">
<meta property="og:title" content="{esc(title)}">
<meta property="og:description" content="{esc(desc)}">
<meta property="og:url" content="https://lineupbeat.com/{SPORT}/draft-value/">
<meta property="og:type" content="website">
<script type="application/ld+json">{ldjson}</script>
<style>{css}{PAGE_CSS}{seo.RELATED_CSS}{seo.TEAMS_CSS}{seo.BYLINE_CSS}</style>
</head>
<body>
{header}
{body}
{footer}
{seo.TEAMS_JS}{seo.TRACKING}{seo.VIEW_CONTENT}
</body>
</html>"""

    out = (Path(args.out) if args.out
           else SITE / SPORT / "draft-value" / "index.html")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(seo.check_page(page, str(out)))
    print(f"\n  wrote {out.relative_to(ROOT)}  ({len(page):,} bytes)")
    if add_to_sitemap(f"https://lineupbeat.com/{SPORT}/draft-value/"):
        print(f"  added to sitemap.xml")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
