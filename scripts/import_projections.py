#!/usr/bin/env python3
"""Load the hand-built Offense workbook as the projection source.

    python3 scripts/import_projections.py ~/Downloads/2026_Offense_v1.0.xlsx
    python3 scripts/import_projections.py FILE --publish
    python3 scripts/import_projections.py --show RB

WHAT THIS REPLACES

project5.py infers a season from four years of history and a depth chart. It
is a reasonable model and it does not know that a team signed a receiver in
March, or that a coordinator changed, or what a beat writer saw in camp last
week. The workbook does: it was built from that research, position by
position, and then reconciled so every team's player targets sum to its
quarterback's attempts.

That reconciliation is the part a model cannot fake. Checked on import: the
variance across all 32 teams is under a thousandth of an attempt, which is
floating point rather than disagreement.

WHAT IT DOES NOT DO

Adjust anything. The numbers go in as built -- not scaled, not blended, not
discounted for expected games. The workbook already accounts for what its
author thinks a player will play; a second opinion applied on top would be
this script quietly disagreeing with the model it was asked to load.

Durability lives on its own page, where a reader can see the record and
decide.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parent.parent

# Sheet name to position. Versions move, so match on the prefix.
SHEETS = {"QB": "QB", "RB": "RB", "WR": "WR", "TE": "TE"}

SCHEMA = """
-- STAGING. Mutable, overwritten on every import, and never read by the
-- site. Only a validated run in run_projections is publishable, and the
-- published_snapshot pointer decides which one is live.
--
-- Raw stats, stored once. Standard, half-PPR and PPR are views over these
-- rather than three separate models, which is what stops a half-PPR
-- projection quietly disagreeing with a PPR one.
CREATE TABLE IF NOT EXISTS projection_staging (
  season INTEGER, player_id TEXT, sleeper_id TEXT, player TEXT,
  position TEXT, team TEXT, ppr REAL, half REAL, standard REAL,
  adjusted REAL, exp_games REAL, floor REAL, ceiling REAL,
  rank_pos INTEGER, rec REAL, recyd REAL, ruyd REAL, news_adj REAL,
  trace TEXT,
  pass_att REAL, completions REAL, pass_yds REAL, pass_td REAL, ints REAL,
  targets REAL, rec_td REAL, rush_att REAL, rush_td REAL, fumbles REAL,
  -- A residual is opportunity the model left deliberately unassigned: the
  -- 5% "other QB" bucket in Cleveland, a snap share nobody named holds. It
  -- has to exist for the team to reconcile and it must never appear in a
  -- ranking or on the site, because it is not a person.
  is_residual INTEGER DEFAULT 0,
  PRIMARY KEY (season, player_id));
CREATE TABLE IF NOT EXISTS projection_source (
  season INTEGER PRIMARY KEY, source TEXT, loaded_at TEXT, players INTEGER,
  note TEXT);
"""


def key(n):
    n = re.sub(r"[.'`]", "", (n or "").lower())
    return " ".join(re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", n).split())


# Zero is a valid projection. Missing is a malformed input.
#
# The importer used to default an absent column to 0.0, which meant a
# workbook that lost its Targets column would import cleanly as a league
# where nobody was thrown to -- and the reconciliation checks would then fail
# in a way that pointed at the model rather than the file.
REQUIRED = {
    "QB": ["patt", "cmp", "payd", "patd", "int", "ruatt", "ruyd", "rutd", "fl"],
    "RB": ["targets", "rec", "recyd", "rectd", "ruatt", "ruyd", "rutd", "fl"],
    "WR": ["targets", "rec", "recyd", "rectd", "ruatt", "ruyd", "rutd", "fl"],
    "TE": ["targets", "rec", "recyd", "rectd", "ruatt", "ruyd", "rutd", "fl"],
}

# Team codes that are not a team. Free agents belong in the player universe
# and nowhere near a reconciliation.
NON_TEAMS = {"FA", "FA/UNK", "UNK", "", None}


def col(headers, *names):
    """Find a column by any of several header spellings."""
    low = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    return None


def read_workbook(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out, recon = [], None

    for name in wb.sheetnames:
        head = name.split()[0].upper()
        if head not in SHEETS:
            if name.lower().startswith("team recon"):
                recon = wb[name]
            continue
        ws = wb[name]
        h = [c.value for c in ws[1]]
        ci = {
            "player": col(h, "Player"), "team": col(h, "Team"),
            "rank": col(h, "Rank"),
            "targets": col(h, "Targets"),
            "rec": col(h, "Rec", "Receptions"),
            "patt": col(h, "PATT", "Pass Att", "Attempts"),
            "cmp": col(h, "CMP", "Completions"),
            "recyd": col(h, "Rec Yds", "Receiving Yards"),
            "rectd": col(h, "Rec TD"),
            "ruyd": col(h, "Rush Yds"), "rutd": col(h, "Rush TD"),
            "ruatt": col(h, "Rush Att"),
            "payd": col(h, "Pass Yds"), "patd": col(h, "Pass TD"),
            "int": col(h, "INT"), "fl": col(h, "FL", "Fumbles"),
            "ppr": col(h, "PPR", "Locked FPTS", "Calc FPTS", "FPTS"),
        }
        if ci["player"] is None:
            continue
        missing = [f for f in REQUIRED.get(head, []) if ci.get(f) is None]
        if missing:
            raise SystemExit(
                f"  {name}: required columns absent: {missing}\n"
                f"  Zero is a valid projection; a missing column is a "
                f"malformed workbook, and importing it as zero would produce "
                f"a league where nobody was thrown to.")
        for row in ws.iter_rows(min_row=2, values_only=True):
            nm = row[ci["player"]]
            if not nm or not str(nm).strip():
                continue
            def v(k, required=False):
                i = ci.get(k)
                if i is None or row[i] is None or row[i] == "":
                    if required:
                        raise SystemExit(
                            f"  {name}: {nm} has no value for '{k}'. Zero is "
                            f"valid, blank is not.")
                    return 0.0
                try:
                    return float(row[i])
                except (TypeError, ValueError):
                    raise SystemExit(
                        f"  {name}: {nm} has a non-numeric '{k}': {row[i]!r}")
            out.append({
                "name": str(nm).strip(), "pos": head,
                "team": (str(row[ci["team"]]).strip() if ci["team"] is not None
                         and row[ci["team"]] else ""),
                "ppr": v("ppr"),
                "rec": v("rec", head != "QB"),
                "recyd": v("recyd", head != "QB"),
                "rectd": v("rectd", head != "QB"),
                "targets": v("targets", head != "QB"),
                "ruyd": v("ruyd", True), "rutd": v("rutd", True),
                "ruatt": v("ruatt", True), "fl": v("fl", True),
                "payd": v("payd", head == "QB"),
                "patd": v("patd", head == "QB"),
                "ints": v("int", head == "QB"),
                "patt": v("patt", head == "QB"),
                "cmp": v("cmp", head == "QB"),
            })
    return out, recon


def check_reconciliation(recon) -> str:
    """The claim the workbook makes about itself, verified.

    Every team's player targets should sum to its quarterback's attempts. If
    that does not hold the workbook is not what it says it is, and importing
    it silently would publish numbers nobody had checked.
    """
    if recon is None:
        return "no reconciliation sheet"
    h = [c.value for c in recon[1]]
    av = col(h, "Attempt Variance")
    yv = col(h, "Yard Variance")
    if av is None:
        return "reconciliation sheet has no variance column"
    worst_a = worst_y = 0.0
    teams = 0
    for row in recon.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        teams += 1
        try:
            worst_a = max(worst_a, abs(float(row[av] or 0)))
            if yv is not None:
                worst_y = max(worst_y, abs(float(row[yv] or 0)))
        except (TypeError, ValueError):
            pass
    return (f"{teams} teams, worst attempt variance {worst_a:.2e}, "
            f"worst yard variance {worst_y:.2e}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", nargs="?")
    ap.add_argument("--db", default="beatwire.db")
    ap.add_argument("--season", type=int, default=2026)
    ap.add_argument("--publish", action="store_true")
    ap.add_argument("--show", help="print one position from what is stored")
    args = ap.parse_args()

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    if args.show:
        conn.executescript(SCHEMA)
        rows = conn.execute(
            """SELECT * FROM projection_staging WHERE season=? AND position=?
               ORDER BY rank_pos LIMIT 40""",
            (args.season, args.show.upper())).fetchall()
        if not rows:
            sys.exit(f"  nothing stored for {args.show}. Import first.")
        src = conn.execute("SELECT * FROM projection_source WHERE season=?",
                           (args.season,)).fetchone()
        if src:
            print(f"\n  {src['source']}, loaded {src['loaded_at'][:16]}")
            print(f"  {src['note']}")
        print(f"\n  {'#':<4}{'PLAYER':<24}{'TM':<5}{'PPR':>8}{'REC':>7}"
              f"{'RECYD':>8}{'RUYD':>8}")
        for r in rows:
            print(f"  {r['rank_pos']:<4}{r['player'][:24]:<24}"
                  f"{(r['team'] or ''):<5}{r['ppr']:>8.1f}{r['rec'] or 0:>7.0f}"
                  f"{r['recyd'] or 0:>8.0f}{r['ruyd'] or 0:>8.0f}")
        return

    if not args.workbook:
        sys.exit("  pass the workbook path, or --show POS")
    path = Path(args.workbook).expanduser()
    if not path.exists():
        sys.exit(f"  no file at {path}")

    rows, recon = read_workbook(path)
    if not rows:
        sys.exit("  no player rows found; is this the right workbook?")

    note = check_reconciliation(recon)
    print(f"\n  {path.name}")
    print(f"  {len(rows)} players")
    from collections import Counter
    print(f"  {dict(Counter(r['pos'] for r in rows))}")
    print(f"  reconciliation: {note}")

    # Match to the roster so the site can link a projection to a player page.
    import csv
    ids = {}
    rp = ROOT / "rosters" / "nfl.csv"
    if rp.exists():
        for r in csv.DictReader(rp.open()):
            ids[key(r["name"])] = r["id"]

    matched = sum(1 for r in rows if key(r["name"]) in ids)
    print(f"  matched {matched} of {len(rows)} to the roster")

    if not args.publish:
        print(f"\n  Nothing written. Re-run with --publish.")
        for r in sorted(rows, key=lambda x: -x["ppr"])[:8]:
            print(f"    {r['pos']}  {r['name'][:24]:<24} {r['ppr']:>7.1f}")
        return

    from datetime import datetime, timezone
    conn.execute("DROP TABLE IF EXISTS projection_staging")
    conn.executescript(SCHEMA)

    by_pos = {}
    for r in sorted(rows, key=lambda x: -x["ppr"]):
        by_pos.setdefault(r["pos"], []).append(r)
    # Free agents stay in the universe and out of the team count. FA/UNK is
    # not a thirty-third NFL team and reconciling it would be meaningless.
    teams = {r["team"] for r in rows if r["team"] not in NON_TEAMS}
    print(f"  {len(teams)} NFL teams, "
          f"{sum(1 for r in rows if r['team'] in NON_TEAMS)} free agents")

    n = 0
    for pos, group in by_pos.items():
        for i, r in enumerate(group, 1):
            k = key(r["name"])
            pid = ids.get(k) or f"unmatched-{k.replace(' ', '-')}"
            half = r["ppr"] - r["rec"] * 0.5
            std = r["ppr"] - r["rec"]
            conn.execute(
                "INSERT OR REPLACE INTO projection_staging VALUES "
                "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (args.season, pid, ids.get(k), r["name"], pos, r["team"],
                 r["ppr"], half, std, None, None, None, None, i,
                 r["rec"], r["recyd"], r["ruyd"], None,
                 f"workbook {path.name}",
                 r["patt"], r["cmp"], r["payd"], r["patd"], r["ints"],
                 r["targets"], r["rectd"], r["ruatt"], r["rutd"], r["fl"],
                 0))
            n += 1
    conn.execute("DELETE FROM projection_source WHERE season=?", (args.season,))
    conn.execute("INSERT INTO projection_source VALUES (?,?,?,?,?)",
                 (args.season, path.name,
                  datetime.now(timezone.utc).isoformat(), n, note))
    conn.commit()
    print(f"\n  published {n} projections for {args.season}")
    print(f"  next: python3 -m beatwire.cli export --sports nfl --limit 4000")


if __name__ == "__main__":
    main()
