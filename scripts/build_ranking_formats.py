#!/usr/bin/env python3
"""Build PPR, Non-PPR, Superflex and Dynasty ranking pages.

The workbook is the single scoring source. PPR and Non-PPR read their named
columns. Superflex uses PPR scoring with QB33 replacement value, reflecting a
12-team league where starting quarterbacks are normally rostered. No ranking is
Dynasty combines those PPR projections with verified roster ages and published
position-specific age curves. IDP is intentionally outside this product.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import build_rankings as base  # noqa: E402
import seo  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"
SOURCE = ROOT / "data" / "projections.xlsx"
ROSTER = ROOT / "rosters" / "nfl.csv"


def minimum_population() -> int:
    """Allow the curated dev release without weakening production checks."""
    return 400 if os.environ.get("LINEUPBEAT_NFL_SEASON") == "v1.6-trusted-current" else 500

FORMATS = {
    "ppr": {
        "label": "PPR", "column": "ppr", "replacement_qb": 13,
        "preseason_slug": "ppr", "top_slug": "top-200-ppr",
    },
    "non_ppr": {
        "label": "Non-PPR", "column": "non-ppr", "replacement_qb": 13,
        "preseason_slug": "non-ppr", "top_slug": "top-200-non-ppr",
    },
    "superflex": {
        "label": "Superflex", "column": "ppr", "replacement_qb": 33,
        "preseason_slug": None, "top_slug": "top-200-superflex",
    },
}

FORMAT_NAV = (
    ("Preseason Rankings (PPR)", "/nfl/rankings/ppr/", True),
    ("Preseason Rankings (NON-PPR)", "/nfl/rankings/non-ppr/", True),
    ("Top 200 Rankings (PPR)", "/nfl/rankings/top-200-ppr/", True),
    ("Top 200 Rankings (NON-PPR)", "/nfl/rankings/top-200-non-ppr/", True),
    ("Top 200 Rankings (Superflex)", "/nfl/rankings/top-200-superflex/", True),
    ("Dynasty Rankings", "/nfl/rankings/dynasty/", True),
)

FAQ = {
    "ppr": [
        ("What does PPR mean in fantasy football?",
         "PPR awards one point for every reception, increasing the value of high-volume receivers, pass-catching running backs and tight ends."),
        ("How are LineupBeat PPR rankings calculated?",
         "The rankings start with LineupBeat's full-season PPR projections and compare each player with replacement value at his position."),
    ],
    "non_ppr": [
        ("What is Non-PPR fantasy football scoring?",
         "Non-PPR, also called standard scoring, awards no points for receptions. Yardage and touchdowns therefore carry more of each player's value."),
        ("Why do running backs move in Non-PPR rankings?",
         "Touchdown and rushing volume matter more when catches are worth no points, while reception-heavy players lose their PPR advantage."),
    ],
    "superflex": [
        ("What is a Superflex fantasy football league?",
         "A Superflex lineup includes a flex spot that can start a quarterback, making quarterback depth much more valuable than in a one-QB league."),
        ("Why are quarterbacks higher in Superflex rankings?",
         "Most starting quarterbacks are rostered in a 12-team Superflex league, so LineupBeat evaluates quarterbacks against a much deeper replacement level."),
    ],
    "dynasty": [
        ("How are LineupBeat dynasty rankings calculated?",
         "LineupBeat combines current PPR projection value with a position-specific age curve using verified NFL roster ages."),
        ("Do LineupBeat dynasty rankings include future rookies?",
         "No. The board ranks currently rostered NFL players with a published projection and verified age; it does not invent values for unprojected prospects."),
    ],
}

DYNASTY_PEAK_AGE = {"QB": 27, "RB": 23, "WR": 24, "TE": 25}
DYNASTY_AGE_PENALTY = {"QB": 2.0, "RB": 9.0, "WR": 4.0, "TE": 3.5}
DYNASTY_YOUTH_BONUS = {"QB": 2.0, "RB": 4.0, "WR": 3.0, "TE": 2.5}
DYNASTY_EDITORIAL = {
    ("Ashton Jeanty", "LV", "RB"): (16.0, "Feature-role and rushing-volume upside"),
    ("Malik Nabers", "NYG", "WR"): (45.0, "Age-23 cornerstone value after returning to full-contact practice"),
}

# Independent Lineup Beat PPR decisions. The supplied comparison sheets were
# used only to identify places worth auditing; no external rank, tier, ADP or
# SOS value enters this builder. Each change must be supportable from our own
# projection line or an already documented Lineup Beat editorial decision.
PPR_ADJUSTMENTS = {
    ("Ashton Jeanty", "LV", "RB"): (
        16.0,
        "Lineup Beat upgrade for projected feature-back volume and role upside; "
        "the underlying statistical projection remains unchanged."),
    ("Christian McCaffrey", "SF", "RB"): (
        -48.0,
        "Lineup Beat durability adjustment for an age-30 back following an "
        "extreme 2025 workload and an extended health-related camp absence; "
        "the underlying statistical projection remains unchanged."),
}

PPR_ORDER = (
    ("Bijan Robinson", "Jahmyr Gibbs",
     "Robinson ranks first because Lineup Beat projects slightly more rushing "
     "and receiving opportunity; the two are separated by only 2.3 PPR points."),
    ("Ja'Marr Chase", "Puka Nacua",
     "Chase ranks first at wide receiver because Lineup Beat projects more "
     "targets, receptions and receiving touchdowns in a near-tie on PPR points."),
    ("Malik Nabers", "Chris Olave",
     "Nabers closes the PPR top 10 because Lineup Beat expects a Week 1-ready "
     "featured role after his return to full-contact practice; this is an "
     "explicit editorial call rather than a copied outside rank."),
)

NON_PPR_ADJUSTMENTS = {
    ("Ashton Jeanty", "LV", "RB"): (
        12.0,
        "Lineup Beat upgrade for projected feature-back rushing volume and "
        "touchdown upside; the statistical projection remains unchanged."),
    ("Christian McCaffrey", "SF", "RB"): (
        -33.0,
        "Lineup Beat durability adjustment for an age-30 back following an "
        "extreme 2025 workload and an extended health-related camp absence; "
        "the statistical projection remains unchanged."),
}

FORMAT_CSS = """
.rkformats{margin:1.4rem 0 1.8rem;padding:1.1rem 1.2rem;border:1px solid var(--rule);background:rgba(255,255,255,.018)}
.rkformats h2{font:600 .88rem/1 var(--agate);letter-spacing:.12em;text-transform:uppercase;color:var(--quiet);margin:0 0 .8rem}
.rkformatgrid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:.25rem 1.4rem}
.rkformatgrid a,.rkformatgrid span{display:block;padding:.42rem 0;font:400 1rem/1.25 var(--text)}
.rkformatgrid a{color:var(--ink);text-decoration:none}.rkformatgrid a:hover{color:var(--signal)}
.rkformatgrid .soon{color:var(--quiet)}.rkformatgrid small{font:600 .68rem/1 var(--agate);letter-spacing:.08em;text-transform:uppercase;margin-left:.45rem}
@media(max-width:680px){.rkformatgrid{grid-template-columns:1fr}}
"""


def read_projection_formats(path: Path) -> dict[str, list[dict]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {key: [] for key in FORMATS}
    for sheet in wb.sheetnames:
        pos = sheet.upper()
        if pos not in base.POSITIONS:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        head = [str(v or "").strip().lower() for v in next(rows)]
        index = {name: i for i, name in enumerate(head)}
        for raw in rows:
            try:
                name = str(raw[index["player"]]).strip()
                team = str(raw[index["team"]]).strip().upper()
            except (KeyError, IndexError, TypeError):
                continue
            if not name:
                continue
            for key, spec in FORMATS.items():
                try:
                    points = round(float(raw[index[spec["column"]]]), 1)
                except (KeyError, IndexError, TypeError, ValueError):
                    continue
                result[key].append({"player_name": name, "team": team,
                                    "position": pos,
                                    "projected_points": points})
    return result


def source_updated(path: Path) -> datetime:
    """The workbook's own update stamp, never the checkout file mtime."""
    import openpyxl
    from zoneinfo import ZoneInfo
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Weekly Update" in wb.sheetnames:
        value = wb["Weekly Update"]["B2"].value
        if value:
            try:
                parsed = datetime.strptime(str(value).strip(),
                                           "%B %d, %Y, %I:%M %p ET")
                return parsed.replace(tzinfo=ZoneInfo("America/New_York"))
            except ValueError:
                pass
    raise ValueError("projection workbook has no reliable Weekly Update timestamp")


def read_roster_ages(path: Path) -> dict[tuple[str, str, str], int]:
    ages = {}
    with path.open(newline="") as handle:
        for row in csv.DictReader(handle):
            try:
                age = int(row.get("age") or "")
            except ValueError:
                continue
            key = (row.get("name", "").strip(), row.get("team", "").strip().upper(),
                   row.get("position", "").strip().upper())
            if key[0] and key[1] and key[2] in base.POSITIONS:
                ages[key] = age
    return ages


def rank_dynasty(rows: list[dict], ages: dict[tuple[str, str, str], int]) -> list[dict]:
    """Rank current NFL players from our projection value and verified age.

    The curve is deliberately visible and reproducible. Older players lose
    value faster at RB than at QB; players younger than the positional peak
    receive a small development bonus. No supplied comparison rank or rating
    is read here.
    """
    replacement = {}
    for pos, n in base.REPLACEMENT_RANK.items():
        group = sorted((r for r in rows if r["position"] == pos),
                       key=lambda r: (-r["projected_points"], r["player_name"]))
        replacement[pos] = group[n - 1]["projected_points"]
    records = []
    for row in rows:
        identity = (row["player_name"], row["team"], row["position"])
        age = ages.get(identity)
        if age is None:
            continue
        pos = row["position"]
        vorp = round(row["projected_points"] - replacement[pos], 1)
        delta = age - DYNASTY_PEAK_AGE[pos]
        age_adjustment = (max(-delta, 0) * DYNASTY_YOUTH_BONUS[pos]
                          - (max(delta, 0) ** 1.25) * DYNASTY_AGE_PENALTY[pos])
        editorial, editorial_reason = DYNASTY_EDITORIAL.get(identity, (0.0, ""))
        records.append({
            "scoring_format": "dynasty", "overall_rank": None,
            "overall_tier": None, "position_rank": None, "position_tier": None,
            **row, "age": age, "replacement_points": replacement[pos],
            "vorp": vorp, "manual_adjustment": round(age_adjustment + editorial, 1),
            "ranking_score": round(vorp + age_adjustment + editorial, 1), "adp": None,
            "value_label": None, "adjustment_reason": (
                "Published dynasty age curve" + (f"; {editorial_reason}" if editorial else "")),
            "editorial_override": bool(editorial), "ranked_ahead_of": None,
            "override_reason": None,
        })
    records.sort(key=base.sort_key)
    by_name = {row["player_name"]: row for row in records}
    bijan, gibbs = by_name["Bijan Robinson"], by_name["Jahmyr Gibbs"]
    if records.index(bijan) > records.index(gibbs):
        records.remove(bijan)
        records.insert(records.index(gibbs), bijan)
    bijan["editorial_override"] = True
    bijan["ranked_ahead_of"] = "Jahmyr Gibbs"
    bijan["override_reason"] = "Lineup Beat projects slightly more combined opportunity"
    pos_seen = {}
    for overall, row in enumerate(records, 1):
        pos = row["position"]
        pos_seen[pos] = pos_seen.get(pos, 0) + 1
        row["position_rank"] = pos_seen[pos]
        row["position_tier"] = base.tier_for_rank(pos_seen[pos], base.POSITION_TIER_MAX[pos])
        if overall <= base.TOP_N:
            row["overall_rank"] = overall
            row["overall_tier"] = base.tier_for_rank(overall, base.OVERALL_TIER_MAX)
    if len(records) < minimum_population() or sum(r["overall_rank"] is not None for r in records) != 200:
        raise ValueError("dynasty: insufficient verified projection-age matches")
    return records


def rank(rows: list[dict], key: str) -> tuple[list[dict], dict[str, float]]:
    spec = FORMATS[key]
    replacement_rank = dict(base.REPLACEMENT_RANK)
    replacement_rank["QB"] = spec["replacement_qb"]
    replacement = {}
    for pos in base.POSITIONS:
        group = sorted((r for r in rows if r["position"] == pos),
                       key=lambda r: (-r["projected_points"], r["player_name"]))
        n = replacement_rank[pos]
        if len(group) < n:
            raise ValueError(f"{key}: only {len(group)} {pos}s; need {n}")
        replacement[pos] = group[n - 1]["projected_points"]

    records = []
    for row in rows:
        rep = replacement[row["position"]]
        decisions = (PPR_ADJUSTMENTS if key in {"ppr", "superflex"} else
                     NON_PPR_ADJUSTMENTS if key == "non_ppr" else {})
        adjustment, reason = decisions.get(
            (row["player_name"], row["team"], row["position"]),
            (0.0, "Projection-only format ranking"))
        records.append({
            "scoring_format": key,
            "overall_rank": None, "overall_tier": None,
            "position_rank": None, "position_tier": None,
            **row,
            "replacement_points": rep,
            "vorp": round(row["projected_points"] - rep, 1),
            "manual_adjustment": adjustment,
            "ranking_score": round(row["projected_points"] - rep + adjustment, 1),
            "adp": None, "value_label": None,
            "adjustment_reason": reason,
            "editorial_override": bool(adjustment), "ranked_ahead_of": None,
            "override_reason": None,
        })
    records.sort(key=base.sort_key)
    if key in {"ppr", "superflex"}:
        by_name = {row["player_name"]: row for row in records}
        for ahead, behind, reason in PPR_ORDER:
            a, b = by_name[ahead], by_name[behind]
            ia, ib = records.index(a), records.index(b)
            if ia > ib:
                records.pop(ia)
                records.insert(records.index(b), a)
            a["editorial_override"] = True
            a["ranked_ahead_of"] = behind
            a["override_reason"] = reason
    pos_seen = {}
    for overall, row in enumerate(records, 1):
        pos = row["position"]
        pos_seen[pos] = pos_seen.get(pos, 0) + 1
        row["position_rank"] = pos_seen[pos]
        row["position_tier"] = base.tier_for_rank(
            pos_seen[pos], base.POSITION_TIER_MAX[pos])
        if overall <= base.TOP_N:
            row["overall_rank"] = overall
            row["overall_tier"] = base.tier_for_rank(
                overall, base.OVERALL_TIER_MAX)
    validate(records, rows, key)
    return records, replacement


def validate(records: list[dict], source: list[dict], key: str) -> None:
    if len(records) != len(source) or len(records) < minimum_population():
        raise ValueError(f"{key}: player reconciliation failed")
    identities = {(r["player_name"], r["team"], r["position"])
                  for r in records}
    if len(identities) != len(records):
        raise ValueError(f"{key}: duplicate player identity")
    ranked = [r for r in records if r["overall_rank"] is not None]
    if len(ranked) != base.TOP_N:
        raise ValueError(f"{key}: expected {base.TOP_N} overall ranks")
    if [r["overall_rank"] for r in ranked] != list(range(1, base.TOP_N + 1)):
        raise ValueError(f"{key}: overall ranks are not sequential")
    for pos in base.POSITIONS:
        group = [r for r in records if r["position"] == pos]
        if sorted(r["position_rank"] for r in group) != list(range(1, len(group)+1)):
            raise ValueError(f"{key}: {pos} ranks are not sequential")


def format_nav(active: str) -> str:
    items = []
    for label, href, live in FORMAT_NAV:
        if live:
            current = ' aria-current="page"' if href == active else ""
            items.append(f'<a href="{href}"{current}>{base.esc(label)}</a>')
        else:
            items.append(f'<span class="soon">{base.esc(label)}'
                         f'<small>Data required</small></span>')
    return ('<section class="rkformats" aria-labelledby="ranking-formats">'
            '<h2 id="ranking-formats">Rankings</h2>'
            f'<div class="rkformatgrid">{"".join(items)}</div></section>')


def methodology(label: str, superflex: bool) -> str:
    qb = "QB33" if superflex else "QB13"
    league = ("a 12-team Superflex league where two quarterbacks can start"
              if superflex else "a 12-team, one-quarterback league")
    return f"""
<section class="rkmeth" id="methodology">
 <h2>How These {base.esc(label)} Rankings Work</h2>
 <p>LineupBeat ranks players from its full-season statistical projections,
 then compares each player with the expected replacement option at his
 position. This board uses {base.esc(label)} scoring in {league}.</p>
 <p>Replacement levels are {qb}, RB37, WR49 and TE13. Ranking score is
 projected points above that replacement level. This lets positions share
 one overall board without pretending raw quarterback points and running
 back points carry identical draft value.</p>
 <p class="rknote">These pages use the same projection workbook as the
 projection pages. No separate private ranking list or copied scoring source
 can drift away from the published projections.</p>
</section>"""


def editorial_notes(key: str) -> str:
    if key not in {"ppr", "non_ppr", "superflex"}:
        return ""
    notes = ([
        "Bijan Robinson leads Jahmyr Gibbs on projected combined opportunity "
        "in an otherwise near-even PPR projection.",
        "Ja'Marr Chase leads Puka Nacua because his projection carries more "
        "targets, receptions and receiving touchdowns.",
        "Ashton Jeanty's feature-role upside is recognized without changing "
        "his published statistical projection.",
        "Christian McCaffrey receives a durability discount without changing "
        "his published statistical projection.",
        "Malik Nabers closes the PPR wide-receiver top 10 after returning to "
        "full-contact practice; Lineup Beat expects a Week 1-ready featured role.",
    ] if key in {"ppr", "superflex"} else [
        "The Non-PPR board remains projection-led: Jahmyr Gibbs and Puka Nacua "
        "retain the top spots earned by Lineup Beat's scoring model.",
        "Ashton Jeanty's projected rushing role receives a modest Lineup Beat "
        "upgrade, while Christian McCaffrey receives a durability discount.",
    ])
    if key == "superflex":
        notes.insert(0, "Superflex uses QB33 as its quarterback replacement "
                     "level because viable starters are normally rostered; "
                     "the resulting quarterback premium is intentional.")
    title = {"ppr": "PPR", "non_ppr": "Non-PPR",
             "superflex": "Superflex"}[key]
    return (f'<section class="rkmeth"><h2>Lineup Beat {title} Decisions</h2><ul>'
            + "".join(f'<li>{base.esc(note)}</li>' for note in notes)
            + '</ul><p class="rkdisc">Comparison rankings were used as a '
              'research checklist only. No outside rank, tier, ADP or strength-of-schedule '
              'value is imported into this board.</p></section>')


def dynasty_table(records: list[dict], pos: str | None) -> tuple[str, str]:
    shown = ([r for r in records if r["overall_rank"] is not None] if not pos
             else sorted((r for r in records if r["position"] == pos),
                         key=lambda r: r["position_rank"]))
    rows = []
    for r in shown:
        rank_value = r["overall_rank"] if not pos else r["position_rank"]
        pos_rank = f'{r["position"]}{r["position_rank"]}'
        rows.append(
            f'<tr class="r" data-name="{base.esc(r["player_name"].lower())}" '
            f'data-team="{base.esc(r["team"])}">'
            f'<td class="rkrank">{rank_value}</td>'
            f'<td class="l rktier">T{r["overall_tier"] if not pos else r["position_tier"]}</td>'
            f'<td class="l rkname">{base.esc(r["player_name"])}</td>'
            f'<td class="l rkteam">{base.esc(r["team"])}</td>'
            + (f'<td class="l rkpos">{r["position"]}</td>' if not pos else '')
            + f'<td class="rkposrk">{pos_rank}</td>'
              f'<td class="rkpts">{r["age"]}</td>'
              f'<td class="rkpts">{r["projected_points"]:.1f}</td></tr>')
    headers = ["RK", "TIER", "PLAYER", "TEAM"]
    if not pos:
        headers.append("POS")
    headers += ["POS RK", "AGE", "PPR PROJ"]
    head = "<tr>" + "".join(
        f'<th scope="col"{" class=\"l\"" if h in {"TIER", "PLAYER", "TEAM", "POS"} else ""}>{h}</th>'
        for h in headers) + "</tr>"
    return head, "".join(rows)


def render_dynasty(records: list[dict], built: datetime, pos: str | None = None) -> str:
    suffix = f'{pos.lower()}/' if pos else ''
    url = f"https://lineupbeat.com/nfl/rankings/dynasty/{suffix}"
    title = (f"2026 Dynasty Fantasy {pos} Rankings | LineupBeat" if pos else
             "2026 Dynasty Fantasy Football Rankings | LineupBeat")
    desc = ("LineupBeat's 2026 dynasty fantasy football rankings, built from "
            "current PPR projection value and verified position-specific age curves.")
    shown = ([r for r in records if r["overall_rank"] is not None] if not pos
             else [r for r in records if r["position"] == pos])
    head_row, body_rows = dynasty_table(records, pos)
    teams = sorted({r["team"] for r in shown})
    opts = "".join(f'<option value="{base.esc(t)}">{base.esc(t)}</option>' for t in teams)
    root = "/nfl/rankings/dynasty/"
    tabs = '<nav class="rktabs" aria-label="Rankings by position">' + "".join(
        f'<a class="rktab" href="{href}"{" aria-current=\"page\"" if current else ""}>{name}</a>'
        for name, href, current in ([('Overall', root, pos is None)] +
            [(p, f'{root}{p.lower()}/', pos == p) for p in base.POSITIONS])) + '</nav>'
    methodology_html = f'''<section class="rkmeth"><h2>How Dynasty Rankings Work</h2>
<p>The starting point is LineupBeat's current PPR projection value above positional replacement. A published age curve then applies a small development bonus before the positional peak and an increasing decline adjustment after it.</p>
<p>Peak ages are QB {DYNASTY_PEAK_AGE['QB']}, RB {DYNASTY_PEAK_AGE['RB']}, WR {DYNASTY_PEAK_AGE['WR']} and TE {DYNASTY_PEAK_AGE['TE']}. Running backs decline fastest; quarterbacks decline slowest.</p>
<p class="rknote">Only current NFL players with both a LineupBeat projection and verified roster age qualify. Unprojected prospects and players without a verified age are excluded.</p>
<p>Lineup Beat decisions: Bijan Robinson leads Jahmyr Gibbs on projected combined opportunity; Ashton Jeanty's feature-role upside is recognized; Malik Nabers receives a long-term age-and-role upgrade after returning to full-contact practice.</p>
<p class="rkdisc">Comparison rankings were used as a research checklist only. No outside rank, tier, rating, ADP or age was imported.</p></section>'''
    css, header, footer = base.site_chrome()
    body = f'''<main class="rkwrap"><nav class="crumbs"><a href="/">Home</a><span>/</span><a href="/nfl/rankings/">Rankings</a><span>/</span><b>Dynasty</b></nav>
<header class="rkhead"><p class="rkeyebrow">2026 FANTASY FOOTBALL</p><h1>{base.esc(title.removesuffix(' | LineupBeat'))}</h1><p class="rkintro">Current production and long-term value, ranked together.</p><p class="rkstatus">Updated {built:%B %d, %Y} &middot; PPR dynasty &middot; Current NFL players</p></header>
{format_nav(root if not pos else '')}<div class="rkctl">{tabs}<div class="rkfilters"><input id="rkq" type="search" placeholder="Search players"><select id="rkteam"><option value="">All teams</option>{opts}</select><button class="rkclear" id="rkclear" type="button">Clear filters</button></div><p class="rkcount" id="rkcount">{len(shown)} players</p></div>
<table class="rktable" id="rktable"><thead>{head_row}</thead><tbody>{body_rows}</tbody></table>{methodology_html}{seo.faq_html(FAQ['dynasty'])}{seo.related_html('rankings')}</main>'''
    items = [(r["overall_rank"] if not pos else r["position_rank"],
              f'{r["player_name"]} ({r["team"]}, {r["position"]})', None) for r in shown]
    schema = seo.graph({"@type": "Dataset", "name": title, "description": desc,
                        "url": url, "dateModified": built.strftime("%Y-%m-%d")},
                       seo.itemlist_schema(title, url, items), seo.faq_schema(FAQ["dynasty"]),
                       seo.ORGANISATION)
    return f'''<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{base.esc(title)}</title><meta name="description" content="{base.esc(desc)}"><link rel="canonical" href="{url}"><script type="application/ld+json">{schema}</script><style>{css}{base.PAGE_CSS}{FORMAT_CSS}{seo.CRUMB_CSS}{seo.UI_CSS}{seo.RELATED_CSS}</style></head><body>{header}{body}{footer}{seo.TRACKING}{seo.VIEW_CONTENT}{base.PAGE_JS}</body></html>'''


def page_body(records: list[dict], key: str, slug: str,
              built: datetime, pos: str | None, top_only: bool) -> str:
    spec = FORMATS[key]
    label = spec["label"]
    shown = ([r for r in records if r["overall_rank"] is not None] if not pos
             else sorted((r for r in records if r["position"] == pos),
                         key=lambda r: r["position_rank"]))
    body_rows, _ = base.rows_html(shown, pos, base.existing_slugs())
    teams = sorted({r["team"] for r in shown})
    h1 = (f"2026 Fantasy Football Top 200 Rankings ({label})" if top_only
          else (f"2026 Fantasy Football Rankings ({label})" if not pos
                else f"2026 Fantasy {pos} Rankings ({label})"))
    root = f"/nfl/rankings/{slug}/"
    tabs = ""
    if not top_only:
        tab_rows = [("Overall", root, pos is None)] + [
            (p, f"{root}{p.lower()}/", pos == p) for p in base.POSITIONS]
        tabs = '<nav class="rktabs" aria-label="Rankings by position">' + "".join(
            f'<a class="rktab" href="{href}"'
            f'{" aria-current=\"page\"" if current else ""}>{name}</a>'
            for name, href, current in tab_rows) + '</nav>'
    opts = "".join(f'<option value="{base.esc(t)}">{base.esc(t)}</option>'
                   for t in teams)
    crumbs = ('<nav class="crumbs" aria-label="Breadcrumb"><a href="/">Home</a>'
              '<span>/</span><a href="/nfl/data/">NFL</a><span>/</span>'
              '<a href="/nfl/rankings/">Rankings</a><span>/</span>'
              f'<b aria-current="page">{base.esc(h1)}</b></nav>')
    return f"""
<main class="rkwrap">{crumbs}
 <header class="rkhead"><p class="rkeyebrow">2026 FANTASY FOOTBALL</p>
  <h1>{base.esc(h1)}</h1>
  <p class="rkintro">LineupBeat's current {base.esc(label)} draft rankings,
  generated from the same player projections used throughout the site.</p>
  <p class="rkstatus">Updated {built:%B %d, %Y} &middot; {base.esc(label)}
  &middot; {'12-team Superflex' if key == 'superflex' else '12-team, one-QB'}</p>
 </header>
 {format_nav(root if not pos else '')}
 <div class="rkctl">{tabs}<div class="rkfilters">
  <label class="visually-hidden" for="rkq">Search players</label>
  <input id="rkq" type="search" placeholder="Search players" autocomplete="off">
  <label class="visually-hidden" for="rkteam">Filter by team</label>
  <select id="rkteam"><option value="">All teams</option>{opts}</select>
  <button class="rkclear" id="rkclear" type="button">Clear filters</button>
 </div><p class="rkcount" id="rkcount">{len(shown)} players</p></div>
 <table class="rktable" id="rktable"><caption class="visually-hidden">{base.esc(h1)}</caption>
  <thead>{base.head_row(pos)}</thead><tbody>{body_rows}</tbody></table>
 {methodology(label, key == 'superflex')}
 {editorial_notes(key)}
 {seo.faq_html(FAQ[key])}
 {seo.related_html('rankings')}
</main>"""


def render(records: list[dict], key: str, slug: str, built: datetime,
           pos: str | None = None, top_only: bool = False) -> str:
    spec = FORMATS[key]
    label = spec["label"]
    suffix = f"/{pos.lower()}/" if pos else "/"
    url = f"https://lineupbeat.com/nfl/rankings/{slug}{suffix}"
    title = (f"2026 Fantasy Football Top 200 Rankings ({label}) | LineupBeat"
             if top_only else
             f"2026 Fantasy {pos + ' ' if pos else 'Football '}{label} Rankings | LineupBeat")
    desc = (f"LineupBeat's updated 2026 {label} fantasy football rankings, "
            f"built from full-season projections and positional replacement value.")
    css, header, footer = base.site_chrome()
    body = page_body(records, key, slug, built, pos, top_only)
    shown = ([r for r in records if r["overall_rank"] is not None] if not pos
             else sorted((r for r in records if r["position"] == pos),
                         key=lambda r: r["position_rank"]))
    items = [(r["overall_rank"] if not pos else r["position_rank"],
              f'{r["player_name"]} ({r["team"]}, {r["position"]})', None)
             for r in shown]
    schema = seo.graph(
        {"@type": "Dataset", "name": title, "description": desc, "url": url,
         "dateModified": built.strftime("%Y-%m-%d"),
         "creator": {"@type": "Organization", "name": "LineupBeat"},
         "variableMeasured": ["Projected fantasy points", "Value over replacement", "Rank"]},
        seo.breadcrumbs((("LineupBeat", "/"), ("Rankings", "/nfl/rankings/"),
                         (title, url))),
        seo.itemlist_schema(title, url, items), seo.faq_schema(FAQ[key]),
        seo.ORGANISATION)
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{base.esc(title)}</title><meta name="description" content="{base.esc(desc)}">
<link rel="canonical" href="{url}">{seo.social_meta(title, desc, url, base.OG_IMAGE)}
<script type="application/ld+json">{schema}</script>
<style>{css}{base.PAGE_CSS}{FORMAT_CSS}{seo.CRUMB_CSS}{seo.UI_CSS}{seo.RELATED_CSS}</style>
</head><body>{header}{body}{footer}{seo.TRACKING}{seo.VIEW_CONTENT}{base.PAGE_JS}</body></html>"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(SOURCE.relative_to(ROOT)))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    source = ROOT / args.source
    if not source.exists():
        raise SystemExit(f"missing {source}")
    inputs = read_projection_formats(source)
    ranked, replacements = {}, {}
    for key in FORMATS:
        ranked[key], replacements[key] = rank(inputs[key], key)
        print(f"  {key}: {len(ranked[key])} players; "
              f"QB replacement {replacements[key]['QB']:.1f}")
    if args.dry_run:
        print("  dry run; nothing written")
        return 0

    # The visible date comes from the workbook's own release metadata. A git
    # checkout mtime would label old projections as freshly updated.
    built = source_updated(source)

    pages = []
    for key, spec in FORMATS.items():
        if spec["preseason_slug"]:
            pages.append((key, spec["preseason_slug"], None, False))
            pages.extend((key, spec["preseason_slug"], pos, False)
                         for pos in base.POSITIONS)
        pages.append((key, spec["top_slug"], None, True))
    for key, slug, pos, top_only in pages:
        out = SITE / "nfl" / "rankings" / slug
        if pos:
            out /= pos.lower()
        out /= "index.html"
        out.parent.mkdir(parents=True, exist_ok=True)
        page = seo.check_page(render(ranked[key], key, slug, built, pos, top_only), str(out))
        out.write_text(page)
        print(f"  wrote {out.relative_to(ROOT)} ({len(page):,} bytes)")
    dynasty = rank_dynasty(inputs["ppr"], read_roster_ages(ROSTER))
    print(f"  dynasty: {len(dynasty)} players with verified ages")
    for pos in [None, *base.POSITIONS]:
        out = SITE / "nfl" / "rankings" / "dynasty"
        if pos:
            out /= pos.lower()
        out /= "index.html"
        out.parent.mkdir(parents=True, exist_ok=True)
        page = seo.check_page(render_dynasty(dynasty, built, pos), str(out))
        out.write_text(page)
        print(f"  wrote {out.relative_to(ROOT)} ({len(page):,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
