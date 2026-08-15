#!/usr/bin/env python3
"""Build the projection changes page from the workbook's Weekly Update tab.

    python3 scripts/build_changes.py
    python3 scripts/build_changes.py --projections data/projections.xlsx

WHY THIS PAGE EXISTS

The board says what we project. It does not say what we changed our mind
about, or why, and that is the more interesting claim: anybody can publish
numbers, far fewer will publish the moment their numbers moved and point at
the report that moved them.

It is also the honest version of "fantasy analysis". Every line here is a
decision somebody made with a stated reason and a source link, rather than
a sentence generated to sound like insight.

WHAT THE DECISIONS MEAN

  UPDATED           the projection moved on this evidence
  RECONCILED        moved because somebody else's did -- a backfield only
                    has so many carries, so raising one back lowers another
  OUT, ZEROED       ruled out for the season, projection set to zero
  WATCH, NO CHANGE  we are aware and have not moved anything yet

That last one matters most. A changelog that only records changes implies
everything else went unexamined; recording the things looked at and left
alone is what makes it a record rather than a highlight reel.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent))
import seo

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"
SPORT = "nfl"
POSITIONS = ["QB", "RB", "WR", "TE"]

DECISION_NOTE = {
    "UPDATED": "The projection moved on this evidence.",
    "RECONCILED": "Moved to stay consistent with another change. A backfield "
                  "only has so many carries.",
    "OUT, ZEROED": "Ruled out for the season. Projection set to zero.",
    "WATCH, NO CHANGE": "Looked at, nothing moved yet.",
}

EVIDENCE_ORDER = {"High": 0, "Medium": 1, "Medium-Low": 2, "Uncertain": 3}


def eastern_now():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/New_York"))
    except Exception:
        return datetime.now(timezone.utc) - timedelta(hours=4)


def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)


def slug(s):
    s = re.sub(r"[^\w\s-]", "", (s or "").lower())
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s)
    return re.sub(r"[\s_]+", "-", s).strip("-")


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_changes(path: Path):
    """The Weekly Update tab, if the workbook has one.

    Returns (rows, meta). A workbook without the tab is not an error: it
    means nothing changed this week, and the page simply is not built.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Weekly Update" not in wb.sheetnames:
        return [], {}
    ws = wb["Weekly Update"]

    meta, header_at = {}, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12,
                                         values_only=True), 1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells and cells[0] == "Updated":
            meta["updated"] = cells[1]
        if cells and cells[0] == "Method":
            meta["method"] = cells[1]
        if cells and cells[0] == "Position" and "Player" in cells:
            header_at = i
            head = [c.lower() for c in cells]
            break
    if header_at is None:
        return [], meta

    def col(*names):
        for n in names:
            if n in head:
                return head.index(n)
        return None

    ci = {"pos": col("position"), "player": col("player"), "team": col("team"),
          "decision": col("decision"), "rb": col("rank before"),
          "ra": col("rank after"), "pb": col("ppr before"),
          "pa": col("ppr after"), "d": col("ppr delta"),
          "ev": col("evidence"), "why": col("reason"), "src": col("source")}

    out = []
    for row in ws.iter_rows(min_row=header_at + 1, values_only=True):
        def g(k):
            i = ci.get(k)
            if i is None or i >= len(row) or row[i] is None:
                return None
            v = str(row[i]).strip()
            return v or None

        name = g("player")
        pos = g("pos")
        if not name or not pos or pos not in POSITIONS:
            continue
        out.append({
            "pos": pos, "name": name, "team": (g("team") or "").upper(),
            "decision": (g("decision") or "").upper(),
            "rank_before": num(g("rb")), "rank_after": num(g("ra")),
            "ppr_before": num(g("pb")), "ppr_after": num(g("pa")),
            "delta": num(g("d")), "evidence": g("ev") or "",
            "reason": g("why") or "", "source": g("src") or "",
            "slug": slug(name),
        })
    return out, meta


PAGE_CSS = """
.topbar .logo,.topbar .vbtn{text-decoration:none}
.topbar .vbtn:hover{text-decoration:none; color:var(--ink)}
.vbtn[aria-current="page"]{color:#0A0C08; background:var(--signal);
  border-color:var(--signal)}

/* ---- changes ----
   A record, not a highlight reel. Grouped by the reason rather than by
   player, because one piece of news usually moves several players and
   showing them apart hides the fact that they are the same decision. */
.chwrap{max-width:1080px; margin:0 auto; padding:0 1rem 4rem}
.chhead h1{font-size:1.7rem; margin:1.6rem 0 0; letter-spacing:-.01em;
  font-family:var(--text)}
.chsub{color:var(--quiet); font-size:.86rem; margin:.4rem 0 0; max-width:70ch;
  line-height:1.55}
.chdate{display:inline-block; margin-left:.4rem; font-family:var(--agate);
  text-transform:uppercase; letter-spacing:.06em; font-size:.7rem;
  color:var(--signal); border:1px solid var(--rule); border-radius:999px;
  padding:.1rem .5rem; vertical-align:.05em}

.chctl{display:flex; gap:.3rem; flex-wrap:wrap; align-items:center;
  margin:1.4rem 0 .3rem}
.chlab{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.07em; font-size:.66rem; color:var(--quiet);
  margin-right:.3rem}
.chtab{font-family:var(--agate); text-transform:uppercase;
  background:transparent; border:1px solid var(--rule); color:var(--quiet);
  font-size:.76rem; padding:.3rem .7rem; border-radius:999px;
  cursor:pointer; letter-spacing:.04em}
.chtab:hover{color:var(--ink); border-color:var(--ink)}
.chtab[aria-pressed="true"]{background:var(--signal);
  border-color:var(--signal); color:#0b0f0a; font-weight:600}

.chgroup{background:var(--card); border:1px solid var(--rule);
  border-radius:10px; padding:1rem 1.1rem; margin:.8rem 0 0}
.chwhy{font-size:.92rem; line-height:1.6; color:var(--ink); margin:0;
  max-width:76ch}
.chmeta{display:flex; gap:.5rem; flex-wrap:wrap; align-items:center;
  margin:.6rem 0 0}
.chev{font-family:var(--agate); text-transform:uppercase; font-size:.62rem;
  letter-spacing:.06em; border:1px solid var(--rule); border-radius:999px;
  padding:.1rem .5rem; color:var(--quiet)}
.chev.e-high{color:#8BE04E; border-color:#8BE04E}
.chev.e-medium{color:#B9DE7E; border-color:#B9DE7E}
.chev.e-uncertain{color:var(--standing); border-color:var(--standing)}
.chsrc{font-size:.74rem; color:var(--quiet); text-decoration:underline}
.chsrc:hover{color:var(--signal)}

.chmoves{display:grid; gap:.3rem; margin:.8rem 0 0}
.chmove{display:flex; gap:.6rem; align-items:baseline; flex-wrap:wrap;
  padding:.35rem 0; border-top:1px solid var(--rule);
  font-variant-numeric:tabular-nums}
.chpos{font-family:var(--data); font-size:.7rem; color:var(--quiet);
  width:2rem; flex:none}
.chname{font-size:.88rem; color:var(--ink); flex:1 1 9rem}
.chname a{color:var(--ink); text-decoration:none}
.chname a:hover{color:var(--signal)}
.chrank{font-family:var(--data); font-size:.76rem; color:var(--quiet);
  flex:none}
.chpts{font-family:var(--data); font-size:.82rem; font-weight:600;
  flex:none; min-width:5rem; text-align:right}
.up{color:#8BE04E} .down{color:#FF6B4A} .flat{color:var(--quiet)}
.chdec{font-family:var(--agate); text-transform:uppercase; font-size:.58rem;
  letter-spacing:.06em; color:var(--quiet); flex:none}

.chnote{color:var(--quiet); font-size:.78rem; margin:2rem 0 0;
  max-width:74ch; line-height:1.55}
.chnote h2{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.07em; font-size:.78rem; color:var(--quiet);
  margin:0 0 .5rem}
.chnote dt{font-family:var(--agate); text-transform:uppercase;
  letter-spacing:.05em; font-size:.66rem; color:var(--ink);
  margin-top:.6rem}
.chnote dd{margin:.1rem 0 0; font-size:.8rem; line-height:1.5}
.chempty{color:var(--quiet); font-size:.86rem; padding:1.2rem 0}

@media (max-width:760px){
  .chtab{min-height:44px; display:inline-flex; align-items:center}
  .chpts{min-width:4.2rem}
}
"""


def site_chrome():
    tpl = SITE / "template.html"
    if not tpl.exists():
        return "", "", ""
    src = tpl.read_text()
    css = re.search(r"<style>(.*?)</style>", src, re.S)
    foot = re.search(r"<footer.*?</footer>", src, re.S)
    header = (
        '<header class="topbar">\n'
        '  <div class="wrap tbrow">\n'
        '    <a class="logo" href="/">Lineup<em>Beat</em></a>\n'
        '    <nav class="views"><a class="vbtn" href="/">The Wire</a>'
        '<a class="vbtn" href="/#v=roster">My Roster</a>'
        f'<a class="vbtn" href="/{SPORT}/data/" aria-current="page">'
        'Fantasy Data</a>'
        + seo.teams_menu()
        + '<a class="vbtn" href="/about/">Who We Are</a></nav>\n'
        '  </div>\n'
        '</header>')
    return (css.group(1) if css else ""), header, (foot.group(0) if foot else "")


def group_changes(rows):
    """By reason, because one piece of news moves several players.

    Pearsall going out raised four other San Francisco receivers. Listing
    them separately would show four changes; listing them together shows
    one decision, which is what actually happened.
    """
    groups = {}
    for r in rows:
        key = (r["reason"] or r["name"])[:400]
        g = groups.setdefault(key, {"reason": r["reason"], "rows": [],
                                    "evidence": r["evidence"],
                                    "source": r["source"]})
        g["rows"].append(r)
        # The strongest evidence in the group describes the group.
        if EVIDENCE_ORDER.get(r["evidence"], 9) < \
                EVIDENCE_ORDER.get(g["evidence"], 9):
            g["evidence"] = r["evidence"]
    out = list(groups.values())
    for g in out:
        g["rows"].sort(key=lambda r: -(abs(r["delta"] or 0)))
        g["biggest"] = max((abs(r["delta"] or 0) for r in g["rows"]),
                           default=0)
    out.sort(key=lambda g: (EVIDENCE_ORDER.get(g["evidence"], 9),
                            -g["biggest"]))
    return out


def move_html(r, links):
    d = r["delta"]
    cls = "flat" if not d else ("up" if d > 0 else "down")
    pts = ("&mdash;" if d is None
           else f'{d:+.1f}'.replace("-", "\u2212"))
    rank = ""
    if r["rank_before"] and r["rank_after"]:
        rank = (f'{r["pos"]}{r["rank_before"]:.0f} '
                f'&rarr; {r["pos"]}{r["rank_after"]:.0f}')
    name = (f'<a href="/{SPORT}/{r["slug"]}/">{esc(r["name"])}</a>'
            if r["slug"] in links else esc(r["name"]))
    dec = ("" if r["decision"] == "UPDATED"
           else f'<span class="chdec">{esc(r["decision"].lower())}</span>')
    return (f'<div class="chmove">'
            f'<span class="chpos">{esc(r["pos"])}</span>'
            f'<span class="chname">{name}</span>'
            f'<span class="chrank">{rank}</span>'
            f'{dec}'
            f'<span class="chpts {cls}">{pts}</span>'
            f'</div>')


def build_html(rows, meta, links, css, header, footer, built):
    groups = group_changes(rows)
    moved = [r for r in rows if r["decision"] != "WATCH, NO CHANGE"]
    watch = [r for r in rows if r["decision"] == "WATCH, NO CHANGE"]
    updated = meta.get("updated") or built.strftime("%B %-d, %Y")

    blocks = []
    for g in groups:
        ev = g["evidence"] or ""
        evcls = "e-" + ev.lower().split("-")[0] if ev else ""
        src = (f'<a class="chsrc" href="{esc(g["source"])}" '
               f'rel="nofollow noopener" target="_blank">Source</a>'
               if g["source"].startswith("http") else "")
        posns = sorted({r["pos"] for r in g["rows"]})
        blocks.append(
            f'<article class="chgroup" data-pos="{esc(",".join(posns))}" '
            f'data-ev="{esc(ev)}">\n'
            f'  <p class="chwhy">{esc(g["reason"])}</p>\n'
            f'  <div class="chmeta">'
            f'{f_ev(ev, evcls)}{src}</div>\n'
            f'  <div class="chmoves">'
            f'{"".join(move_html(r, links) for r in g["rows"])}</div>\n'
            f'</article>')

    body = f"""<main class="chwrap">
  <nav class="crumbs" aria-label="Breadcrumb">
    <a href="/">LineupBeat</a><span>/</span>
    <a href="/{SPORT}/data/">Fantasy data</a><span>/</span>
    <a href="/{SPORT}/projections/">Projections</a><span>/</span>
    <b>Changes</b></nav>

  <div class="chhead">
    <h1>What changed in our projections</h1>
    <p class="chsub">Every projection we moved this week, what moved it, and
      what we looked at and left alone.
      <span class="chdate">{esc(updated)}</span></p>
  </div>

  <div class="chctl" role="group" aria-label="Position">
    <span class="chlab">Position</span>
    <button class="chtab" data-f="ALL" aria-pressed="true">All</button>
    {''.join(f'<button class="chtab" data-f="{p}" aria-pressed="false">{p}</button>'
             for p in POSITIONS)}
  </div>

  <p class="chsub" style="margin-top:.8rem">
    <b>{len(moved)}</b> projections moved across
    <b>{len(groups)}</b> decisions.
    {f"<b>{len(watch)}</b> looked at and left alone." if watch else ""}</p>

{chr(10).join(blocks)}

  <p class="chempty" id="chempty" hidden>Nothing changed at that
    position.</p>

  <section class="chnote">
    <h2>How to read this</h2>
    <p>Changes are grouped by what caused them, because one piece of news
       usually moves several players. A receiver ruled out for the season
       raises everybody else in that receiving corps, and those are one
       decision rather than five.</p>
    <dl>
      <div><dt>Updated</dt>
        <dd>The projection moved on this evidence.</dd></div>
      <div><dt>Reconciled</dt>
        <dd>Moved to stay consistent with another change. A backfield only
            has so many carries, so raising one back lowers another.</dd></div>
      <div><dt>Out, zeroed</dt>
        <dd>Ruled out for the season. The projection is set to zero and his
            opportunity is redistributed.</dd></div>
      <div><dt>Watch, no change</dt>
        <dd>We looked at it and moved nothing. Uncertain injuries and
            possible discipline are monitored rather than automatically
            deducted.</dd></div>
    </dl>
    <p style="margin-top:.9rem">Evidence is our own read on how firm the
       information is, not a measure of how much a projection moved. A
       confirmed starting change is high; a camp report that a player looks
       good is not.</p>
  </section>
{seo.faq_html(CHANGES_FAQ)}{seo.related_html('projections')}
</main>

<script>
let posFilter = "ALL";
const groups = [...document.querySelectorAll(".chgroup")];
const empty = document.getElementById("chempty");

function draw(){{
  let shown = 0;
  groups.forEach(g => {{
    const ok = posFilter === "ALL" ||
      (g.dataset.pos || "").split(",").includes(posFilter);
    g.hidden = !ok;
    if(ok) shown++;
    // Inside a shown group, dim the moves that are not this position:
    // a Pearsall decision that moved four receivers and one back should
    // still explain itself when you filter to backs.
    g.querySelectorAll(".chmove").forEach(m => {{
      const p = m.querySelector(".chpos").textContent.trim();
      m.style.opacity = (posFilter === "ALL" || p === posFilter) ? "" : ".38";
    }});
  }});
  empty.hidden = shown > 0;
}}

document.querySelectorAll("[data-f]").forEach(b =>
  b.addEventListener("click", () => {{
    posFilter = b.dataset.f;
    document.querySelectorAll("[data-f]").forEach(x =>
      x.setAttribute("aria-pressed", x === b ? "true" : "false"));
    draw();
  }}));
draw();
</script>"""
    return body


def f_ev(ev, cls):
    return (f'<span class="chev {cls}">{esc(ev)} evidence</span>'
            if ev else "")


CHANGES_FAQ = [
    ("How often do the projections change?",
     "Whenever the information behind them does. In the preseason that is "
     "usually a few times a week, driven by starting decisions, confirmed "
     "injuries and depth chart moves. This page records each one with the "
     "reason and a source."),
    ("Why did a player's projection change when nothing happened to him?",
     "Because something happened to a team-mate. A backfield has a fixed "
     "number of carries and a passing game a fixed number of targets, so "
     "raising one player lowers another. Those changes are marked "
     "Reconciled and grouped with the decision that caused them."),
    ("Does a camp report automatically change a projection?",
     "No. A confirmed starting change or a season-ending injury moves a "
     "projection. A report that somebody looks good in camp is weaker "
     "evidence and usually moves a projection less, or not at all. The "
     "evidence label on each change says which."),
    ("What does Watch, no change mean?",
     "That we looked at something and did not move anything. Uncertain "
     "injuries and possible discipline are monitored rather than "
     "automatically deducted, because guessing at a suspension length is "
     "not analysis."),
]


def add_to_sitemap(url):
    sm = SITE / "sitemap.xml"
    if not sm.exists():
        return False
    text = sm.read_text()
    if url in text:
        return False
    today = eastern_now().strftime("%Y-%m-%d")
    sm.write_text(text.replace(
        "</urlset>",
        f"  <url><loc>{url}</loc><lastmod>{today}</lastmod>"
        f"<changefreq>weekly</changefreq><priority>0.7</priority></url>\n"
        "</urlset>"))
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--projections", default="data/projections.xlsx")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    wb = ROOT / args.projections
    if not wb.exists():
        sys.exit(f"  no {args.projections}")
    rows, meta = read_changes(wb)
    if not rows:
        print(f"\n  no Weekly Update tab in {args.projections}, "
              f"so no changes page")
        return 0

    # Link only where a player page exists, the same rule the board uses.
    links = {p.name for p in (SITE / SPORT).glob("*") if p.is_dir()}

    built = eastern_now()
    css, header, footer = site_chrome()
    body = build_html(rows, meta, links, css, header, footer, built)

    moved = [r for r in rows if r["decision"] != "WATCH, NO CHANGE"]
    updated = meta.get("updated") or built.strftime("%B %-d, %Y")
    title = "What Changed in Our Fantasy Projections | LineupBeat"
    desc = (f"Every 2026 fantasy football projection we changed and why, "
            f"with the report behind each one. {len(moved)} projections "
            f"updated as of {updated}.")

    schema = {
        "@type": "Dataset",
        "name": "LineupBeat fantasy projection changes",
        "description": desc,
        "url": f"{seo.SITE_URL}/{SPORT}/projections/changes/",
        "dateModified": built.strftime("%Y-%m-%d"),
        "creator": {"@type": "Organization", "name": "LineupBeat"},
        **seo.dataset_extras(temporal="2026"),
    }
    crumbs = seo.breadcrumbs([
        ("LineupBeat", "/"), ("Fantasy data", f"/{SPORT}/data/"),
        ("Projections", f"/{SPORT}/projections/"),
        ("Changes", f"/{SPORT}/projections/changes/")])

    page = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical"
      href="{seo.SITE_URL}/{SPORT}/projections/changes/">
<meta property="og:title" content="{esc(title)}">
<meta property="og:description" content="{esc(desc)}">
<meta property="og:url"
      content="{seo.SITE_URL}/{SPORT}/projections/changes/">
<meta property="og:type" content="website">
<script type="application/ld+json">{seo.graph(
    schema, crumbs, seo.faq_schema(CHANGES_FAQ), seo.ORGANISATION)}</script>
<style>{css}{PAGE_CSS}{seo.RELATED_CSS}{seo.TEAMS_CSS}{seo.BYLINE_CSS}</style>
</head>
<body>
{header}
{body}
{footer}
{seo.TEAMS_JS}{seo.TRACKING}
</body>
</html>"""

    out = (Path(args.out) if args.out
           else SITE / SPORT / "projections" / "changes" / "index.html")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(seo.check_page(page, str(out)))

    import collections
    dec = collections.Counter(r["decision"] for r in rows)
    print(f"\n  {len(rows)} changes, {len(group_changes(rows))} decisions")
    print("  " + "  ".join(f"{n} {d.lower()}" for d, n in dec.most_common()))
    print(f"\n  wrote {out.relative_to(ROOT)}  ({len(page):,} bytes)")
    if add_to_sitemap(f"{seo.SITE_URL}/{SPORT}/projections/changes/"):
        print(f"  added to sitemap.xml")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
