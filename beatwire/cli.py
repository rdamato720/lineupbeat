"""Command line interface.

  python -m beatwire.cli run    --sport nfl --offline --stub
  python -m beatwire.cli run    --sport nfl
  python -m beatwire.cli feed   --sport nfl --min 2
  python -m beatwire.cli feed   --sport nfl --roster nfl-0007,nfl-0019 --html out.html
  python -m beatwire.cli verify --sport nfl
  python -m beatwire.cli resolve --sport nfl --name "Allen" --team NYJ
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

from . import render
from .models import in_draft_season
from .pipeline import run as run_pipeline
from .registry import Registry
from .resolve import Resolver
from .store import Store

ROOT = Path(__file__).resolve().parent.parent


def _pslug(s):
    """Name to slug, suffixes removed.

    The projection sheet and the roster disagree about suffixes -- "Luther
    Burden III" against "Luther Burden" -- so both sides are normalised the
    same way rather than hoping they match.
    """
    import re as _re
    s = _re.sub(r"[^\w\s-]", "", (s or "").lower())
    s = _re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s)
    return _re.sub(r"[\s_]+", "-", s).strip("-")


def _client(stub: bool):
    if stub:
        return None
    # A local model needs no client. The check was running before extraction
    # got far enough to notice BEATWIRE_LOCAL was set, so the whole pipeline
    # refused to start on a machine that was never going to call the API.
    from . import local_model
    if local_model.enabled():
        return None
    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("Set ANTHROPIC_API_KEY, set BEATWIRE_LOCAL to point at an "
                 "ollama host, or pass --stub.")
    import anthropic
    return anthropic.Anthropic()


def cmd_run(args):
    store = Store(args.db)
    client = _client(args.stub)
    report = run_pipeline(
        args.sport, store, client=client, stub=args.stub, offline=args.offline,
        x_daily_cap=args.x_daily_cap,
        tapi_daily_cap=args.tapi_daily_cap, only=args.only,
    )
    print(report)
    print(f"  totals: {store.stats()}")


def cmd_feed(args):
    store = Store(args.db)
    ids = args.roster.split(",") if args.roster else None
    rows = store.feed(
        sport=args.sport,
        player_ids=ids,
        min_actionability=args.min,
        limit=args.limit,
    )
    if args.json:
        print(render.to_json(rows))
    elif args.html:
        Path(args.html).write_text(render.to_html(rows, f"{args.sport.upper()} beat feed"))
        print(f"wrote {args.html} ({len(rows)} nuggets)")
    else:
        print(render.to_terminal(rows))


def _check_page(html, where):
    """Refuse to write a document a browser cannot parse."""
    bad = []
    if html.count("<style") != html.count("</style>"):
        bad.append(f"{html.count('<style')} <style> vs "
                   f"{html.count('</style>')} </style>")
    if "<body" not in html:
        bad.append("no <body>")
    if "</html>" not in html:
        bad.append("no </html>")
    if html.count("<script") != html.count("</script>"):
        bad.append(f"{html.count('<script')} <script> vs "
                   f"{html.count('</script>')} </script>")
    if bad:
        raise SystemExit(f"\n  REFUSING TO WRITE {where}:\n"
                         + "".join(f"    {x}\n" for x in bad)
                         + "  Nothing written.\n")


def cmd_export(args):
    """Write the data bundle the static site reads.

    One file, all sports, plus a player index so the roster picker can work
    without a backend. This is what the GitHub Action publishes.
    """
    import json

    store = Store(args.db)
    sports = args.sports.split(",")
    bundle = {"generated_at": None, "sports": {}, "players": []}

    # ADP is a draft-season number. Outside the window it is zeroed here, so
    # the site never renders a stale badge and never has to know the date.
    # `--adp on|off` overrides for testing or an unusual year.
    draft_season = (in_draft_season() if getattr(args, "adp", "auto") == "auto"
                    else args.adp == "on")

    # Projections, from the same workbook the projections page is built
    # from.
    #
    # This read published_snapshot -> run_projections, which is the engine's
    # output and exists only where the engine has been run. CI has neither
    # table, so the wire's player cards carried no projection in production
    # while the standalone board showed all 614 -- two sources for one
    # number, and only one of them deployed.
    #
    # The workbook is committed, so both now read it and cannot disagree.
    proj = {}
    _wb = ROOT / "data" / "projections.xlsx"
    try:
        if _wb.exists():
            import openpyxl, re as _re
            _slug = _pslug

            # Keyed by name slug, because the registry is built per sport
            # further down and this runs before it. The roster id is
            # attached at the point the players are written out.
            book = openpyxl.load_workbook(_wb, data_only=True)
            for sheet in book.sheetnames:
                pos = sheet.split()[0].upper()
                if pos not in ("QB", "RB", "WR", "TE"):
                    continue
                ws = book[sheet]
                head = [str(c.value or "").strip().lower() for c in ws[1]]
                def col(*names):
                    for n in names:
                        if n in head:
                            return head.index(n)
                    return None
                ci = {"player": col("player", "name"),
                      "rank": col("rank"),
                      "ppr": col("ppr", "ppr points"),
                      "half": col("half ppr", "half"),
                      "std": col("non-ppr", "standard", "std"),
                      "rec": col("receptions", "rec"),
                      "recyd": col("rec yds", "receiving yards"),
                      "ruyd": col("rush yds", "rushing yards")}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if ci["player"] is None or ci["player"] >= len(row):
                        continue
                    nm = row[ci["player"]]
                    if not nm:
                        continue
                    key = _slug(str(nm))
                    if key in proj:
                        continue
                    def val(k, cast=float):
                        i = ci.get(k)
                        if i is None or i >= len(row) or row[i] is None:
                            return None
                        try:
                            return cast(row[i])
                        except (TypeError, ValueError):
                            return None
                    ppr = val("ppr")
                    if ppr is None:
                        continue
                    proj[key] = {
                        "ppr": round(ppr, 1),
                        "half": round(val("half") or ppr, 1),
                        "std": round(val("std") or ppr, 1),
                        "healthy": round(ppr, 1),
                        "adjusted": None, "games": None,
                        "floor": None, "ceil": None,
                        "rank": val("rank", int),
                        "rec": round(val("rec") or 0, 1),
                        "recyd": round(val("recyd") or 0),
                        "ruyd": round(val("ruyd") or 0),
                        "adj": None, "trace": [], "season": 2026,
                    }
    except Exception as exc:
        print(f"  projections unavailable: {exc}")

    if proj:
        print(f"  {len(proj)} projections attached")
    if not draft_season:
        print("  outside draft season, ADP suppressed (--adp on to force)")

    for sport in sports:
        reg = Registry(sport)
        bundle["sports"][sport] = {
            "display": reg.profile.display,
            "nuggets": store.feed(sport=sport, limit=args.limit),
        }
        bundle["players"].extend(
            {"id": p.id, "sport": sport, "name": p.name,
             "team": p.team, "pos": p.position, "espn": p.espn_id,
             "rank": p.rank, "depth": p.depth_pos, "order": p.depth_order,
             "status": p.injury_status, "exp": p.years_exp,
             # Zeroed outside draft season so the site never has to decide.
             "adp": p.adp if draft_season else 0.0,
             # Matched on the name, since the workbook has no roster id.
             # Suffixes differ between the sheet and the roster -- the board
             # says Luther Burden III and the roster says Luther Burden --
             # so the slug is stripped of them on both sides.
             **({"proj": proj[_pslug(p.name)]}
                if _pslug(p.name) in proj else {})}
            for p in reg.players
        )

    from datetime import datetime, timezone
    bundle["generated_at"] = datetime.now(timezone.utc).isoformat()

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(bundle, indent=2))
    total = sum(len(s["nuggets"]) for s in bundle["sports"].values())
    print(f"wrote {out} ({total} nuggets, {len(bundle['players'])} players)")

    # Moving Now, written into the HTML.
    #
    # The panel is the clearest demonstration of what the site does -- a
    # player, what changed, who reported it -- and a crawler saw an empty
    # div. If something is worth putting on the homepage it is worth being
    # in the source; JavaScript re-renders it on load either way.
    def _moving_now_html(bundle, sport="nfl", limit=8):
        import html as _h
        from datetime import datetime as _dt, timezone as _tz
        s = (bundle.get("sports") or {}).get(sport) or {}
        nug = s.get("nuggets") or []
        by = {}
        for n in nug:
            pid = n.get("player_id")
            if not pid or not n.get("resolved"):
                continue
            cur = by.setdefault(pid, {"n": 0, "top": None,
                                      "name": n.get("player_name"),
                                      "team": n.get("team")})
            cur["n"] += 1
            # Newest item is the reason line, matching what the client does.
            if (not cur["top"] or (n.get("published_at") or "")
                    > (cur["top"].get("published_at") or "")):
                cur["top"] = n
        # Two or more reports today: a developing story rather than the
        # newest thing, which is what the wire beneath it already answers.
        rows = [v for v in by.values() if v["n"] >= 2 and v["top"]]
        rows.sort(key=lambda x: x["top"].get("published_at") or "",
                  reverse=True)
        out = []
        for r in rows[:limit]:
            top = r["top"]
            attrs = top.get("attributions") or []
            who = attrs[0].get("source_name", "") if attrs else ""
            when = top.get("published_at") or ""
            out.append(
                '<button class="trend">'
                '<span class="trow">'
                f'<span class="tname">{_h.escape(r["name"] or "")}</span>'
                f'<span class="twhy">{_h.escape(top.get("claim") or "")}</span>'
                '<span class="tmeta">'
                f'{_h.escape(r["team"] or "")} &middot; '
                f'{_h.escape(who)} &middot; '
                f'{r["n"]} reports'
                f'<time datetime="{_h.escape(when)}"></time>'
                '</span></span></button>')
        return "\n".join(out) or ""

    # Render the static site by inlining the bundle into the template. One
    # self-contained file: no server, no CORS, no loading state to design for.
    tpl = Path(args.template)
    if tpl.exists():
        site = Path(args.site)
        site.parent.mkdir(parents=True, exist_ok=True)
        payload = json.dumps(bundle).replace("</", "<\\/")  # keep it out of </script>
        html_out = tpl.read_text().replace(
            '/*__DATA__*/ {"generated_at":new Date().toISOString(),'
            '"sports":{},"players":[]}',
            payload,
        ).replace("<!--__MOVINGNOW__-->", _moving_now_html(bundle))
        # The homepage that shipped empty had a 200, 1.2MB and a valid
        # sitemap: the only signal was looking at it. An unclosed style
        # tag swallows the entire body, so these assertions are cheap and
        # would all have fired.
        _check_page(html_out, str(site))
        site.write_text(html_out)
        print(f"wrote {site}")


def cmd_spend(args):
    """What the metered sources cost, and what they returned for it.

    The column that matters is $/nugget. A writer who costs $3 a month and
    produces two usable notes is not worth polling; one who costs $3 and
    produces sixty is. This is how you allocate a budget instead of guessing.
    """
    store = Store(args.db)
    daily = store.spend_report(args.provider, days=args.days)
    if not daily:
        print(f"  no recorded {args.provider} spend")
        return

    total = sum(d["cost"] for d in daily)
    units = sum(d["units"] for d in daily)
    print(f"  last {len(daily)} days: {units} units, ${total:.2f}")
    print(f"  run rate: ${total / max(len(daily), 1) * 30:.2f}/month\n")
    for d in daily:
        print(f"    {d['day']}  {d['units']:>6} units  ${d['cost']:>7.2f}")

    rows = store.spend_by_source(args.provider)
    if rows:
        print(f"\n  {'source':<30} {'units':>7} {'cost':>8} {'nuggets':>8} {'$/nugget':>9}")
        for r in rows[:args.limit]:
            per = f"${r['cost']/r['nuggets']:.3f}" if r["nuggets"] else "-"
            print(f"  {r['source_id']:<30} {r['units']:>7} "
                  f"${r['cost']:>7.2f} {r['nuggets']:>8} {per:>9}")


def cmd_preflight(args):
    """One command, GO or NO-GO. Run it before every deploy.

    Each check is something that fails silently in production if you skip it.
    None of them are style opinions.
    """
    import subprocess
    from datetime import datetime, timedelta, timezone

    ok, warn, fail = [], [], []
    root = Path(__file__).resolve().parent.parent

    # 1. roster looks real, not the checked-in sample
    reg = Registry(args.sport)
    n_players, n_teams = len(reg.players), len({p.team for p in reg.players})
    floor = {"nfl": 1500, "mlb": 900}.get(args.sport, 200)
    if n_players < floor:
        fail.append(f"roster has {n_players} players across {n_teams} teams, "
                    f"expected {floor}+. Run scripts/import_rosters.py")
    else:
        ok.append(f"roster: {n_players} players, {n_teams} teams")

    # 2. aliases actually populated, or every accented name will miss
    with_alias = sum(1 for p in reg.players if p.aliases)
    if n_players and with_alias / n_players < 0.02:
        warn.append(f"only {with_alias} players have aliases. The importer "
                    f"generates them; a near-zero count means it did not run")
    else:
        ok.append(f"aliases: {with_alias} players")

    # 3. team codes line up
    reg_teams = {t for s in reg.sources for t in s.teams}
    roster_teams = {p.team for p in reg.players}
    drift = (reg_teams - roster_teams) | (roster_teams - reg_teams)
    if drift:
        fail.append(f"team code drift on {sorted(drift)[:6]}. Run doctor")
    else:
        ok.append(f"team codes: {len(reg_teams)} teams aligned")

    # 4. no enabled source still pointing at a TODO
    todo_live = [s.id for s in reg.enabled_sources
                 if "TODO" in s.url or not (s.url or s.handle)]
    if todo_live:
        fail.append(f"{len(todo_live)} enabled sources have TODO urls")
    unfilled = [s for s in reg.sources if "TODO" in s.url]
    if unfilled:
        warn.append(f"{len(unfilled)} hand-research slots unfilled "
                    f"({len(reg.enabled_sources)} sources live)")
    else:
        ok.append(f"sources: {len(reg.enabled_sources)} live, none pending")

    from collections import Counter as _C
    mix = _C(s.outlet or s.kind for s in reg.enabled_sources)
    if mix:
        top, n = mix.most_common(1)[0]
        share = n / len(reg.enabled_sources)
        if share > 0.7:
            warn.append(f"{share:.0%} of live sources are '{top}'. One outlet "
                        f"changing its mind takes out most of your coverage")
        else:
            ok.append(f"source mix: {dict(mix)}")

    # 5. resolver regression suite
    r = subprocess.run([sys.executable, str(root / "scripts" / "test_resolve.py")],
                       capture_output=True, text=True)
    if r.returncode:
        fail.append("resolver tests failing. Run scripts/test_resolve.py")
    else:
        ok.append("resolver tests pass")

    # 6. extraction is really running, not stubbed
    if not os.environ.get("ANTHROPIC_API_KEY"):
        fail.append("ANTHROPIC_API_KEY not set, so you are still on --stub")
    else:
        ok.append("extraction key present")

    # 7. data is fresh and resolving
    store = Store(args.db)
    st = store.stats()
    if not st["nuggets"]:
        fail.append("no nuggets in the database")
    else:
        # Measured over the last week, not all time.
        #
        # The lifetime rate is a ratchet: unresolvable mentions accumulate
        # and never leave, so it only climbs. It crossed eight percent and
        # then blocked every run, and the wire stopped for a reason that was
        # not a fault.
        #
        # And an unresolved nugget is not a broken one. It publishes,
        # unlinked, saying honestly that we do not know who this is. That is
        # worth watching, and it is not worth stopping the news for -- so it
        # warns loudly and only fails when the resolver is clearly broken
        # rather than merely refusing things it should refuse.
        base = st.get("recent_nuggets") or 0
        if base < 50:
            ok.append("too few recent nuggets to judge resolution")
        else:
            rate = st["recent_unresolved"] / base * 100
            if rate > 25:
                fail.append(f"{rate:.1f}% of the last week unresolved, which "
                            f"means the resolver or the roster is broken. Run "
                            f"`unresolved --sport {args.sport}`")
            elif rate > 10:
                warn.append(f"{rate:.1f}% of the last week unresolved. Run "
                            f"`unresolved --sport {args.sport}` and add "
                            f"aliases; the wire still publishes them unlinked")
            else:
                ok.append(f"unresolved rate {rate:.1f}% over the last week")

        x_today = store.spend_today("x")
        if x_today:
            ok.append(f"x spend today ${x_today:.2f}")

        recent = store.feed(sport=args.sport, limit=1)
        if recent:
            age = datetime.now(timezone.utc) - datetime.fromisoformat(
                recent[0]["published_at"]
            ).astimezone(timezone.utc)
            if age > timedelta(hours=36):
                warn.append(f"newest nugget is {age.days}d old, feeds may be dead")
            else:
                ok.append(f"freshest nugget {int(age.total_seconds()//3600)}h old")

    for m in ok:
        print(f"  [ ok ] {m}")
    for m in warn:
        print(f"  [warn] {m}")
    for m in fail:
        print(f"  [FAIL] {m}")

    print()
    if fail:
        print(f"  NO-GO: {len(fail)} blocking issue(s)")
        sys.exit(1)
    print("  GO" + (f" ({len(warn)} warnings)" if warn else ""))


def cmd_unresolved(args):
    """Show the mentions the resolver could not place.

    This is a work queue, not an error log. The top entries are almost always
    either a missing alias or a player signed since your last roster import.
    Watch the rate, not just the list: a spike means your roster went stale.
    """
    store = Store(args.db)
    rows = store.unresolved_mentions(args.sport, args.limit)
    if not rows:
        print("  nothing unresolved")
        return
    print(f"  {'mention':<28} {'team':<6} count")
    for mention, team, n in rows:
        print(f"  {mention:<28} {team or '-':<6} {n}")
    print(f"\n  Fix by adding an alias to rosters/{args.sport}.csv or "
          f"re-running the roster import.")


def cmd_doctor(args):
    """Cross-check the registry against the roster.

    Exists because a team code mismatch is the worst class of bug in this
    system: entirely silent. If the registry says ATH and the roster says OAK,
    that team's sources carry no usable team hint, every bare surname in them
    goes unresolved, and the feed just looks a bit quiet all season. Nothing
    errors. Run this after any roster import or registry regeneration.
    """
    import sys as _sys

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
    try:
        from build_registry import ALIASES
    except ImportError:
        ALIASES = {}

    reg = Registry(args.sport)
    reg_teams = {t for s in reg.sources for t in s.teams}
    roster_teams = {p.team for p in reg.players}
    aliases = ALIASES.get(args.sport, {})

    orphan_sources = sorted(reg_teams - roster_teams)
    orphan_roster = sorted(roster_teams - reg_teams)
    matched = reg_teams & roster_teams

    print(f"  registry teams: {len(reg_teams)}   roster teams: {len(roster_teams)}"
          f"   matched: {len(matched)}")

    problems = 0
    for code in orphan_sources:
        hint = ""
        for old_code, new_code in aliases.items():
            if new_code == code and old_code in roster_teams:
                hint = f"  <- roster uses '{old_code}' for this team"
            elif old_code == code and new_code in roster_teams:
                hint = f"  <- roster uses '{new_code}' for this team"
        print(f"  [MISMATCH] registry '{code}' has no players in the roster{hint}")
        problems += 1
    for code in orphan_roster:
        print(f"  [NO SOURCE] roster '{code}' has no sources in the registry")
        problems += 1

    # A teamless source is a warning, not an error. National insiders
    # (Schefter, Rapoport, Pelissero) legitimately have no team, and they
    # write full names rather than bare surnames, so the loss of team scoping
    # costs little. Worth surfacing, not worth failing a build over.
    empty = [s.id for s in reg.enabled_sources if not s.teams]
    for sid in empty:
        print(f"  [no team] '{sid}' has no team hint, so bare surnames from "
              f"it will not resolve. Fine for national insiders.")

    todo = [s for s in reg.sources if "TODO" in s.url]
    print(f"\n  {len(todo)} hand-research slots still unfilled")
    if problems:
        print(f"  {problems} problems. Fix these before running the pipeline.")
        _sys.exit(1)
    print("  no team code problems")


def cmd_verify(args):
    """Check every feed URL in a registry actually returns entries.

    Run this before you trust any source registry, including the one shipped
    with this repo. Feeds rot constantly and a silently dead source looks
    exactly like a quiet news day.

    --fix rewrites the yaml to disable whatever did not respond, so a
    verification pass is one command instead of hand-editing a 200 line file.
    """
    import concurrent.futures as cf

    import feedparser
    import yaml as _yaml

    reg = Registry(args.sport)
    targets = [s for s in reg.sources
               if "TODO" not in s.url and (s.url or s.handle)]
    skipped = len(reg.sources) - len(targets)

    def check(s):
        try:
            if s.kind == "x":
                # Deliberately not verified here: every check costs money.
                return s, -1, "metered, not checked"
            if s.kind == "bluesky":
                from .ingest import fetch_bluesky
                return s, len(fetch_bluesky(s, limit=5)), ""
            feed = feedparser.parse(s.url)
            return s, len(feed.entries), ""
        except Exception as exc:
            return s, 0, str(exc)[:60]

    results = []
    with cf.ThreadPoolExecutor(max_workers=args.workers) as pool:
        for s, n, err in pool.map(check, targets):
            results.append((s, n, err))
            status = "ok  " if n > 0 else ("skip" if n < 0 else "DEAD")
            where = s.url or f"@{s.handle}"
            print(f"  [{status}] {s.id:<26} {n:>3} entries  {where} {err}")

    dead = [s.id for s, n, _ in results if n == 0]
    live = sum(1 for _, n, _ in results if n > 0)
    print(f"\n  {live}/{len(results)} returning entries "
          f"({skipped} TODO slots skipped)")

    if dead and args.fix:
        path = Path(__file__).resolve().parent.parent / "sources" / f"{args.sport}.yaml"
        doc = _yaml.safe_load(path.read_text())
        for src in doc["sources"]:
            if src["id"] in dead:
                src["enabled"] = False
        head = "\n".join(
            l for l in path.read_text().splitlines() if l.startswith("#")
        )
        path.write_text(
            head + "\n\n"
            + _yaml.dump(doc, sort_keys=False, allow_unicode=True, width=100)
        )
        print(f"  disabled {len(dead)} dead sources in {path}")
    elif dead:
        print(f"  rerun with --fix to disable them")


def cmd_resolve(args):
    """Interactive check on the entity resolver, which is where trust dies."""
    reg = Registry(args.sport)
    r = Resolver(reg.players, reg.profile.position_groups)
    player, conf = r.resolve(args.name, args.team, args.pos)
    if player:
        print(f"  '{args.name}' + team={args.team} -> {player.name} "
              f"({player.team} {player.position}) conf={conf:.2f}")
    else:
        print(f"  '{args.name}' + team={args.team} -> UNRESOLVED (best={conf:.2f})")


def main():
    p = argparse.ArgumentParser(prog="beatwire")
    p.add_argument("--db", default="beatwire.db")
    sub = p.add_subparsers(dest="cmd", required=True)

    r = sub.add_parser("run")
    r.add_argument("--sport", required=True)
    r.add_argument("--stub", action="store_true", help="skip the model, use keyword extractor")
    r.add_argument("--offline", action="store_true", help="read fixtures/ instead of the network")
    r.add_argument("--tapi-daily-cap", type=float, default=12.0,
                   help="ceiling on twitterapi spend for the day. 174 handles "
                        "across twenty runs costs about ten dollars; two was "
                        "sized for half as many and ran out before ten in "
                        "the morning.")
    r.add_argument("--x-daily-cap", type=float, default=5.0,
                   help="hard local ceiling on X spend per day")
    r.add_argument("--only", help="poll only sources whose id or handle "
                                  "contains this, e.g. --only profootballdoc")
    r.set_defaults(func=cmd_run)

    f = sub.add_parser("feed")
    f.add_argument("--sport")
    f.add_argument("--roster", help="comma separated player ids")
    f.add_argument("--min", type=int, default=0, help="minimum actionability 0-3")
    f.add_argument("--limit", type=int, default=100)
    f.add_argument("--json", action="store_true")
    f.add_argument("--html")
    f.set_defaults(func=cmd_feed)

    e = sub.add_parser("export")
    e.add_argument("--sports", default="nfl", help="comma separated")
    e.add_argument("--out", default="site/data/feed.json")
    e.add_argument("--limit", type=int, default=500)
    e.add_argument("--adp", choices=["auto", "on", "off"], default="auto",
                   help="ADP is shown only in draft season; force it either way")
    e.add_argument("--template", default="site/template.html")
    e.add_argument("--site", default="site/index.html")
    e.set_defaults(func=cmd_export)

    sp = sub.add_parser("spend")
    sp.add_argument("--provider", default="x")
    sp.add_argument("--days", type=int, default=14)
    sp.add_argument("--limit", type=int, default=25)
    sp.set_defaults(func=cmd_spend)

    pf = sub.add_parser("preflight")
    pf.add_argument("--sport", required=True)
    pf.set_defaults(func=cmd_preflight)

    u = sub.add_parser("unresolved")
    u.add_argument("--sport", required=True)
    u.add_argument("--limit", type=int, default=40)
    u.set_defaults(func=cmd_unresolved)

    d = sub.add_parser("doctor")
    d.add_argument("--sport", required=True)
    d.set_defaults(func=cmd_doctor)

    v = sub.add_parser("verify")
    v.add_argument("--sport", required=True)
    v.add_argument("--fix", action="store_true",
                   help="disable sources that return nothing")
    v.add_argument("--workers", type=int, default=12)
    v.set_defaults(func=cmd_verify)

    s = sub.add_parser("resolve")
    s.add_argument("--sport", required=True)
    s.add_argument("--name", required=True)
    s.add_argument("--team")
    s.add_argument("--pos", help="position hint, e.g. P")
    s.set_defaults(func=cmd_resolve)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
